"""Admin CRUD routes — Courses, Papers, Lecturers, Students, Enrollment, Audit."""

import re
import random
import secrets
import os
import time
import csv
from io import BytesIO, StringIO
import openpyxl
from datetime import datetime, timedelta, timezone
import cv2
import numpy as np
from threading import Lock, Thread
from uuid import uuid4

from flask import Blueprint, request, jsonify, current_app, send_file, has_app_context
from bson import ObjectId
from pymongo.errors import DuplicateKeyError

try:
    import redis
except Exception:  # pragma: no cover - optional dependency at runtime
    redis = None

from app.extensions import get_collection
from app.models.attendance import log_attendance, get_approved_leave_dates, session_date_str
from app.models.audit import log_action, get_audit_logs, get_audit_log_by_id, mark_audit_log_rolled_back
from app.models.course import (
    create_course,
    get_all_courses,
    get_course_by_id,
    get_course_by_code,
    update_course,
    delete_course,
    is_course_active,
)
from app.models.enrollment import (
    create_student_profile,
    get_profile_by_user,
    get_profile_by_id,
    add_face_embedding,
    set_face_embeddings,
    enroll_in_papers,
    get_all_profiles,
    update_profile,
    delete_profile,
)
from app.models.paper import (
    create_paper,
    get_all_papers,
    get_paper_by_id,
    get_paper_by_code,
    get_papers_by_course,
    update_paper,
    delete_paper,
    bulk_assign_lecturer,
    bulk_assign_course,
)
from app.models.user import (
    create_user,
    get_users_by_role,
    get_users_by_ids,
    update_user,
    delete_user,
    find_user_by_id,
    find_user_by_email,
    reset_user_password,
    set_user_pin,
)
from app.services.face_detection import get_detector
from app.services.face_recognition import generate_embedding, normalize_embedding
from app.services.capture_upload import capture_faces_for_user, save_student_upload, save_cropped_face_dataset
from utilities.train_model import train_and_save_face_model
from app.utils.auth_decorators import role_required, super_admin_required, validate_ids
from app.security.rbac import (
    dept_scope_filter,
    is_super_admin,
    is_any_admin,
    validate_department_access,
    validate_role_assignment,
    get_user_department_id,
    ADMIN_ROLES,
)
from app.models.department import (
    create_department,
    get_all_departments,
    get_department_by_id,
    get_department_by_code,
    update_department,
    delete_department as soft_delete_department,
    hard_delete_department,
)
from app.utils.helpers import sanitise_mongo_doc, sanitise_many, decode_base64_image
from app.utils.timezone import india_timestamp_token
from app.utils.validation import validate_password_strength
from app.services.email_service import (
    send_welcome_email,
    send_password_reset_email,
    send_shortage_alert_email,
    is_email_delivery_enabled,
)

admin_bp = Blueprint("admin", __name__)

_QUERY_CACHE = {}
_QUERY_CACHE_LOCK = Lock()
_QUERY_CACHE_TTL_SECONDS = 30
_QUERY_CACHE_MAX_ENTRIES_DEFAULT = 500
_ELIGIBILITY_CACHE_TTL_SECONDS = 20
_QUEUE_CLIENT = None
_QUEUE_CLIENT_LOCK = Lock()
_QUEUE_UNAVAILABLE_LOGGED = False
_AUDIT_EXCLUDED_ACTIONS = [
    "paper_profile_read",
    "paper_profile_count",
    "student_profile_read",
    "student_profile_read_by_id",
    "student_profiles_bulk_read",
]


# ---------------------------------------------------------------------------
# Department scoping helpers
# ---------------------------------------------------------------------------

def _dept_filter(user):
    """Build a MongoDB query filter for department-level data isolation.

    • super_admin → {} (all data) unless they pass ?department_id=…
    • department_admin / lecturer → {department_id: <their dept>}
    """
    return dept_scope_filter(user)


def _user_dept_id(user):
    """Return the user's department_id as ObjectId (or None for super_admin)."""
    return get_user_department_id(user)

def _get_paginated_data(collection, filter_query, page=1, per_page=10, sort=None, project=None):
    """Refactored pagination helper to ensure consistency across all admin routes."""
    try:
        page = max(1, int(page or 1))
        per_page = max(1, int(per_page or 10))
    except (ValueError, TypeError):
        page = 1
        per_page = 10
        
    skip = (page - 1) * per_page
    
    cursor = collection.find(filter_query, project)
    if sort:
        cursor = cursor.sort(sort)
    
    total = collection.count_documents(filter_query)
    data = list(cursor.skip(skip).limit(per_page))
    
    return {
        "data": data,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page
    }


class _JobCancelledError(Exception):
    """Raised when a background job cancellation request is detected."""


def _utcnow():
    return datetime.now(timezone.utc)


def _temp_pass_display_enabled():
    return bool(current_app.config.get("TEMP_PASS_DISPLAY_ENABLED", False))


def _cache_get(key):
    now = time.monotonic()
    with _QUERY_CACHE_LOCK:
        item = _QUERY_CACHE.get(key)
        if not item:
            return None
        if item.get("expires_at", 0) <= now:
            _QUERY_CACHE.pop(key, None)
            return None
        return item.get("value")


def _cache_set(key, value, ttl_seconds):
    ttl = max(1, int(ttl_seconds or _QUERY_CACHE_TTL_SECONDS))
    max_entries = _QUERY_CACHE_MAX_ENTRIES_DEFAULT
    if has_app_context():
        try:
            max_entries = max(50, int(current_app.config.get("QUERY_CACHE_MAX_ENTRIES", _QUERY_CACHE_MAX_ENTRIES_DEFAULT)))
        except Exception:
            max_entries = _QUERY_CACHE_MAX_ENTRIES_DEFAULT

    with _QUERY_CACHE_LOCK:
        # Evict the oldest cache entries first to keep memory bounded.
        while len(_QUERY_CACHE) >= max_entries:
            oldest_key = min(
                _QUERY_CACHE,
                key=lambda cache_key: _QUERY_CACHE.get(cache_key, {}).get("expires_at", float("inf")),
            )
            _QUERY_CACHE.pop(oldest_key, None)

        _QUERY_CACHE[key] = {
            "value": value,
            "expires_at": time.monotonic() + ttl,
        }


def _cache_payload(key, ttl_seconds, builder):
    cached = _cache_get(key)
    if cached is not None:
        return cached
    value = builder()
    _cache_set(key, value, ttl_seconds)
    return value


def _clear_query_cache():
    with _QUERY_CACHE_LOCK:
        _QUERY_CACHE.clear()


def _update_training_job_progress(
    job_id,
    *,
    total_faces=None,
    processed_faces=None,
    trained_faces=None,
    failed_faces=None,
    stage=None,
    message=None,
):
    updates = {}
    if total_faces is not None:
        updates["training_total_faces"] = max(0, _to_int(total_faces, 0))
    if processed_faces is not None:
        updates["training_processed_faces"] = max(0, _to_int(processed_faces, 0))
    if trained_faces is not None:
        updates["training_trained_faces"] = max(0, _to_int(trained_faces, 0))
    if failed_faces is not None:
        updates["training_failed_faces"] = max(0, _to_int(failed_faces, 0))
    if stage is not None:
        updates["training_stage"] = _as_text(stage)
    if message is not None:
        updates["training_message"] = _as_text(message)

    current_job = _get_background_job(job_id) or {}
    total = updates.get("training_total_faces", _to_int(current_job.get("training_total_faces"), 0))
    processed = updates.get("training_processed_faces", _to_int(current_job.get("training_processed_faces"), 0))

    if total > 0:
        updates["training_progress_percent"] = round((processed / total) * 100, 2)
    else:
        updates["training_progress_percent"] = 0

    _update_background_job(job_id, **updates)


def _create_background_job(job_type, payload=None):
    job_id = str(uuid4())
    now = _utcnow()
    max_attempts = max(1, _to_int(current_app.config.get("TASK_QUEUE_MAX_RETRIES", 3), 3))
    get_collection("attendance", "background_jobs").insert_one(
        {
            "job_id": job_id,
            "job_type": job_type,
            "status": "queued",
            "created_at": now,
            "updated_at": now,
            "next_attempt_at": now,
            "attempts": 0,
            "max_attempts": max_attempts,
            "payload": payload or {},
            "retry_count": 0,
            "retry_in_seconds": None,
            "last_error_at": None,
            "dead_lettered_at": None,
            "error_history": [],
            "training_total_faces": 0,
            "training_processed_faces": 0,
            "training_trained_faces": 0,
            "training_failed_faces": 0,
            "training_stage": "queued",
            "training_message": "Queued",
            "training_progress_percent": 0,
            "cancel_requested": False,
            "cancelled_at": None,
        }
    )
    return job_id


def _update_background_job(job_id, **updates):
    updates["updated_at"] = _utcnow()
    get_collection("attendance", "background_jobs").update_one(
        {"job_id": job_id},
        {"$set": updates},
    )


def _get_background_job(job_id):
    return get_collection("attendance", "background_jobs").find_one({"job_id": job_id})

def _is_job_cancel_requested(job_id):
    job = _get_background_job(job_id)
    if not job:
        return False
    return bool(job.get("cancel_requested", False))

def _raise_if_job_cancelled(job_id):
    if job_id and _is_job_cancel_requested(job_id):
        raise _JobCancelledError("Job cancelled by user")


def _get_queue_names():
    queue_name = current_app.config.get("TASK_QUEUE_NAME", "biometric:jobs")
    delayed_queue_name = f"{queue_name}:delayed"
    return queue_name, delayed_queue_name


def _get_task_queue_client():
    global _QUEUE_CLIENT, _QUEUE_UNAVAILABLE_LOGGED
    if not current_app.config.get("TASK_QUEUE_ENABLED", False):
        return None

    if redis is None:
        return None

    with _QUEUE_CLIENT_LOCK:
        if _QUEUE_CLIENT is not None:
            return _QUEUE_CLIENT

        queue_url = current_app.config.get("TASK_QUEUE_REDIS_URL")
        if not queue_url:
            return None

        try:
            client = redis.Redis.from_url(queue_url, decode_responses=True)
            # Validate connectivity once so callers don't explode on first command.
            client.ping()
            _QUEUE_CLIENT = client
            _QUEUE_UNAVAILABLE_LOGGED = False
            return _QUEUE_CLIENT
        except Exception as exc:
            _QUEUE_CLIENT = None
            if not _QUEUE_UNAVAILABLE_LOGGED:
                current_app.logger.warning(
                    "Redis unavailable (%s). Falling back to local in-process queue mode.",
                    exc,
                )
                _QUEUE_UNAVAILABLE_LOGGED = True
            return None


def _enqueue_background_job(job_id, delay_seconds=0):
    client = _get_task_queue_client()
    if client is None:
        return False

    queue_name, delayed_queue_name = _get_queue_names()
    delay_seconds = max(0, int(delay_seconds or 0))
    if delay_seconds > 0:
        run_at = int(time.time()) + delay_seconds
        client.zadd(delayed_queue_name, {job_id: run_at})
    else:
        client.lpush(queue_name, job_id)
    return True


def promote_due_delayed_jobs(max_items=100):
    client = _get_task_queue_client()
    if client is None:
        return 0

    queue_name, delayed_queue_name = _get_queue_names()
    now_ts = int(time.time())
    moved = 0
    candidates = client.zrangebyscore(delayed_queue_name, 0, now_ts, start=0, num=max(1, int(max_items)))
    for job_id in candidates:
        if client.zrem(delayed_queue_name, job_id):
            client.lpush(queue_name, job_id)
            moved += 1
    return moved


def _compute_retry_delay_seconds(attempt_number):
    base = max(1, _to_int(current_app.config.get("TASK_QUEUE_BASE_BACKOFF_SECONDS", 10), 10))
    cap = max(base, _to_int(current_app.config.get("TASK_QUEUE_MAX_BACKOFF_SECONDS", 300), 300))
    exponent = max(0, int(attempt_number) - 1)
    raw_delay = min(base * (2 ** exponent), cap)
    jitter_ratio = _to_float(current_app.config.get("TASK_QUEUE_BACKOFF_JITTER_RATIO", 0.25), 0.25)
    jitter_ratio = max(0.0, min(jitter_ratio, 0.9))
    jitter_multiplier = random.uniform(1.0 - jitter_ratio, 1.0 + jitter_ratio)  # nosec B311
    return max(1, int(round(raw_delay * jitter_multiplier)))


def recover_stuck_background_jobs(max_items=50):
    jobs = get_collection("attendance", "background_jobs")
    timeout_seconds = max(30, _to_int(current_app.config.get("TASK_QUEUE_RUNNING_TIMEOUT_SECONDS", 900), 900))
    cutoff = _utcnow() - timedelta(seconds=timeout_seconds)
    stale_jobs = list(
        jobs.find(
            {"status": "running", "updated_at": {"$lte": cutoff}},
            {"job_id": 1},
        ).limit(max(1, int(max_items)))
    )

    recovered = 0
    for row in stale_jobs:
        job_id = _as_text(row.get("job_id"))
        if not job_id:
            continue

        now = _utcnow()
        res = jobs.update_one(
            {"job_id": job_id, "status": "running"},
            {
                "$set": {
                    "status": "queued",
                    "next_attempt_at": now,
                    "updated_at": now,
                    "error": "Recovered from stale running state",
                }
            },
        )
        if not res.modified_count:
            continue

        recovered += 1
        try:
            enqueued = _enqueue_background_job(job_id)
        except Exception:
            current_app.logger.exception("Recovery enqueue failed for stale job %s", job_id)
            enqueued = False

        if not enqueued:
            _schedule_local_retry(job_id, 1)

    return recovered


def _schedule_local_retry(job_id, delay_seconds):
    app = current_app._get_current_object()

    def _runner():
        time.sleep(max(0, int(delay_seconds or 0)))
        with app.app_context():
            process_background_job(job_id)

    Thread(target=_runner, daemon=True).start()


def _execute_background_job(job):
    job_type = _as_text(job.get("job_type"))
    payload = job.get("payload") or {}
    job_id = _as_text(job.get("job_id"))

    if job_type == "train_face_from_dataset":
        actor_id = _as_text(payload.get("actor_id"))
        user_id = _as_text(payload.get("user_id"))
        return _train_single_face_job(actor_id, user_id, job_id=job_id)

    if job_type == "bulk_train_face":
        actor_id = _as_text(payload.get("actor_id"))
        user_ids = payload.get("user_ids") or []
        return _train_bulk_faces_job(actor_id, user_ids, job_id=job_id)

    if job_type == "rebuild_all_face_embeddings":
        actor_id = _as_text(payload.get("actor_id"))
        return _rebuild_all_faces_job(actor_id, job_id=job_id)

    raise ValueError(f"Unsupported background job type: {job_type}")


def process_background_job(job_id):
    job = _get_background_job(job_id)
    if not job:
        return {"status": "missing"}

    status = _as_text(job.get("status")).lower()
    if status in {"completed", "dead_letter", "cancelled"}:
        return {"status": "skipped", "reason": status}

    if bool(job.get("cancel_requested", False)):
        _update_background_job(
            job_id,
            status="cancelled",
            training_stage="cancelled",
            training_message="Cancelled by user",
            finished_at=_utcnow(),
            cancelled_at=_utcnow(),
            retry_in_seconds=None,
        )
        return {"status": "cancelled"}

    max_attempts = max(1, _to_int(job.get("max_attempts"), 3))
    attempts = max(0, _to_int(job.get("attempts"), 0))
    current_attempt = attempts + 1

    _update_background_job(
        job_id,
        status="running",
        attempts=current_attempt,
        started_at=job.get("started_at") or _utcnow(),
        next_attempt_at=None,
        training_stage="running",
    )
    try:
        _raise_if_job_cancelled(job_id)
        result = _execute_background_job(job)
        _update_background_job(
            job_id,
            status="completed",
            training_stage="completed",
            training_message="Training complete",
            training_progress_percent=100,
            result=result,
            error=None,
            finished_at=_utcnow(),
            retry_in_seconds=None,
        )
        _clear_query_cache()
        return {"status": "completed", "result": result}
    except _JobCancelledError:
        _update_background_job(
            job_id,
            status="cancelled",
            training_stage="cancelled",
            training_message="Cancelled by user",
            finished_at=_utcnow(),
            cancelled_at=_utcnow(),
            error=None,
            retry_in_seconds=None,
        )
        return {"status": "cancelled"}
    except Exception as exc:
        current_app.logger.exception("Background job %s failed", job_id)
        error_text = str(exc)
        error_now = _utcnow()
        existing_history = job.get("error_history") or []
        if not isinstance(existing_history, list):
            existing_history = []

        next_history = list(existing_history)
        next_history.append(
            {
                "attempt": current_attempt,
                "error": error_text,
                "at": error_now,
            }
        )
        if len(next_history) > 10:
            next_history = next_history[-10:]

        if current_attempt < max_attempts:
            delay_seconds = _compute_retry_delay_seconds(current_attempt)
            next_attempt = error_now + timedelta(seconds=delay_seconds)
            _update_background_job(
                job_id,
                status="queued",
                error=error_text,
                next_attempt_at=next_attempt,
                retry_count=current_attempt,
                retry_in_seconds=delay_seconds,
                last_error_at=error_now,
                error_history=next_history,
            )

            try:
                enqueued = _enqueue_background_job(job_id, delay_seconds=delay_seconds)
            except Exception:
                current_app.logger.exception("Retry enqueue failed for job %s", job_id)
                enqueued = False

            if not enqueued:
                _schedule_local_retry(job_id, delay_seconds)

            return {
                "status": "retry_scheduled",
                "attempt": current_attempt,
                "max_attempts": max_attempts,
                "retry_in_seconds": delay_seconds,
            }

        _update_background_job(
            job_id,
            status="dead_letter",
            error=error_text,
            finished_at=error_now,
            dead_lettered_at=error_now,
            retry_count=max(0, current_attempt - 1),
            retry_in_seconds=None,
            last_error_at=error_now,
            error_history=next_history,
        )
        return {
            "status": "dead_letter",
            "attempt": current_attempt,
            "max_attempts": max_attempts,
            "error": error_text,
        }


def _run_background_job(app, job_id):
    with app.app_context():
        process_background_job(job_id)


def _launch_background_job(app, job_type, payload):
    job_id = _create_background_job(job_type, payload)

    if current_app.config.get("TASK_QUEUE_ENABLED", False):
        try:
            if _enqueue_background_job(job_id):
                return job_id
        except Exception:
            app.logger.exception("Queue enqueue failed for job %s; falling back to local thread", job_id)

    thread = Thread(target=_run_background_job, args=(app, job_id), daemon=True)
    thread.start()
    return job_id


def _as_text(value):
    return str(value or "").strip()


def _normalise_year(value):
    if value is None:
        return ""
    text = _as_text(value)
    return text


def _to_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _as_object_id(value):
    if isinstance(value, ObjectId):
        return value
    text = _as_text(value)
    if not text:
        return None
    try:
        return ObjectId(text)
    except Exception:
        return None


def _id_variants(value):
    """Return equivalent ID representations to handle mixed string/ObjectId legacy data."""
    variants = []
    text = _as_text(value)
    if text:
        variants.append(text)
    oid = _as_object_id(value)
    if oid is not None and oid not in variants:
        variants.append(oid)
    return variants


def _normalise_filter_ids(filter_doc):
    if not isinstance(filter_doc, dict):
        return filter_doc
    out = dict(filter_doc)
    if "_id" in out:
        if isinstance(out["_id"], dict):
            oid_map = dict(out["_id"])
            if "$in" in oid_map and isinstance(oid_map["$in"], list):
                oid_map["$in"] = [_as_object_id(v) or v for v in oid_map["$in"]]
            out["_id"] = oid_map
        else:
            out["_id"] = _as_object_id(out["_id"]) or out["_id"]
    return out


def _normalise_document_ids(doc):
    if not isinstance(doc, dict):
        return doc
    out = dict(doc)
    if "_id" in out:
        out["_id"] = _as_object_id(out["_id"]) or out["_id"]
    return out


def _rb_delete(db, collection, filter_doc):
    return {
        "type": "delete_document",
        "db": db,
        "collection": collection,
        "filter": filter_doc,
    }


def _rb_restore(db, collection, document):
    return {
        "type": "restore_document",
        "db": db,
        "collection": collection,
        "document": document,
    }


def _rb_replace(db, collection, filter_doc, previous_document):
    return {
        "type": "replace_document",
        "db": db,
        "collection": collection,
        "filter": filter_doc,
        "previous_document": previous_document,
    }


def _rb_batch(operations):
    return {"type": "batch", "operations": operations}


def _execute_rollback_operation(operation):
    op_type = operation.get("type")
    if op_type == "batch":
        for op in operation.get("operations") or []:
            _execute_rollback_operation(op)
        return

    db = operation.get("db")
    collection = operation.get("collection")
    if not db or not collection:
        return

    col = get_collection(db, collection)

    if op_type == "delete_document":
        filt = _normalise_filter_ids(operation.get("filter") or {})
        if filt:
            col.delete_many(filt)
        return

    if op_type == "replace_document":
        filt = _normalise_filter_ids(operation.get("filter") or {})
        prev_doc = _normalise_document_ids(operation.get("previous_document") or {})
        if filt and prev_doc:
            col.replace_one(filt, prev_doc, upsert=True)
        return

    if op_type == "restore_document":
        doc = _normalise_document_ids(operation.get("document") or {})
        if not doc:
            return

        if doc.get("_id") is not None:
            col.replace_one({"_id": doc.get("_id")}, doc, upsert=True)
        elif doc.get("user_id"):
            col.replace_one({"user_id": doc.get("user_id")}, doc, upsert=True)
        else:
            col.insert_one(doc)


def _derive_academic_session(enrollment_year, course_duration):
    """Build session label like 2024-26 from start year and duration."""
    start_year = _to_int(enrollment_year, datetime.now(timezone.utc).year)
    duration_years = max(1, _to_int(course_duration, 1))
    end_year_short = str(start_year + duration_years)[-2:]
    return f"{start_year}-{end_year_short}"


def _safe_find_user(user_id):
    try:
        return find_user_by_id(user_id)
    except Exception:
        return None


def _safe_get_profile_by_id(profile_id):
    try:
        return get_profile_by_id(profile_id)
    except Exception:
        return None


def _get_profile_by_user_any(user_id):
    """Resolve student profile by user_id across legacy string/ObjectId storage."""
    variants = _id_variants(user_id)
    if not variants:
        return None
    profiles = get_collection("academic", "student_profiles")
    return profiles.find_one({"user_id": {"$in": variants}})


def _safe_get_course(course_id):
    try:
        return get_course_by_id(course_id)
    except Exception:
        return None


def _course_is_inactive(course):
    if not course:
        return True
    return str(course.get("status") or "active").lower() != "active"


def _get_active_course_or_error(course_id):
    course = _safe_get_course(course_id)
    if not course:
        return None, (jsonify({"error": "Course not found"}), 404)
    if _course_is_inactive(course):
        return None, (jsonify({"error": "Course is inactive. Reassign entities to an active course first."}), 409)
    return course, None


def _ensure_student_course_active(user_id):
    profile = _get_profile_by_user_any(user_id)
    if not profile:
        return None, (jsonify({"error": "Student profile not found"}), 404)
    course = _safe_get_course(profile.get("course_id"))
    if _course_is_inactive(course):
        return profile, (jsonify({"error": "Student is linked to an inactive course and is read-only"}), 409)
    return profile, None


def _ensure_paper_course_active(paper):
    course_id = _as_text((paper or {}).get("course_id"))
    if not course_id:
        return jsonify({"error": "Paper is not linked to a valid course"}), 409
    if not is_course_active(course_id):
        return jsonify({"error": "Paper is linked to an inactive course and is read-only"}), 409
    return None


def _detach_lecturers_from_course_papers(course_id):
    """Clear lecturer assignments for all papers under a given course."""
    papers_col = get_collection("academic", "papers")
    course_keys = {_as_text(v) for v in _id_variants(course_id) if _as_text(v)}
    if not course_keys:
        return 0, []

    assigned_papers = list(
        papers_col.find({"lecturer_id": {"$nin": [None, ""]}})
    )

    affected_papers = [
        p for p in assigned_papers
        if _as_text(p.get("course_id")) in course_keys
    ]
    if not affected_papers:
        return 0, []

    affected_ids = [p.get("_id") for p in affected_papers if p.get("_id") is not None]
    if not affected_ids:
        return 0, []

    res = papers_col.update_many(
        {"_id": {"$in": affected_ids}},
        {"$set": {"lecturer_id": None}},
    )
    return int(res.modified_count or 0), affected_papers


def _legacy_safe_name(raw_value):
    text = _as_text(raw_value)
    cleaned = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in text)
    return cleaned.strip("_") or "unknown"


def _resolve_dataset_dir_for_user(user_id):
    user_id_text = _as_text(user_id)
    dataset_dir = os.path.join("dataset", user_id_text)
    if os.path.isdir(dataset_dir):
        return dataset_dir

    # Backward compatibility for previously created name-based dataset folders.
    student = find_user_by_id(user_id) or {}
    legacy_name = _as_text(student.get("name"))
    legacy_dataset_dir = os.path.join("dataset", _legacy_safe_name(legacy_name))
    if os.path.isdir(legacy_dataset_dir):
        return legacy_dataset_dir

    return dataset_dir


def _train_embeddings_from_dataset_for_user(user_id):
    dataset_dir = _resolve_dataset_dir_for_user(user_id)
    if not os.path.isdir(dataset_dir):
        raise ValueError(
            f"Dataset folder not found for this student. Please run Enroll Face first to capture dataset images. Expected: {dataset_dir}"
        )

    allowed_ext = {".jpg", ".jpeg", ".png"}
    image_files = [
        os.path.join(dataset_dir, name)
        for name in sorted(os.listdir(dataset_dir))
        if os.path.splitext(name.lower())[1] in allowed_ext
    ]

    if not image_files:
        raise ValueError("No dataset images found for training")

    detector = get_detector()
    embeddings = []
    skipped = 0
    seen_signatures = set()

    for file_path in image_files:
        img_bgr = cv2.imread(file_path, cv2.IMREAD_COLOR)
        if img_bgr is None:
            skipped += 1
            continue

        try:
            img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
            faces = detector.detect_faces(img_rgb)
            if not faces:
                skipped += 1
                continue

            embedding = normalize_embedding(generate_embedding(faces[0]["crop"]))
            signature = tuple(np.round(np.asarray(embedding, dtype=np.float32), 3))
            if signature in seen_signatures:
                skipped += 1
                continue

            seen_signatures.add(signature)
            embeddings.append(embedding)
        except Exception:
            skipped += 1
            continue

    if not embeddings:
        raise ValueError("Training failed: no valid faces found in dataset images")

    # Keep the strongest set of embeddings, but cap the storage size to reduce noisy duplicates.
    if len(embeddings) > 25:
        embeddings = embeddings[:25]

    set_face_embeddings(user_id, embeddings)
    return {
        "dataset_dir": dataset_dir,
        "trained_embeddings": len(embeddings),
        "skipped_images": skipped,
    }


def _refresh_face_trainer_artifact():
    """Rebuild the showcase .h5 trainer from the current dataset tree."""
    try:
        return train_and_save_face_model(
            dataset_root="dataset",
            trainer_dir="trainer",
            model_filename="face_trainer.keras",
        )
    except Exception as exc:
        current_app.logger.exception("Face trainer export failed")
        return {
            "model_path": None,
            "error": str(exc),
        }


def _resolve_user_identity(user_identifier):
    """Resolve route id that may be either user_id or profile_id."""
    profile = _get_profile_by_user_any(user_identifier)
    if profile:
        return _as_text(profile.get("user_id")) or _as_text(user_identifier), profile

    profile = _safe_get_profile_by_id(user_identifier)
    if profile:
        return _as_text(profile.get("user_id")), profile

    user = _safe_find_user(user_identifier)
    if user and user.get("role") == "student":
        return _as_text(user_identifier), _get_profile_by_user_any(user_identifier)

    return None, None


def _generate_registration_number(course, academic_session, exclude_user_id=None):
    """Generate a unique registration number, compatible with legacy and new session fields."""
    prefix = re.sub(r"[^A-Za-z0-9]", "", (course or {}).get("code", "STU")).upper() or "STU"
    session = _as_text(academic_session) or "NA"
    course_id = _as_text((course or {}).get("_id"))

    profiles = get_collection("academic", "student_profiles")
    query = {
        "course_id": course_id,
        "$or": [
            {"academic_session": session},
            {"academic_year": session},
            {"year": session},
        ],
    }

    pattern = re.compile(rf"^{re.escape(prefix)}-{re.escape(session)}-(\\d+)$")
    max_seq = 0
    for row in profiles.find(query, {"reg_number": 1}):
        reg = _as_text(row.get("reg_number"))
        m = pattern.match(reg)
        if m:
            max_seq = max(max_seq, _to_int(m.group(1), 0))

    seq = max_seq + 1
    while True:
        candidate = f"{prefix}-{session}-{seq:03d}"
        existing = profiles.find_one({"reg_number": candidate}, {"user_id": 1})
        if not existing:
            return candidate
        if exclude_user_id and _as_text(existing.get("user_id")) == _as_text(exclude_user_id):
            return candidate
        seq += 1


def _enrich_paper(paper, course_map, lecturer_map):
    item = sanitise_mongo_doc(paper)
    course = course_map.get(item.get("course_id"))
    lecturer = lecturer_map.get(item.get("lecturer_id"))
    item["course_name"] = course.get("name") if course else None
    item["course_code"] = course.get("code") if course else None
    item["course_status"] = _as_text((course or {}).get("status") or "active").lower() or "active"
    item["is_course_inactive"] = item["course_status"] != "active"
    item["semester"] = item.get("semester")
    item["academic_year"] = item.get("academic_session") or item.get("academic_year")
    item["lecturer_name"] = lecturer.get("name") if lecturer else None
    item["lecturer_email"] = lecturer.get("email") if lecturer else None
    return item


def _to_bool(value):
    if isinstance(value, bool):
        return value
    text = _as_text(value).lower()
    return text in {"1", "true", "yes", "y"}


def _get_requested_pagination():
    page_raw = request.args.get("page")
    per_page_raw = request.args.get("per_page")

    if not _as_text(page_raw) and not _as_text(per_page_raw):
        return None

    page = max(1, _to_int(page_raw, 1))
    per_page = max(1, min(_to_int(per_page_raw, 20), 100))
    return page, per_page


def _paginate_items(items):
    pagination = _get_requested_pagination()
    if not pagination:
        return jsonify(items)

    page, per_page = pagination
    total = len(items)
    start = (page - 1) * per_page
    end = start + per_page

    return jsonify({
        "items": items[start:end],
        "total": total,
        "page": page,
        "per_page": per_page,
    })


# ─── Courses ────────────────────────────────────────────────────────────────


@admin_bp.route("/courses", methods=["GET"])
@role_required("super_admin", "department_admin")
def list_courses(user):
    # Unified department filter logic
    dept_id = None
    if is_super_admin(user):
        dept_id = _as_text(request.args.get("department_id", "")).strip() or None
    else:
        dept_id = _user_dept_id(user)

    # Fetch all courses first; apply scoped filtering with legacy fallback below.
    courses = sanitise_many(get_all_courses(["name", "code", "department", "course_duration", "status", "department_id"], department_id=None))

    if dept_id:
        selected_dept_id = _as_text(dept_id).strip()
        selected_dept_name = ""
        selected_dept = None
        try:
            selected_dept = get_department_by_id(selected_dept_id)
        except Exception:
            selected_dept = None
        if selected_dept:
            selected_dept_name = _as_text(selected_dept.get("name", "")).strip().lower()

        scoped_courses = []
        for course in courses:
            course_dept_id = _as_text(course.get("department_id", "")).strip()
            course_dept_name = _as_text(course.get("department", "")).strip().lower()
            if course_dept_id and course_dept_id == selected_dept_id:
                scoped_courses.append(course)
                continue
            # Legacy fallback for old course records that only stored department name.
            if selected_dept_name and course_dept_name and course_dept_name == selected_dept_name:
                scoped_courses.append(course)
        courses = scoped_courses

    q = _as_text(request.args.get("q", "")).lower()
    course_duration = _as_text(request.args.get("course_duration", ""))
    status = _as_text(request.args.get("status", "")).lower()

    filtered = []
    for c in courses:
        c["status"] = _as_text(c.get("status") or "active").lower() or "active"
        if course_duration and str(c.get("course_duration", "")) != course_duration:
            continue
        if status and c.get("status") != status:
            continue
        if q and not (
            q in _as_text(c.get("name")).lower()
            or q in _as_text(c.get("code")).lower()
            or q in _as_text(c.get("department")).lower()
        ):
            continue
        filtered.append(c)

    return _paginate_items(filtered)


@admin_bp.route("/courses", methods=["POST"])
@role_required("super_admin", "department_admin")
def add_course(user):
    d = request.get_json(silent=True) or {}
    if not d.get("name") or not d.get("code") or not d.get("course_duration"):
        return jsonify({"error": "name, code and course_duration are required"}), 400
    # Resolve department_id: dept admins use their own, super admins may pass in body
    dept_id = None
    body_dept_id = _as_text(d.get("department_id", "")).strip()
    if is_super_admin(user):
        dept_id = body_dept_id or None
    else:
        dept_id = _user_dept_id(user)
    course = create_course(
        d["name"],
        d["code"],
        d.get("department", ""),
        _to_int(d.get("course_duration"), 0),
        department_id=dept_id,
    )
    log_action(
        "CREATE_COURSE",
        str(user["_id"]),
        details=f"Course {d['code']}",
        rollback=_rb_delete("academic", "courses", {"_id": course.get("_id")}),
    )
    _clear_query_cache()
    return jsonify(sanitise_mongo_doc(course)), 201


@admin_bp.route("/courses/<cid>/semesters", methods=["GET"])
@role_required("super_admin", "department_admin")
def list_course_semesters(user, cid):
    """Return available semesters for a course (duration-derived + paper-derived)."""
    course = _safe_get_course(cid)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    semesters = set()

    duration_years = max(1, _to_int(course.get("course_duration"), 1))
    for sem in range(1, duration_years * 2 + 1):
        semesters.add(sem)

    papers = get_papers_by_course(cid)
    for paper in papers:
        sem = _to_int(paper.get("semester"), 0)
        if sem > 0:
            semesters.add(sem)

    return jsonify(sorted(list(semesters)))


@admin_bp.route("/courses/<cid>", methods=["GET"])
@role_required("department_admin")
@validate_ids("cid")
def get_course_details(user, cid):
    course = get_course_by_id(cid)
    if not course:
        return jsonify({"error": "Course not found"}), 404
    return jsonify(sanitise_mongo_doc(course))


@admin_bp.route("/courses/<cid>/sessions", methods=["GET"])
@role_required("department_admin")
def list_course_sessions(user, cid):
    """Return distinct academic sessions for a course."""
    course = _safe_get_course(cid)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    profiles = get_collection("academic", "student_profiles")
    sessions = set()
    course_duration = max(1, _to_int(course.get("course_duration"), 1))
    for row in profiles.aggregate([
        {"$match": {"course_id": cid}},
        {
            "$project": {
                "session": {
                    "$ifNull": [
                        "$academic_session",
                        {
                            "$ifNull": [
                                "$academic_year",
                                {
                                    "$ifNull": [
                                        "$year",
                                        {"$toString": {"$year": "$created_at"}},
                                    ]
                                },
                            ]
                        },
                    ]
                }
            }
        },
        {"$group": {"_id": "$session"}},
    ]):
        session = _as_text(row.get("_id"))
        if session:
            sessions.add(session)

    # Ensure at least current derived session appears when course has active profiles but no stored session field.
    if not sessions and profiles.count_documents({"course_id": cid}) > 0:
        now_year = datetime.now(timezone.utc).year
        sessions.add(_derive_academic_session(now_year, course_duration))

    return jsonify(sorted(sessions))


def _normalise_course_semester(course_id, semester):
    cid = _as_text(course_id)
    if not cid:
        return None, None, "course_id is required"

    course = _safe_get_course(cid)
    if not course:
        return None, None, "Course not found"
    if _course_is_inactive(course):
        return None, None, "Course is inactive. Reassign entities to an active course first"

    sem = _to_int(semester, 0)
    if sem <= 0:
        return None, None, "semester must be a positive integer"

    max_sem = max(1, _to_int(course.get("course_duration"), 1) * 2)
    if sem > max_sem:
        return None, None, f"semester must be between 1 and {max_sem} for selected course"

    return cid, sem, None


@admin_bp.route("/courses/<cid>", methods=["PUT"])
@role_required("department_admin")
@validate_ids("cid")
def edit_course(user, cid):
    d = request.get_json(silent=True) or {}
    allowed = {"name", "code", "department", "course_duration", "status"}
    fields = {k: v for k, v in d.items() if k in allowed}
    if "course_duration" in fields:
        fields["course_duration"] = _to_int(fields.get("course_duration"), 0)
    if "status" in fields:
        next_status = _as_text(fields.get("status")).lower()
        if next_status not in {"active", "inactive"}:
            return jsonify({"error": "status must be active or inactive"}), 400
        fields["status"] = next_status

    previous = get_course_by_id(cid)
    if not previous:
        return jsonify({"error": "Course not found"}), 404

    prev_status = _as_text(previous.get("status") or "active").lower() or "active"
    next_status = _as_text(fields.get("status") or prev_status).lower() or "active"

    detached_count = 0
    detached_papers = []
    if prev_status == "active" and next_status == "inactive":
        detached_count, detached_papers = _detach_lecturers_from_course_papers(cid)

    updated = update_course(cid, fields)

    rollback_ops = [
        _rb_replace("academic", "courses", {"_id": cid}, previous),
    ]
    for doc in detached_papers:
        rollback_ops.append(_rb_replace("academic", "papers", {"_id": doc.get("_id")}, doc))

    details = f"Course {cid}"
    if detached_count > 0:
        details = f"Course {cid}; detached lecturers from {detached_count} paper(s)"

    log_action(
        "UPDATE_COURSE",
        str(user["_id"]),
        details=details,
        rollback=_rb_batch(rollback_ops),
    )
    _clear_query_cache()
    return jsonify(sanitise_mongo_doc(updated))


@admin_bp.route("/courses/<cid>", methods=["DELETE"])
@role_required("department_admin")
@validate_ids("cid")
def remove_course(user, cid):
    previous = get_course_by_id(cid)
    if not previous:
        return jsonify({"error": "Course not found"}), 404

    detached_count, detached_papers = _detach_lecturers_from_course_papers(cid)
    delete_course(cid)

    rollback_ops = [_rb_restore("academic", "courses", previous)]
    for doc in detached_papers:
        rollback_ops.append(_rb_replace("academic", "papers", {"_id": doc.get("_id")}, doc))

    details = f"Course {cid}"
    if detached_count > 0:
        details = f"Course {cid}; detached lecturers from {detached_count} paper(s)"

    log_action(
        "DEACTIVATE_COURSE",
        str(user["_id"]),
        details=details,
        rollback=_rb_batch(rollback_ops),
    )
    _clear_query_cache()
    return jsonify({
        "message": "Course marked inactive",
        "detached_lecturer_assignments": detached_count,
    }), 200


# ─── Papers ─────────────────────────────────────────────────────────────────

@admin_bp.route("/papers", methods=["GET"])
@role_required("super_admin", "department_admin")
def list_papers(user):
    dept_id = None
    if is_super_admin(user):
        dept_id = _as_text(request.args.get("department_id", "")).strip() or None
    else:
        dept_id = _user_dept_id(user)

    # Fetch full sets first; apply scoped filtering with legacy fallback below.
    papers = get_all_papers(["name", "code", "course_id", "lecturer_id", "semester", "total_classes", "created_at", "department_id"], department_id=None)
    courses = sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year", "department_id"], department_id=None))

    if dept_id:
        selected_dept_id = _as_text(dept_id).strip()
        selected_dept_name = ""
        selected_dept = None
        try:
            selected_dept = get_department_by_id(selected_dept_id)
        except Exception:
            selected_dept = None
        if selected_dept:
            selected_dept_name = _as_text(selected_dept.get("name", "")).strip().lower()

        scoped_course_ids = set()
        scoped_courses = []
        for course in courses:
            course_dept_id = _as_text(course.get("department_id", "")).strip()
            course_dept_name = _as_text(course.get("department", "")).strip().lower()
            if course_dept_id and course_dept_id == selected_dept_id:
                scoped_courses.append(course)
                scoped_course_ids.add(course.get("_id"))
                continue
            # Legacy fallback for old records linked only by department name.
            if selected_dept_name and course_dept_name and course_dept_name == selected_dept_name:
                scoped_courses.append(course)
                scoped_course_ids.add(course.get("_id"))

        courses = scoped_courses
        papers = [paper for paper in papers if paper.get("course_id") in scoped_course_ids]

    lecturers = sanitise_many(get_users_by_role("lecturer", department_id=dept_id))
    course_map = {c["_id"]: c for c in courses}
    lecturer_map = {l["_id"]: l for l in lecturers}

    department_filter = _as_text(request.args.get("department", ""))
    q = _as_text(request.args.get("q", "")).lower()
    course_id = _as_text(request.args.get("course_id", ""))
    lecturer_id = _as_text(request.args.get("lecturer_id", ""))
    semester = _as_text(request.args.get("semester", ""))
    academic_year = _normalise_year(request.args.get("academic_year", ""))

    # Filter by department name on the associated course
    if department_filter:
        dept_course_ids = {
            c["_id"] for c in courses
            if _as_text(c.get("department") or "").lower() == department_filter.lower()
        }
        # Restrict course_map and papers to matching department courses
        course_map = {cid: c for cid, c in course_map.items() if cid in dept_course_ids}
        papers = [p for p in papers if p.get("course_id") in dept_course_ids]

    result = []
    for paper in papers:
        item = _enrich_paper(paper, course_map, lecturer_map)
        if course_id and item.get("course_id") != course_id:
            continue
        if lecturer_id and item.get("lecturer_id") != lecturer_id:
            continue
        if semester and str(item.get("semester", "")) != semester:
            continue
        if academic_year and _normalise_year(item.get("academic_year")) != academic_year:
            continue
        if q and not (
            q in _as_text(item.get("name")).lower()
            or q in _as_text(item.get("code")).lower()
            or q in _as_text(item.get("course_name")).lower()
            or q in _as_text(item.get("lecturer_name")).lower()
        ):
            continue
        result.append(item)

    return _paginate_items(sanitise_many(result))


@admin_bp.route("/papers/<pid>", methods=["GET"])
@role_required("department_admin")
@validate_ids("pid")
def get_paper_details(user, pid):
    paper = get_paper_by_id(pid)
    if not paper:
        return jsonify({"error": "Paper not found"}), 404
    return jsonify(sanitise_mongo_doc(paper))


@admin_bp.route("/papers", methods=["POST"])
@role_required("department_admin")
def add_paper(user):
    d = request.get_json(silent=True) or {}
    if not d.get("name") or not d.get("code") or not d.get("course_id") or not d.get("semester"):
        return jsonify({"error": "name, code, course_id and semester are required"}), 400

    course_id, semester, error = _normalise_course_semester(d.get("course_id"), d.get("semester"))
    if error:
        return jsonify({"error": error}), 400

    paper = create_paper(
        d["name"],
        d["code"],
        course_id,
        d.get("lecturer_id") or None,
        semester,
        d.get("total_classes", 0),
        department_id=_user_dept_id(user),
    )
    log_action(
        "CREATE_PAPER",
        str(user["_id"]),
        details=f"Paper {d['code']}",
        rollback=_rb_delete("academic", "papers", {"_id": paper.get("_id")}),
    )
    _clear_query_cache()
    return jsonify(sanitise_mongo_doc(paper)), 201


@admin_bp.route("/papers/<pid>", methods=["PUT"])
@role_required("department_admin")
@validate_ids("pid")
def edit_paper(user, pid):
    d = request.get_json(silent=True) or {}
    fields = dict(d)
    if "lecturer_id" in fields and not fields["lecturer_id"]:
        fields["lecturer_id"] = None

    previous = get_paper_by_id(pid)
    if not previous:
        return jsonify({"error": "Paper not found"}), 404

    lock_error = _ensure_paper_course_active(previous)
    if lock_error:
        return lock_error

    if "course_id" in fields or "semester" in fields:
        next_course_id = fields.get("course_id", previous.get("course_id"))
        next_semester = fields.get("semester", previous.get("semester"))
        course_id, semester, error = _normalise_course_semester(next_course_id, next_semester)
        if error:
            return jsonify({"error": error}), 400
        fields["course_id"] = course_id
        fields["semester"] = semester

    updated = update_paper(pid, fields)
    log_action(
        "UPDATE_PAPER",
        str(user["_id"]),
        details=f"Paper {pid}",
        rollback=_rb_replace("academic", "papers", {"_id": pid}, previous) if previous else None,
    )
    _clear_query_cache()
    return jsonify(sanitise_mongo_doc(updated))


@admin_bp.route("/papers/<pid>", methods=["DELETE"])
@role_required("department_admin")
@validate_ids("pid")
def remove_paper(user, pid):
    previous = get_paper_by_id(pid)
    if not previous:
        return jsonify({"error": "Paper not found"}), 404

    lock_error = _ensure_paper_course_active(previous)
    if lock_error:
        return lock_error

    delete_paper(pid)
    log_action(
        "DELETE_PAPER",
        str(user["_id"]),
        details=f"Paper {pid}",
        rollback=_rb_restore("academic", "papers", previous) if previous else None,
    )
    _clear_query_cache()
    return jsonify({"message": "Deleted"}), 200


@admin_bp.route("/papers/bulk-assign", methods=["POST"])
@role_required("super_admin", "department_admin")
def bulk_assign(user):
    """Assign multiple papers to a lecturer or course in one click."""
    d = request.get_json(silent=True) or {}

    # Student enrollment flow: assign one paper to many students.
    paper_id = _as_text(d.get("paper_id"))
    user_ids = [_as_text(sid) for sid in (d.get("user_ids") or []) if _as_text(sid)]
    if paper_id and user_ids:
        paper = get_paper_by_id(paper_id)
        if not paper:
            return jsonify({"error": "Paper not found"}), 404

        lock_error = _ensure_paper_course_active(paper)
        if lock_error:
            return lock_error

        updated_count = 0
        for sid in user_ids:
            uid, _ = _resolve_user_identity(sid)
            if not uid:
                continue
            _, student_lock_error = _ensure_student_course_active(uid)
            if student_lock_error:
                continue
            changed = enroll_in_papers(uid, [paper_id])
            if changed > 0:
                updated_count += 1

        if updated_count <= 0:
            return jsonify({"error": "No eligible students could be assigned"}), 400

        log_action(
            "BULK_ENROLL_STUDENTS",
            str(user["_id"]),
            details=f"Paper {paper_id}, students {updated_count}",
        )
        _clear_query_cache()
        return jsonify({"message": "Students enrolled successfully", "updated_count": updated_count, "assigned_paper_count": 1}), 200

    # Student enrollment flow: assign many papers to many students.
    # This branch is intentionally prioritized whenever user_ids are present,
    # even if course_id is also included by the frontend payload.
    paper_ids_for_students = [_as_text(pid) for pid in (d.get("paper_ids") or []) if _as_text(pid)]
    if user_ids and paper_ids_for_students and not d.get("lecturer_id"):
        valid_paper_ids = []
        for pid in paper_ids_for_students:
            paper = get_paper_by_id(pid)
            if not paper:
                continue
            lock_error = _ensure_paper_course_active(paper)
            if lock_error:
                continue
            valid_paper_ids.append(pid)

        if not valid_paper_ids:
            return jsonify({"error": "No active papers found for assignment"}), 400

        updated_count = 0
        for sid in user_ids:
            uid, _ = _resolve_user_identity(sid)
            if not uid:
                continue
            _, student_lock_error = _ensure_student_course_active(uid)
            if student_lock_error:
                continue
            changed = enroll_in_papers(uid, valid_paper_ids)
            if changed > 0:
                updated_count += 1

        if updated_count <= 0:
            return jsonify({"error": "No eligible students could be assigned"}), 400

        log_action(
            "BULK_ENROLL_STUDENTS",
            str(user["_id"]),
            details=f"Papers {len(valid_paper_ids)}, students {updated_count}",
        )
        _clear_query_cache()
        return jsonify(
            {
                "message": "Students enrolled successfully",
                "updated_count": updated_count,
                "assigned_paper_count": len(valid_paper_ids),
            }
        ), 200

    paper_ids = d.get("paper_ids", [])
    lecturer_id = d.get("lecturer_id")
    course_id = d.get("course_id")

    if not paper_ids:
        return jsonify({"error": "paper_ids or (paper_id + user_ids) is required"}), 400

    if lecturer_id:
        for pid in paper_ids:
            paper = get_paper_by_id(pid)
            if not paper:
                continue
            lock_error = _ensure_paper_course_active(paper)
            if lock_error:
                return lock_error

        bulk_assign_lecturer(paper_ids, lecturer_id)
        log_action("BULK_ASSIGN_LECTURER", str(user["_id"]),
                   details=f"Papers {paper_ids} → Lecturer {lecturer_id}")
        _clear_query_cache()
    if course_id:
        course = _safe_get_course(course_id)
        if not course:
            return jsonify({"error": "Course not found"}), 404
        if _course_is_inactive(course):
            return jsonify({"error": "Cannot assign papers to an inactive course"}), 409
        max_sem = max(1, _to_int(course.get("course_duration"), 1) * 2)

        invalid = []
        for paper_id in paper_ids:
            paper = get_paper_by_id(paper_id)
            if not paper:
                continue
            psem = _to_int(paper.get("semester"), 0)
            if psem > max_sem:
                invalid.append({"paper_id": paper_id, "paper_code": paper.get("code", ""), "semester": psem})

        if invalid:
            return jsonify({
                "error": f"One or more papers have semester above course limit (max {max_sem})",
                "invalid_papers": invalid,
            }), 400

        bulk_assign_course(paper_ids, course_id)
        log_action("BULK_ASSIGN_COURSE", str(user["_id"]),
                   details=f"Papers {paper_ids} → Course {course_id}")
        _clear_query_cache()
    return jsonify({"message": "Assigned"}), 200


@admin_bp.route("/lecturers/<lid>/papers", methods=["GET"])
@role_required("department_admin")
def get_lecturer_papers(user, lid):
    papers = get_all_papers(["name", "code", "course_id", "lecturer_id", "semester", "total_classes", "created_at"])
    courses = sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year"]))
    lecturers = sanitise_many(get_users_by_role("lecturer"))
    course_map = {c["_id"]: c for c in courses}
    lecturer_map = {l["_id"]: l for l in lecturers}

    all_papers = [_enrich_paper(p, course_map, lecturer_map) for p in papers]
    assigned = [p for p in all_papers if p.get("lecturer_id") == lid]
    return jsonify({"assigned": assigned, "all": all_papers})


@admin_bp.route("/lecturers/<lid>/papers", methods=["PUT"])
@role_required("department_admin")
def set_lecturer_papers(user, lid):
    d = request.get_json(silent=True) or {}
    paper_ids = set(d.get("paper_ids") or [])
    object_ids = []
    for pid in paper_ids:
        try:
            object_ids.append(ObjectId(pid))
        except Exception:
            continue  # nosec B112

    # Unassign papers currently tied to lecturer but not selected now.
    papers = get_collection("academic", "papers")
    papers.update_many(
        {"lecturer_id": lid, "_id": {"$nin": object_ids}},
        {"$set": {"lecturer_id": None}},
    )

    # Assign selected papers to lecturer.
    if object_ids:
        papers.update_many(
            {"_id": {"$in": object_ids}},
            {"$set": {"lecturer_id": lid}},
        )

    log_action(
        "UPDATE_LECTURER_ASSIGNMENTS",
        str(user["_id"]),
        target_user=lid,
        details=f"Assigned papers: {sorted(list(paper_ids))}",
    )
    _clear_query_cache()
    return jsonify({"message": "Lecturer paper assignments updated"}), 200


# ─── Lecturers ──────────────────────────────────────────────────────────────

@admin_bp.route("/lecturers", methods=["GET"])
@role_required("super_admin", "department_admin")
def list_lecturers(user):
    dept_id = None
    if is_super_admin(user):
        dept_id = _as_text(request.args.get("department_id", "")).strip() or None
    else:
        dept_id = _user_dept_id(user)

    # Primary filter by department_id with legacy fallback by department name.
    # Some older lecturer records were stored without department_id.
    if dept_id:
        all_lecturers = sanitise_many(get_users_by_role("lecturer"))
        selected_dept_id = _as_text(dept_id).strip()
        selected_dept_name = ""
        selected_dept = None
        try:
            selected_dept = get_department_by_id(selected_dept_id)
        except Exception:
            selected_dept = None
        if selected_dept:
            selected_dept_name = _as_text(selected_dept.get("name", "")).strip().lower()

        lecturers = []
        for lec in all_lecturers:
            lec_dept_id = _as_text(lec.get("department_id", "")).strip()
            lec_dept_name = _as_text(lec.get("department", "")).strip().lower()
            if lec_dept_id and lec_dept_id == selected_dept_id:
                lecturers.append(lec)
                continue
            if selected_dept_name and lec_dept_name and lec_dept_name == selected_dept_name:
                lecturers.append(lec)
    else:
        lecturers = sanitise_many(get_users_by_role("lecturer", department_id=dept_id))

    papers = sanitise_many(get_all_papers(["name", "code", "lecturer_id", "course_id", "semester", "total_classes", "created_at"]))
    courses = sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year"]))
    course_map = {c["_id"]: c for c in courses}

    department_filter = _as_text(request.args.get("department", ""))
    q = _as_text(request.args.get("q", "")).lower()
    course_id = _as_text(request.args.get("course_id", ""))
    semester = _as_text(request.args.get("semester", ""))
    paper_id = _as_text(request.args.get("paper_id", ""))

    # Filter by department name on courses assigned to the lecturer
    if department_filter:
        dept_course_ids = {
            c["_id"] for c in courses
            if _as_text(c.get("department") or "").lower() == department_filter.lower()
        }
        papers = [p for p in papers if p.get("course_id") in dept_course_ids]
    academic_year = _normalise_year(request.args.get("academic_year", ""))

    result = []
    for lec in lecturers:
        assigned = [p for p in papers if p.get("lecturer_id") == lec["_id"]]
        assigned_paper_ids = [p["_id"] for p in assigned]
        assigned_course_ids = list({p.get("course_id") for p in assigned if p.get("course_id")})
        assigned_papers = [f"{p.get('name', '')} ({p.get('code', '')})" for p in assigned]
        assigned_semesters = list({str(_to_int(p.get("semester"), 0)) for p in assigned if _to_int(p.get("semester"), 0) > 0})

        years = []
        for p in assigned:
            p_year = _normalise_year(p.get("academic_year", ""))
            if not p_year:
                p_year = _normalise_year((course_map.get(p.get("course_id")) or {}).get("year", ""))
            if p_year:
                years.append(p_year)

        if course_id and course_id not in assigned_course_ids:
            continue
        if semester and semester not in assigned_semesters:
            continue
        if paper_id and paper_id not in assigned_paper_ids:
            continue
        if academic_year and academic_year not in years:
            continue
        if q and not (
            q in _as_text(lec.get("name")).lower()
            or q in _as_text(lec.get("email")).lower()
            or any(q in _as_text(paper).lower() for paper in assigned_papers)
        ):
            continue

        lec["paper_count"] = len(assigned)
        lec["assigned_paper_ids"] = assigned_paper_ids
        lec["assigned_course_ids"] = assigned_course_ids
        lec["assigned_papers"] = assigned_papers
        lec["assigned_semesters"] = sorted(assigned_semesters, key=lambda v: int(v))
        lec["academic_years"] = sorted(list(set(years)))
        result.append(lec)

    return _paginate_items(sanitise_many(result))


@admin_bp.route("/lecturers", methods=["POST"])
@role_required("department_admin")
def add_lecturer(user):
    d = request.get_json(silent=True) or {}
    initial_password = str(d.get("initial_password", "")).strip()
    if not initial_password:
        return jsonify({"error": "initial_password is required and must be delivered out-of-band."}), 400

    is_strong, msg = validate_password_strength(initial_password)
    if not is_strong:
        return jsonify({"error": msg}), 400

    lec = create_user(d["name"], d["email"], initial_password,
                      "lecturer", d.get("department", ""),
                      must_change_password=True,
                      department_id=_user_dept_id(user))
    log_action(
        "CREATE_LECTURER",
        str(user["_id"]),
        target_user=lec["_id"],
        rollback=_rb_delete("auth", "users", {"_id": lec.get("_id")}),
    )
    _clear_query_cache()
    lec_clean = sanitise_mongo_doc(lec)

    email_delivery_enabled = is_email_delivery_enabled()
    temp_pass_display_enabled = _temp_pass_display_enabled()
    if email_delivery_enabled:
        send_welcome_email(
            to_email=d["email"],
            name=d["name"],
            temp_password=initial_password,
            role="lecturer",
        )
    message = (
        "Lecturer created. Credentials have been emailed."
        if email_delivery_enabled
        else "Lecturer created. Email delivery is not configured; share the initial password securely."
    )

    payload = {
        **lec_clean,
        "message": message,
        "email_delivery_enabled": email_delivery_enabled,
        "temp_pass_display_enabled": temp_pass_display_enabled,
    }
    if temp_pass_display_enabled:
        payload["temp_password"] = initial_password
    return jsonify(payload), 201


@admin_bp.route("/lecturers/<lid>", methods=["PUT"])
@role_required("department_admin")
@validate_ids("lid")
def edit_lecturer(user, lid):
    d = request.get_json(silent=True) or {}
    previous = find_user_by_id(lid)
    updated = update_user(lid, d)
    log_action(
        "UPDATE_LECTURER",
        str(user["_id"]),
        target_user=lid,
        rollback=_rb_replace("auth", "users", {"_id": lid}, previous) if previous else None,
    )
    _clear_query_cache()
    return jsonify(sanitise_mongo_doc(updated))


@admin_bp.route("/lecturers/<lid>", methods=["DELETE"])
@role_required("department_admin")
@validate_ids("lid")
def remove_lecturer(user, lid):
    previous = find_user_by_id(lid)
    delete_user(lid)
    log_action(
        "DELETE_LECTURER",
        str(user["_id"]),
        target_user=lid,
        rollback=_rb_restore("auth", "users", previous) if previous else None,
    )
    _clear_query_cache()
    return jsonify({"message": "Deleted"}), 200


@admin_bp.route("/lecturers/<lid>/reset-password", methods=["POST"])
@role_required("department_admin")
def reset_lecturer_password(user, lid):
    d = request.get_json(silent=True) or {}
    temp_password = reset_user_password(lid, temp_password=str(d.get("temp_password", "")).strip() or None)
    log_action("RESET_PASSWORD", str(user["_id"]), target_user=lid,
               details="Lecturer password reset")

    email_delivery_enabled = is_email_delivery_enabled()
    temp_pass_display_enabled = _temp_pass_display_enabled()

    # Send reset email
    lec_user = find_user_by_id(lid)
    if email_delivery_enabled and lec_user and lec_user.get("email"):
        send_password_reset_email(
            to_email=lec_user["email"],
            name=lec_user.get("name", "Lecturer"),
            temp_password=temp_password,
            role="lecturer",
        )
    message = (
        "Lecturer password reset. New credentials have been emailed."
        if email_delivery_enabled
        else "Lecturer password reset. Email delivery is not configured; share the new password securely."
    )

    payload = {
        "message": message,
        "email_delivery_enabled": email_delivery_enabled,
        "temp_pass_display_enabled": temp_pass_display_enabled,
    }
    if temp_pass_display_enabled:
        payload["temp_password"] = temp_password
    return jsonify(payload)


@admin_bp.route("/lecturers/<lid>/reset-pin", methods=["POST"])
@role_required("department_admin")
@validate_ids("lid")
def reset_lecturer_pin(user, lid):
    new_pin = f"{secrets.randbelow(10000):04d}"
    set_user_pin(lid, new_pin)
    log_action("RESET_LECTURER_PIN", str(user["_id"]), target_user=lid,
               details="Admin reset lecturer PIN")
    return jsonify({"pin": new_pin, "message": "Lecturer PIN reset"})


@admin_bp.route("/lecturers/<lid>/pin", methods=["PUT"])
@role_required("department_admin")
@validate_ids("lid")
def update_lecturer_pin(user, lid):
    return jsonify({"error": "Admins cannot set lecturer PIN. Lecturer must manage PIN from dashboard."}), 403


# ─── Students ───────────────────────────────────────────────────────────────

@admin_bp.route("/students", methods=["GET"])
@role_required("super_admin", "department_admin")

def list_students(user):
    page = max(1, _to_int(request.args.get("page", 1), 1))
    per_page = max(1, min(_to_int(request.args.get("per_page", 20), 20), 100))
    skip = (page - 1) * per_page

    # Scope courses to department for department admins
    dept_id = None
    dept_filter_id = _as_text(request.args.get("department_id", ""))  # from super admin filter
    if is_super_admin(user):
        dept_id = dept_filter_id or None
    else:
        dept_id = _user_dept_id(user)

    courses = sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year", "department_id"], department_id=dept_id))
    papers = sanitise_many(get_all_papers(["name", "code", "semester", "course_id", "lecturer_id"]))
    course_map = {c.get("_id"): c for c in courses}
    paper_map = {p.get("_id"): p for p in papers}

    # Set of course IDs visible to this user (enforces dept isolation)
    visible_course_ids = set(course_map.keys())

    department_filter = _as_text(request.args.get("department", ""))
    q = _as_text(request.args.get("q", "")).lower()
    course_id = _as_text(request.args.get("course_id", ""))
    paper_id = _as_text(request.args.get("paper_id", ""))
    academic_session = _as_text(request.args.get("academic_session", "")) or _normalise_year(request.args.get("academic_year", ""))
    semester = _as_text(request.args.get("semester", ""))
    include_inactive = _to_bool(request.args.get("include_inactive", False))

    student_profiles = get_collection("academic", "student_profiles")
    users_col = get_collection("auth", "users")

    filters = []

    # Always restrict to courses visible to this user (dept admin isolation)
    filters.append({"course_id": {"$in": list(visible_course_ids)}})

    # Filter by department name (super admin picks a specific dept from the dropdown)
    if department_filter:
        dept_course_ids = [
            _as_text(c.get("_id"))
            for c in courses
            if _as_text(c.get("department") or "").lower() == department_filter.lower()
        ]
        if dept_course_ids:
            filters.append({"course_id": {"$in": dept_course_ids}})
        else:
            filters.append({"course_id": "never_match"})
    if course_id:
        filters.append({"course_id": course_id})
    if paper_id:
        filters.append({"enrolled_papers": paper_id})
    if academic_session:
        filters.append(
            {
                "$or": [
                    {"academic_session": academic_session},
                    {"academic_year": academic_session},
                    {"year": academic_session},
                ]
            }
        )

    if semester:
        semester_int = _to_int(semester, 0)
        if semester_int > 0:
            semester_paper_ids = [p.get("_id") for p in papers if _to_int(p.get("semester"), 0) == semester_int]
            semester_or = [{"current_semester": semester_int}]
            if semester_paper_ids:
                semester_or.append({"enrolled_papers": {"$in": semester_paper_ids}})
            filters.append({"$or": semester_or})

    if not include_inactive:
        active_course_ids = {
            _as_text(c.get("_id"))
            for c in courses
            if _as_text(c.get("status") or "active").lower() == "active"
        }
        filters.append({"course_id": {"$in": list(active_course_ids)}})

    if q:
        regex = {"$regex": re.escape(q), "$options": "i"}
        matching_user_ids = {
            _as_text(row.get("_id"))
            for row in users_col.find({"$or": [{"name": regex}, {"email": regex}]}, {"_id": 1})
        }
        q_filters = [{"reg_number": regex}, {"roll_number": regex}]
        if matching_user_ids:
            q_filters.append({"user_id": {"$in": list(matching_user_ids)}})
        filters.append({"$or": q_filters})

    query = {"$and": filters} if filters else {}
    projection = {
        "user_id": 1,
        "course_id": 1,
        "enrolled_papers": 1,
        "reg_number": 1,
        "roll_number": 1,
        "academic_session": 1,
        "academic_year": 1,
        "year": 1,
        "enrollment_year": 1,
        "current_semester": 1,
        "face_embeddings": 1,
        "created_at": 1,
    }

    paginated = _get_paginated_data(
        student_profiles,
        query,
        page=page,
        per_page=per_page,
        sort=[("created_at", -1)],
        project=projection
    )
    profiles = paginated["data"]
    total = paginated["total"]
    total_pages = paginated["total_pages"]
    user_map = get_users_by_ids(p.get("user_id") for p in profiles)

    result = []
    for p in profiles:
        u = user_map.get(_as_text(p.get("user_id", "")))
        course = course_map.get(_as_text(p.get("course_id", "")))
        enrolled_papers = p.get("enrolled_papers", [])

        item = sanitise_mongo_doc(p)
        if u:
            item["name"] = u["name"]
            item["email"] = u["email"]

        item["reg_number"] = item.get("reg_number") or item.get("roll_number") or item.get("reg_number")
        enrollment_year = item.get("enrollment_year") or (p.get("created_at") or datetime.now(timezone.utc)).year
        duration_years = _to_int((course or {}).get("course_duration"), 1)
        item["academic_session"] = (
            _as_text(item.get("academic_session"))
            or _as_text(item.get("academic_year"))
            or _derive_academic_session(enrollment_year, duration_years)
        )
        item["academic_year"] = item.get("academic_session")
        item["year"] = item.get("academic_session")
        item["enrollment_year"] = enrollment_year
        item["mobile_no"] = (u or {}).get("mobile_no", "")
        item["course_name"] = (course or {}).get("name")
        item["course_code"] = (course or {}).get("code")
        item["course_status"] = _as_text((course or {}).get("status") or "active").lower() or "active"
        item["is_course_inactive"] = item["course_status"] != "active"
        item["course_department"] = (course or {}).get("department")
        item["course_duration"] = (course or {}).get("course_duration")
        item["current_semester"] = _to_int(item.get("current_semester"), 0) or None
        item["has_face"] = bool(item.get("face_embeddings"))
        item["enrolled_papers"] = [
            {
                "paper_id": pid,
                "paper_name": (paper_map.get(pid) or {}).get("name", "Unknown"),
                "paper_code": (paper_map.get(pid) or {}).get("code", ""),
            }
            for pid in enrolled_papers
        ]

        # Don't send raw embeddings to the frontend
        item.pop("face_embeddings", None)
        result.append(item)

    return jsonify(
        {
            "items": sanitise_many(result),
            "total": total,
            "page": page,
            "per_page": per_page,
        }
    )


@admin_bp.route("/students/options", methods=["GET"])
@role_required("super_admin", "department_admin")
def student_options(user):
    """Return a lightweight student list for select inputs and lookups."""
    course_id = _as_text(request.args.get("course_id", ""))
    department_filter = _as_text(request.args.get("department", ""))
    department_id_filter = _as_text(request.args.get("department_id", ""))
    academic_session = _as_text(request.args.get("academic_session", "")) or _normalise_year(request.args.get("academic_year", ""))
    semester = _as_text(request.args.get("semester", ""))
    q = _as_text(request.args.get("q", "")).lower()
    limit = max(1, min(_to_int(request.args.get("limit", 200), 200), 500))
    include_inactive = _to_bool(request.args.get("include_inactive", False))

    # Build dept scope for isolation
    if is_super_admin(user):
        dept_scope_id = department_id_filter or None
    else:
        dept_scope_id = _user_dept_id(user)

    scoped_courses = sanitise_many(get_all_courses(["department"], department_id=dept_scope_id))
    visible_course_ids = [_as_text(c.get("_id")) for c in scoped_courses]

    profiles_col = get_collection("academic", "student_profiles")
    query = {}

    # Always restrict to visible courses (enforces dept isolation)
    query["course_id"] = {"$in": visible_course_ids}
    
    if department_filter:
        dept_course_ids = [
            _as_text(c.get("_id"))
            for c in scoped_courses
            if _as_text(c.get("department") or "").lower() == department_filter.lower()
        ]
        query["course_id"] = {"$in": dept_course_ids} if dept_course_ids else "never_match"
        
    if course_id:
        if "course_id" in query and isinstance(query["course_id"], dict) and "$in" in query["course_id"]:
            if course_id in query["course_id"]["$in"]:
                query["course_id"] = course_id
            else:
                query["course_id"] = "never_match"
        else:
            query["course_id"] = course_id
    if academic_session:
        query["$or"] = [
            {"academic_session": academic_session},
            {"academic_year": academic_session},
            {"year": academic_session},
        ]
    semester_int = _to_int(semester, 0)
    if semester_int > 0:
        query["current_semester"] = semester_int

    cursor = profiles_col.find(
        query,
        {
            "user_id": 1,
            "course_id": 1,
            "academic_session": 1,
            "academic_year": 1,
            "year": 1,
            "current_semester": 1,
            "reg_number": 1,
            "roll_number": 1,
            "created_at": 1,
        },
    )
    # When searching by q (name/email/reg no), apply limit only after matching,
    # otherwise valid students outside the first N profiles are never considered.
    profiles = list(cursor if q else cursor.limit(limit))

    course_map = {}
    if profiles:
        course_ids = {str(p.get("course_id", "")) for p in profiles if p.get("course_id")}
        if course_ids:
            course_map = {c["_id"]: c for c in sanitise_many(get_all_courses(["name", "code", "status", "course_duration"])) if c.get("_id") in course_ids}

    user_map = get_users_by_ids(profile.get("user_id") for profile in profiles)
    result = []

    for profile in profiles:
        uid = _as_text(profile.get("user_id", ""))
        student = user_map.get(uid)
        if not student:
            continue

        course = course_map.get(_as_text(profile.get("course_id", "")))
        is_course_inactive = _as_text((course or {}).get("status") or "active").lower() != "active"
        if is_course_inactive and not include_inactive:
            continue
        enrollment_year = _to_int((profile.get("created_at") or datetime.now(timezone.utc)).year if hasattr(profile.get("created_at"), "year") else None, 0)
        duration_years = _to_int((course or {}).get("course_duration"), 1)
        resolved_session = (
            _as_text(profile.get("academic_session"))
            or _as_text(profile.get("academic_year"))
            or _as_text(profile.get("year"))
            or _derive_academic_session(enrollment_year or datetime.now(timezone.utc).year, duration_years)
        )
        current_semester = _to_int(profile.get("current_semester"), 0)

        reg_number = _as_text(profile.get("reg_number") or profile.get("roll_number"))
        name = _as_text(student.get("name"))
        email = _as_text(student.get("email"))
        if q and not (q in name.lower() or q in email.lower() or q in reg_number.lower()):
            continue

        result.append({
            "_id": uid,
            "user_id": uid,
            "name": student.get("name"),
            "email": student.get("email"),
            "reg_number": reg_number,
            "roll_number": _as_text(profile.get("roll_number")),
            "academic_session": resolved_session,
            "course_id": _as_text(profile.get("course_id", "")),
            "current_semester": current_semester or None,
            "course_name": (course or {}).get("name"),
            "is_course_inactive": is_course_inactive,
        })

        if q and len(result) >= limit:
            break

    return jsonify(sanitise_many(result))


@admin_bp.route("/students", methods=["POST"])
@role_required("super_admin", "department_admin")
def add_student(user):
    d = request.get_json(silent=True) or {}
    required_fields = ["name", "email", "course_id"]
    missing = [field for field in required_fields if not _as_text(d.get(field))]
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    # Check if email already exists
    existing_user = find_user_by_email(d["email"])
    if existing_user:
        return jsonify({"error": "Email already in use. Please use a different email."}), 409

    course_id = _as_text(d.get("course_id", ""))
    course = _safe_get_course(course_id) if course_id else None
    if not course:
        return jsonify({"error": "Course not found or invalid."}), 404
    if _course_is_inactive(course):
        return jsonify({"error": "Cannot create student under inactive course"}), 409
    
    enrollment_year = _to_int(d.get("enrollment_year"), datetime.now(timezone.utc).year)
    course_duration = _to_int((course or {}).get("course_duration"), 1)
    academic_session = _derive_academic_session(enrollment_year, course_duration)

    try:
        initial_password = str(d.get("initial_password", "")).strip()
        if not initial_password:
            return jsonify({"error": "initial_password is required and must be delivered out-of-band."}), 400

        is_strong, msg = validate_password_strength(initial_password)
        if not is_strong:
            return jsonify({"error": msg}), 400

        stu = create_user(
            d["name"],
            d["email"],
            initial_password,
            "student",
            d.get("department", (course or {}).get("department", "")),
            must_change_password=True,
            department_id=_user_dept_id(user) or (course or {}).get("department_id"),
        )
    except Exception as exc:
        current_app.logger.exception("User creation failed")
        return jsonify({"error": f"Failed to create user: {str(exc)}"}), 500

    mobile_no = _as_text(d.get("mobile_no", ""))
    if mobile_no:
        try:
            update_user(str(stu["_id"]), {"mobile_no": mobile_no})
        except Exception as exc:
            current_app.logger.exception("Failed to update mobile number")
            delete_user(str(stu["_id"]))
            return jsonify({"error": f"Failed to update user: {str(exc)}"}), 500

    profile = None
    for attempt in range(5):
        reg_number = d.get("reg_number") or _generate_registration_number(course, academic_session)
        try:
            profile = create_student_profile(str(stu["_id"]), reg_number, course_id, academic_session)
            break
        except DuplicateKeyError:
            if attempt == 4:  # Last attempt
                break
            continue
        except Exception as exc:
            current_app.logger.exception(f"Profile creation attempt {attempt + 1} failed")
            continue

    if not profile:
        delete_user(str(stu["_id"]))
        return jsonify({"error": "Could not generate a unique registration number. Please try again."}), 409

    try:
        update_profile(
            str(stu["_id"]),
            {
                "enrollment_year": enrollment_year,
                "current_semester": 1,
                "academic_session": academic_session,
                "academic_year": academic_session,
                "year": academic_session,
            },
        )
    except Exception as exc:
        current_app.logger.exception("Failed to update student profile")
        delete_user(str(stu["_id"]))
        delete_profile(str(stu["_id"]), user=stu)
        return jsonify({"error": f"Failed to update profile: {str(exc)}"}), 500

    profile = get_profile_by_user(str(stu["_id"]))

    if d.get("enrolled_papers"):
        enroll_in_papers(str(stu["_id"]), d["enrolled_papers"])

    log_action(
        "CREATE_STUDENT",
        str(user["_id"]),
        target_user=stu["_id"],
        rollback=_rb_batch([
            _rb_delete("academic", "student_profiles", {"user_id": str(stu["_id"])}),
            _rb_delete("auth", "users", {"_id": str(stu["_id"])}),
        ]),
    )
    _clear_query_cache()
    
    # Sanitize before returning to ensure ObjectId is serializable
    stu_clean = sanitise_mongo_doc(stu)
    profile_clean = sanitise_mongo_doc(profile) if profile else None
    
    email_delivery_enabled = is_email_delivery_enabled()
    temp_pass_display_enabled = _temp_pass_display_enabled()
    if email_delivery_enabled:
        send_welcome_email(
            to_email=d["email"],
            name=d["name"],
            temp_password=initial_password,
            role="student",
        )
    message = (
        "Student created. Credentials have been emailed."
        if email_delivery_enabled
        else "Student created. Email delivery is not configured; share the initial password securely."
    )

    payload = {
        **stu_clean,
        "profile": profile_clean,
        "message": message,
        "email_delivery_enabled": email_delivery_enabled,
        "temp_pass_display_enabled": temp_pass_display_enabled,
    }
    if temp_pass_display_enabled:
        payload["temp_password"] = initial_password
    return jsonify(payload), 201


@admin_bp.route("/students/<sid>", methods=["PUT"])
@role_required("super_admin", "department_admin")
@validate_ids("sid")
def edit_student(user, sid):
    d = request.get_json(silent=True) or {}
    user_id, profile = _resolve_user_identity(sid)
    if not user_id:
        return jsonify({"error": "Student not found"}), 404

    _, student_lock_error = _ensure_student_course_active(user_id)
    if student_lock_error:
        return student_lock_error

    prev_user = find_user_by_id(user_id)
    prev_profile = get_profile_by_user(user_id)

    user_fields = {}
    profile_fields = {}
    for k in ["name", "email", "department", "mobile_no"]:
        if k in d:
            user_fields[k] = d[k]
    for k in ["roll_number", "reg_number", "course_id", "enrolled_papers", "academic_year", "year", "academic_session", "enrollment_year", "current_semester"]:
        if k in d:
            profile_fields[k] = d[k]
    if "year" in profile_fields and "academic_year" not in profile_fields:
        profile_fields["academic_year"] = profile_fields["year"]
    if "academic_session" in profile_fields and "academic_year" not in profile_fields:
        profile_fields["academic_year"] = _as_text(profile_fields["academic_session"])

    if "academic_year" in profile_fields:
        profile_fields["academic_year"] = _as_text(profile_fields["academic_year"])
        profile_fields["year"] = profile_fields["academic_year"]
        profile_fields["academic_session"] = profile_fields["academic_year"]

    if "roll_number" in profile_fields:
        profile_fields["roll_number"] = _as_text(profile_fields.get("roll_number"))
    if "reg_number" in profile_fields:
        profile_fields["reg_number"] = _as_text(profile_fields.get("reg_number"))
    if "roll_number" in d and not profile_fields.get("roll_number"):
        return jsonify({"error": "Roll number cannot be empty"}), 400
    if "reg_number" in d and not profile_fields.get("reg_number"):
        return jsonify({"error": "Registration number cannot be empty"}), 400

    current_course_id = (profile or {}).get("course_id")
    next_course_id = profile_fields.get("course_id", current_course_id)
    current_enrollment_year = _to_int((profile or {}).get("enrollment_year"), (profile or {}).get("created_at", datetime.now(timezone.utc)).year)
    next_enrollment_year = _to_int(profile_fields.get("enrollment_year"), current_enrollment_year)
    next_course = _safe_get_course(next_course_id) if next_course_id else None
    if next_course_id and _course_is_inactive(next_course):
        return jsonify({"error": "Cannot move student to an inactive course"}), 409
    next_course_duration = _to_int((next_course or {}).get("course_duration"), 1)
    next_session = profile_fields.get("academic_session") or _derive_academic_session(next_enrollment_year, next_course_duration)
    profile_fields["enrollment_year"] = next_enrollment_year
    if "current_semester" in profile_fields:
        profile_fields["current_semester"] = _to_int(profile_fields.get("current_semester"), 0) or None
    profile_fields["academic_session"] = next_session
    profile_fields["academic_year"] = next_session
    profile_fields["year"] = next_session

    current_session = _as_text((profile or {}).get("academic_session") or (profile or {}).get("academic_year") or (profile or {}).get("year"))
    requested_reg_update = "reg_number" in d
    requested_roll_update = "roll_number" in d
    course_changed = "course_id" in d and _as_text(current_course_id) != _as_text(next_course_id)
    enrollment_changed = "enrollment_year" in d and next_enrollment_year != current_enrollment_year
    session_changed = any(k in d for k in ["academic_session", "academic_year", "year"]) and _as_text(next_session) != current_session

    if (course_changed or enrollment_changed or session_changed) and not (requested_reg_update or requested_roll_update):
        new_reg = _generate_registration_number(next_course, next_session, exclude_user_id=user_id)
        profile_fields["reg_number"] = new_reg
        profile_fields["roll_number"] = new_reg

    if "reg_number" in profile_fields and "roll_number" not in profile_fields:
        profile_fields["roll_number"] = profile_fields["reg_number"]
    if "roll_number" in profile_fields and "reg_number" not in profile_fields:
        profile_fields["reg_number"] = profile_fields["roll_number"]

    if user_fields:
        update_user(user_id, user_fields)
    if profile_fields:
        try:
            update_profile(user_id, profile_fields)
        except DuplicateKeyError:
            return jsonify({"error": "Registration number already exists. Please use a unique value."}), 409
    rollback_ops = []
    if prev_user:
        rollback_ops.append(_rb_replace("auth", "users", {"_id": user_id}, prev_user))
    if prev_profile:
        rollback_ops.append(_rb_replace("academic", "student_profiles", {"user_id": user_id}, prev_profile))

    log_action(
        "UPDATE_STUDENT",
        str(user["_id"]),
        target_user=user_id,
        rollback=_rb_batch(rollback_ops) if rollback_ops else None,
    )
    _clear_query_cache()
    return jsonify({"message": "Updated"})


@admin_bp.route("/students/<sid>", methods=["DELETE"])
@role_required("super_admin", "department_admin")
@validate_ids("sid")
def remove_student(user, sid):
    user_id, _ = _resolve_user_identity(sid)
    if not user_id:
        return jsonify({"error": "Student not found"}), 404

    _, student_lock_error = _ensure_student_course_active(user_id)
    if student_lock_error:
        return student_lock_error

    prev_user = find_user_by_id(user_id)
    prev_profile = get_profile_by_user(user_id)

    attendance_logs = get_collection("attendance", "attendance_logs")
    attendance_logs.delete_many({"user_id": {"$in": _id_variants(user_id)}})

    delete_user(user_id)
    delete_profile(user_id, user=prev_user)
    rollback_ops = []
    if prev_user:
        rollback_ops.append(_rb_restore("auth", "users", prev_user))
    if prev_profile:
        rollback_ops.append(_rb_restore("academic", "student_profiles", prev_profile))

    log_action(
        "DELETE_STUDENT",
        str(user["_id"]),
        target_user=user_id,
        rollback=_rb_batch(rollback_ops) if rollback_ops else None,
    )
    _clear_query_cache()
    return jsonify({"message": "Deleted"}), 200


@admin_bp.route("/students/bulk-promote", methods=["POST"])
@admin_bp.route("/student-bulk-promote", methods=["POST"])
@role_required("department_admin")
def bulk_promote_students(user):
    """Promote selected students to the next semester or an optional target semester."""
    d = request.get_json(silent=True) or {}
    raw_ids = d.get("user_ids") or []
    from_semester = _to_int(d.get("from_semester"), 0)
    target_semester = _to_int(d.get("target_semester"), 0)

    if d.get("target_semester") is not None and target_semester <= 0:
        return jsonify({"error": "target_semester must be a positive integer"}), 400

    user_ids = [sid for sid in raw_ids if _as_text(sid)]
    if not user_ids:
        return jsonify({"error": "user_ids is required"}), 400

    paper_map = {p.get("_id"): p for p in sanitise_many(get_all_papers(["name", "code", "semester", "course_id", "lecturer_id"]))}
    course_map = {c.get("_id"): c for c in sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year"]))}

    promoted = 0
    skipped = 0
    skipped_max_semester = 0
    skipped_target_semester = 0
    removed_papers = 0
    rollback_ops = []
    for sid in user_ids:
        user_id, profile = _resolve_user_identity(_as_text(sid))
        if not user_id or not profile:
            skipped += 1
            continue

        course_for_lock = _safe_get_course((profile or {}).get("course_id"))
        if _course_is_inactive(course_for_lock):
            skipped += 1
            continue

        current_sem = _to_int((profile or {}).get("current_semester"), 0)
        if current_sem <= 0:
            current_sem = from_semester if from_semester > 0 else 1

        course_id = _as_text((profile or {}).get("course_id"))
        course = course_map.get(course_id) or {}
        max_semester = max(1, _to_int(course.get("course_duration"), 1) * 2)

        if target_semester > 0:
            if target_semester > max_semester:
                skipped_max_semester += 1
                continue
            # Temporary rule: allow selecting semester 1 to force reset/demotion to first semester.
            if target_semester == 1:
                next_sem = 1
            elif target_semester <= current_sem:
                skipped_target_semester += 1
                continue
            else:
                next_sem = target_semester
        else:
            if current_sem >= max_semester:
                skipped_max_semester += 1
                continue
            next_sem = current_sem + 1

        enrolled_papers = list((profile or {}).get("enrolled_papers") or [])
        kept_papers = []
        for pid in enrolled_papers:
            pdoc = paper_map.get(pid) or {}
            psem = _to_int(pdoc.get("semester"), 0)
            # Keep unknown-semester papers and papers in/after the promoted semester.
            if psem == 0 or psem >= next_sem:
                kept_papers.append(pid)

        removed_papers += max(0, len(enrolled_papers) - len(kept_papers))
        rollback_ops.append(_rb_replace("academic", "student_profiles", {"user_id": user_id}, profile))
        update_profile(user_id, {"current_semester": next_sem, "enrolled_papers": kept_papers})
        promoted += 1

    log_action(
        "BULK_PROMOTE_STUDENTS",
        str(user["_id"]),
        details=(
            f"Promoted {promoted}, skipped {skipped}, skipped_max={skipped_max_semester}, "
            f"skipped_target={skipped_target_semester}, removed_papers={removed_papers}, "
            f"from_semester={from_semester or 'auto'}, target_semester={target_semester or 'auto'}"
        ),
        rollback=_rb_batch(rollback_ops) if rollback_ops else None,
    )
    _clear_query_cache()

    return jsonify(
        {
            "message": f"Promoted {promoted} students, removed {removed_papers} old-semester paper assignments, skipped {skipped_max_semester} already at max semester",
            "promoted_count": promoted,
            "skipped_count": skipped,
            "skipped_max_semester_count": skipped_max_semester,
            "skipped_target_semester_count": skipped_target_semester,
            "removed_papers_count": removed_papers,
            "target_semester": target_semester or None,
        }
    )


# ─── Excel Import ────────────────────────────────────────────────────────────

@admin_bp.route("/students/import-excel", methods=["POST"])
@role_required("department_admin")
def import_students_excel(user):
    """Bulk-import students from an uploaded Excel file.

    Expects multipart/form-data with:
      - file    : .xlsx file
      - course_id : required
      - semester  : required (integer)

    Excel columns (case-insensitive, stripped):
      Name, RollNo / Roll No, RegdNo / Regd No / Regd. No., Email, PhoneNo / Phone No (optional)
    """
    course_id = _as_text(request.form.get("course_id"))
    semester = _to_int(request.form.get("semester"), 0)

    if not course_id:
        return jsonify({"error": "course_id is required"}), 400
    if semester <= 0:
        return jsonify({"error": "semester is required and must be a positive integer"}), 400

    course, err = _get_active_course_or_error(course_id)
    if err:
        return err

    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "No file uploaded"}), 400
    if not uploaded.filename.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(uploaded.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return jsonify({"error": f"Could not parse Excel file: {str(exc)}"}), 400

    if not rows:
        return jsonify({"error": "Excel file is empty"}), 400

    # Normalise header row
    header_raw = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    COL_ALIASES = {
        "name": ["name"],
        "roll_number": ["rollno", "roll no", "roll_no", "roll number"],
        "reg_number": ["regdno", "regd no", "regd. no.", "regd.no.", "reg no", "reg_no", "reg number", "regno"],
        "email": ["email"],
        "mobile_no": ["phoneno", "phone no", "phone_no", "phone number", "mobile", "mobile_no", "mobileno"],
    }

    col_idx = {}
    for field, aliases in COL_ALIASES.items():
        for idx, h in enumerate(header_raw):
            if h in aliases:
                col_idx[field] = idx
                break

    required_cols = ["name", "email", "reg_number"]
    missing_cols = [c for c in required_cols if c not in col_idx]
    if missing_cols:
        return jsonify({"error": f"Missing required columns: {', '.join(missing_cols)}. Found headers: {header_raw}"}), 400

    enrollment_year = datetime.now(timezone.utc).year
    course_duration = _to_int((course or {}).get("course_duration"), 1)
    academic_session = _derive_academic_session(enrollment_year, course_duration)

    results = []
    created_count = 0
    skipped_count = 0
    error_count = 0
    temp_pass_display_enabled = _temp_pass_display_enabled()

    for row_num, row in enumerate(rows[1:], start=2):
        def _cell(field):
            idx = col_idx.get(field)
            if idx is None:
                return ""
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else ""

        name = _cell("name")
        email = _cell("email")
        reg_number = _cell("reg_number")
        roll_number = _cell("roll_number") or reg_number
        mobile_no = _cell("mobile_no")

        if not name or not email or not reg_number:
            skipped_count += 1
            results.append({"row": row_num, "status": "skipped", "reason": "Missing required field (Name, Email, or RegdNo)"})
            continue

        if find_user_by_email(email):
            skipped_count += 1
            results.append({"row": row_num, "name": name, "email": email, "status": "skipped", "reason": "Email already exists"})
            continue

        try:
            initial_password = _generate_import_temp_password()

            stu = create_user(
                name,
                email,
                initial_password,
                "student",
                (course or {}).get("department", ""),
                must_change_password=True,
            )

            if mobile_no:
                try:
                    update_user(str(stu["_id"]), {"mobile_no": mobile_no})
                except Exception:
                    pass  # nosec B110

            profile = None
            for attempt in range(3):
                try:
                    use_reg = reg_number if attempt == 0 else _generate_registration_number(course, academic_session)
                    profile = create_student_profile(str(stu["_id"]), use_reg, course_id, academic_session)
                    break
                except DuplicateKeyError:
                    continue
                except Exception:
                    break

            if not profile:
                delete_user(str(stu["_id"]))
                error_count += 1
                results.append({"row": row_num, "name": name, "email": email, "status": "error", "reason": "Could not create profile (duplicate reg number?)"})
                continue

            update_profile(str(stu["_id"]), {
                "enrollment_year": enrollment_year,
                "current_semester": semester,
                "roll_number": roll_number,
                "reg_number": reg_number,
                "academic_session": academic_session,
                "academic_year": academic_session,
                "year": academic_session,
            })

            log_action(
                "CREATE_STUDENT",
                str(user["_id"]),
                target_user=stu["_id"],
                rollback=_rb_batch([
                    _rb_delete("academic", "student_profiles", {"user_id": str(stu["_id"])}),
                    _rb_delete("auth", "users", {"_id": str(stu["_id"])}),
                ]),
            )
            created_count += 1
            row_result = {"row": row_num, "name": name, "email": email, "status": "created"}
            if temp_pass_display_enabled:
                row_result["temp_password"] = initial_password
            results.append(row_result)

            # Send welcome email (fire-and-forget)
            send_welcome_email(
                to_email=email,
                name=name,
                temp_password=initial_password,
                role="student",
            )
        except DuplicateKeyError:
            skipped_count += 1
            results.append({"row": row_num, "name": name, "email": email, "status": "skipped", "reason": "Duplicate email or registration number"})
        except Exception as exc:
            error_count += 1
            results.append({"row": row_num, "name": name, "email": email, "status": "error", "reason": str(exc)})

    _clear_query_cache()
    return jsonify({
        "message": f"Import complete: {created_count} created, {skipped_count} skipped, {error_count} errors",
        "created": created_count,
        "skipped": skipped_count,
        "errors": error_count,
        "email_delivery_enabled": is_email_delivery_enabled(),
        "temp_pass_display_enabled": temp_pass_display_enabled,
        "results": results,
    }), 207 if (skipped_count + error_count) > 0 else 201


@admin_bp.route("/lecturers/import-excel", methods=["POST"])
@role_required("department_admin")
def import_lecturers_excel(user):
    """Bulk-import lecturers from an uploaded Excel file.

    Expects multipart/form-data with:
      - file : .xlsx file

    Excel columns (case-insensitive):
            Department, Name, Email, Courses, Papers
    """
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "No file uploaded"}), 400
    if not uploaded.filename.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(uploaded.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return jsonify({"error": f"Could not parse Excel file: {str(exc)}"}), 400

    if not rows:
        return jsonify({"error": "Excel file is empty"}), 400

    header_raw = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    department_idx = next((i for i, h in enumerate(header_raw) if h in ["department", "dept"]), None)
    name_idx = next((i for i, h in enumerate(header_raw) if h in ["name", "full name", "fullname"]), None)
    email_idx = next((i for i, h in enumerate(header_raw) if h in ["email", "email address", "e-mail"]), None)
    courses_idx = next((i for i, h in enumerate(header_raw) if h in ["courses", "course"]), None)
    papers_idx = next((i for i, h in enumerate(header_raw) if h in ["papers", "paper"]), None)

    if department_idx is None or name_idx is None or email_idx is None:
        return jsonify({"error": f"Missing required columns: Department, Name and/or Email. Found headers: {header_raw}"}), 400

    results = []
    created_count = 0
    skipped_count = 0
    error_count = 0
    temp_pass_display_enabled = _temp_pass_display_enabled()
    departments_col = get_collection("academic", "departments")

    def _cell(row, idx):
        value = row[idx] if idx is not None and idx < len(row) else None
        return str(value).strip() if value is not None else ""

    def _parse_csv_list(raw_value):
        if not raw_value:
            return []
        return [item.strip() for item in str(raw_value).split(",") if item and str(item).strip()]

    def _find_department(raw_department):
        if not raw_department:
            return None
        escaped = re.escape(raw_department.strip())
        return departments_col.find_one({
            "$or": [
                {"name": {"$regex": f"^{escaped}$", "$options": "i"}},
                {"code": {"$regex": f"^{escaped}$", "$options": "i"}},
            ]
        })

    def _resolve_courses(raw_courses):
        resolved = []
        seen_ids = set()
        for course_code in _parse_csv_list(raw_courses):
            course = get_course_by_code(course_code)
            if not course:
                continue
            course_id = str(course.get("_id"))
            if course_id in seen_ids:
                continue
            seen_ids.add(course_id)
            resolved.append(course)
        return resolved

    def _resolve_papers(raw_papers):
        resolved = []
        seen_ids = set()
        for paper_code in _parse_csv_list(raw_papers):
            paper = get_paper_by_code(paper_code)
            if not paper:
                continue
            paper_id = str(paper.get("_id"))
            if paper_id in seen_ids:
                continue
            seen_ids.add(paper_id)
            resolved.append(paper)
        return resolved

    for row_num, row in enumerate(rows[1:], start=2):
        department = _cell(row, department_idx)
        name = _cell(row, name_idx)
        email = _cell(row, email_idx)
        raw_courses = _cell(row, courses_idx)
        raw_papers = _cell(row, papers_idx)

        if not department or not name or not email:
            skipped_count += 1
            results.append({"row": row_num, "status": "skipped", "reason": "Missing Department, Name, or Email"})
            continue

        if find_user_by_email(email):
            skipped_count += 1
            results.append({"row": row_num, "name": name, "email": email, "status": "skipped", "reason": "Email already exists"})
            continue

        try:
            department_doc = _find_department(department)
            department_value = department_doc.get("name") if department_doc else department
            department_id_value = department_doc.get("_id") if department_doc else None
            matched_courses = _resolve_courses(raw_courses)
            matched_papers = _resolve_papers(raw_papers)

            initial_password = _generate_import_temp_password()
            lec = create_user(
                name,
                email,
                initial_password,
                "lecturer",
                department_value,
                must_change_password=True,
                department_id=department_id_value,
            )
            log_action(
                "CREATE_LECTURER",
                str(user["_id"]),
                target_user=lec["_id"],
                rollback=_rb_delete("auth", "users", {"_id": lec.get("_id")}),
            )

            assigned_course_ids = []
            assigned_paper_ids = []
            scoped_papers = get_all_papers(["_id", "course_id"], department_id=department_id_value)
            scoped_paper_map = {}
            for paper in scoped_papers:
                scoped_paper_map.setdefault(str(paper.get("course_id") or ""), []).append(str(paper.get("_id")))

            for course in matched_courses:
                course_id = str(course.get("_id"))
                course_paper_ids = scoped_paper_map.get(course_id, [])
                if course_paper_ids:
                    bulk_assign_lecturer(course_paper_ids, lec["_id"])
                    assigned_course_ids.append(course_id)
                    assigned_paper_ids.extend(course_paper_ids)

            direct_paper_ids = [str(paper.get("_id")) for paper in matched_papers]
            if direct_paper_ids:
                bulk_assign_lecturer(direct_paper_ids, lec["_id"])
                assigned_paper_ids.extend(direct_paper_ids)

            assigned_course_ids = sorted(set(assigned_course_ids))
            assigned_paper_ids = sorted(set(assigned_paper_ids))

            created_count += 1
            row_result = {
                "row": row_num,
                "name": name,
                "email": email,
                "status": "created",
                "department": department_value,
                "matched_courses": [course.get("code") for course in matched_courses],
                "matched_papers": [paper.get("code") for paper in matched_papers],
                "assigned_course_count": len(assigned_course_ids),
                "assigned_paper_count": len(assigned_paper_ids),
            }
            if not department_doc:
                row_result["department_warning"] = "Department not found; stored raw department value"
            if temp_pass_display_enabled:
                row_result["temp_password"] = initial_password
            results.append(row_result)

            send_welcome_email(
                to_email=email,
                name=name,
                temp_password=initial_password,
                role="lecturer",
            )
        except DuplicateKeyError:
            skipped_count += 1
            results.append({"row": row_num, "name": name, "email": email, "status": "skipped", "reason": "Duplicate email"})
        except Exception as exc:
            error_count += 1
            results.append({"row": row_num, "name": name, "email": email, "status": "error", "reason": str(exc)})

    _clear_query_cache()
    return jsonify({
        "message": f"Import complete: {created_count} created, {skipped_count} skipped, {error_count} errors",
        "created": created_count,
        "skipped": skipped_count,
        "errors": error_count,
        "email_delivery_enabled": is_email_delivery_enabled(),
        "temp_pass_display_enabled": temp_pass_display_enabled,
        "results": results,
    }), 207 if (skipped_count + error_count) > 0 else 201


def _generate_import_temp_password(length=14):
    """Generate a cryptographically random temporary password for bulk imports."""
    import string
    upper = string.ascii_uppercase.replace("I", "").replace("O", "")
    lower = string.ascii_lowercase.replace("l", "").replace("o", "")
    digits = "23456789"
    symbols = "!@#$%^&*"
    all_chars = upper + lower + digits + symbols
    chars = [
        secrets.choice(upper),
        secrets.choice(lower),
        secrets.choice(digits),
        secrets.choice(symbols),
    ]
    while len(chars) < length:
        chars.append(secrets.choice(all_chars))
    secrets.SystemRandom().shuffle(chars)
    return "".join(chars)


@admin_bp.route("/students/<sid>/reset-password", methods=["POST"])
@role_required("department_admin")
@validate_ids("sid")
def reset_student_password(user, sid):
    user_id, _ = _resolve_user_identity(sid)
    if not user_id:
        return jsonify({"error": "Student not found"}), 404

    d = request.get_json(silent=True) or {}
    temp_password = reset_user_password(user_id, temp_password=str(d.get("temp_password", "")).strip() or None)
    log_action("RESET_PASSWORD", str(user["_id"]), target_user=user_id,
               details="Student password reset")

    email_delivery_enabled = is_email_delivery_enabled()
    temp_pass_display_enabled = _temp_pass_display_enabled()

    # Send reset email
    stu_user = find_user_by_id(user_id)
    if email_delivery_enabled and stu_user and stu_user.get("email"):
        send_password_reset_email(
            to_email=stu_user["email"],
            name=stu_user.get("name", "Student"),
            temp_password=temp_password,
            role="student",
        )
    message = (
        "Student password reset. New credentials have been emailed."
        if email_delivery_enabled
        else "Student password reset. Email delivery is not configured; share the new password securely."
    )

    payload = {
        "message": message,
        "email_delivery_enabled": email_delivery_enabled,
        "temp_pass_display_enabled": temp_pass_display_enabled,
    }
    if temp_pass_display_enabled:
        payload["temp_password"] = temp_password
    return jsonify(payload)


# ─── Student Enrollment (Photo → Embedding) ────────────────────────────────

@admin_bp.route("/students/enroll", methods=["POST"])
@role_required("department_admin")
def enroll_student_face(user):
    """Accept a student photo, extract FaceNet embedding, and store it."""
    d = request.get_json(silent=True) or {}
    user_id = d.get("user_id")
    photo_b64 = d.get("photo")  # base64 encoded image
    dataset_photos = d.get("dataset_photos") or []

    if not user_id or not photo_b64:
        return jsonify({"error": "user_id and photo are required"}), 400

    resolved_user_id, _ = _resolve_user_identity(user_id)
    if not resolved_user_id:
        return jsonify({"error": "Student not found"}), 404

    try:
        img = decode_base64_image(photo_b64)
    except ValueError:
        # User error: invalid/corrupt image, do not log traceback
        return jsonify({"error": "Invalid image format. Please upload a valid PNG or JPEG photo."}), 400
    except Exception:
        # Unexpected error: log traceback
        current_app.logger.exception("Unexpected error during image decoding")
        return jsonify({"error": "Unexpected error while processing image. Please try again or contact support."}), 500

    try:
        detector = get_detector()
        faces = detector.detect_faces(img)
    except Exception as exc:
        current_app.logger.exception("Face detector failed")
        return jsonify({"error": f"Face detector unavailable: {exc}"}), 500

    if not faces:
        return jsonify({"error": "No face detected in the photo"}), 400

    # Use the first (largest confidence) face
    face_crop = faces[0]["crop"]
    try:
        embedding = normalize_embedding(generate_embedding(face_crop))
        add_face_embedding(resolved_user_id, embedding)
    except Exception as exc:
        current_app.logger.exception("Embedding persistence failed")
        return jsonify({"error": f"Failed to store face embedding: {exc}"}), 500

    dataset_saved_count = 0
    dataset_warning = None
    if isinstance(dataset_photos, list) and dataset_photos:
        try:
            dataset_user_key = _as_text(resolved_user_id)

            dataset_crops = []
            last_valid_crop = face_crop
            for frame_b64 in dataset_photos[:50]:
                if not isinstance(frame_b64, str) or not frame_b64:
                    if last_valid_crop is not None:
                        dataset_crops.append(last_valid_crop)
                    continue
                try:
                    frame_img = decode_base64_image(frame_b64)
                    frame_faces = detector.detect_faces(frame_img)
                    if frame_faces:
                        last_valid_crop = frame_faces[0]["crop"]
                        dataset_crops.append(last_valid_crop)
                    elif last_valid_crop is not None:
                        dataset_crops.append(last_valid_crop)
                except Exception:
                    if last_valid_crop is not None:
                        dataset_crops.append(last_valid_crop)
                    continue

            if not dataset_crops and last_valid_crop is not None:
                dataset_crops.append(last_valid_crop)

            while len(dataset_crops) < 50 and last_valid_crop is not None:
                dataset_crops.append(last_valid_crop)

            if dataset_crops:
                saved_paths = save_cropped_face_dataset(
                    dataset_user_key,
                    dataset_crops,
                    dataset_root="dataset",
                    max_images=50,
                )
                dataset_saved_count = len(saved_paths)
        except Exception as exc:
            current_app.logger.exception("Dataset save failed during face enrollment")
            dataset_warning = f"Dataset save failed: {exc}"

    log_action("ENROLL_FACE", str(user["_id"]), target_user=resolved_user_id,
               details="Face embedding added")
    _clear_query_cache()

    response = {
        "message": "Face enrolled successfully",
        "faces_detected": len(faces),
        "dataset_saved_count": dataset_saved_count,
    }
    if dataset_warning:
        response["dataset_warning"] = dataset_warning

    return jsonify(response), 200


@admin_bp.route("/students/upload-photo", methods=["POST"])
@role_required("department_admin")
def upload_student_photo(user):
    """Upload and store a single student photo in uploads folder."""
    student_name = _as_text(request.form.get("student_name", ""))
    if not student_name:
        return jsonify({"error": "student_name is required"}), 400

    if "image" not in request.files:
        return jsonify({"error": "image file is required"}), 400

    file = request.files["image"]
    if not file or file.filename == "":
        return jsonify({"error": "No image file selected"}), 400

    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "Uploaded image is empty"}), 400

    arr = np.frombuffer(file_bytes, dtype=np.uint8)
    image = cv2.imdecode(arr, cv2.IMREAD_UNCHANGED)
    if image is None:
        return jsonify({"error": "Invalid image file"}), 400

    uploads_dir = current_app.config.get("UPLOAD_FOLDER", "uploads")
    saved_path = save_student_upload(student_name, image, uploads_dir=uploads_dir)

    log_action(
        "UPLOAD_STUDENT_PHOTO",
        str(user["_id"]),
        details=f"Stored photo for {student_name} as {saved_path}",
    )

    return jsonify({
        "message": "Student photo uploaded successfully",
        "file_path": saved_path,
        "file_name": os.path.basename(saved_path),
    }), 201


def _train_single_face_job(actor_id, user_id, job_id=None):
    _raise_if_job_cancelled(job_id)
    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=1,
            processed_faces=0,
            trained_faces=0,
            failed_faces=0,
            stage="training",
            message="Training 1 of 1 face",
        )

    train_result = _train_embeddings_from_dataset_for_user(user_id)
    _raise_if_job_cancelled(job_id)
    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=1,
            processed_faces=1,
            trained_faces=1,
            failed_faces=0,
            stage="saving",
            message="Saving trainer artifact",
        )
    trainer_result = _refresh_face_trainer_artifact()
    _raise_if_job_cancelled(job_id)
    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=1,
            processed_faces=1,
            trained_faces=1,
            failed_faces=0,
            stage="completed",
            message="Training complete",
        )
    log_action(
        "TRAIN_FACE_FROM_DATASET",
        actor_id,
        target_user=user_id,
        details=(
            f"dataset={train_result['dataset_dir']}, "
            f"trained={train_result['trained_embeddings']}, "
            f"skipped={train_result['skipped_images']}, "
            f"trainer={trainer_result.get('model_path') or 'not_saved'}"
        ),
    )
    return {
        "message": "Face training completed",
        "trained_embeddings": train_result["trained_embeddings"],
        "skipped_images": train_result["skipped_images"],
        "dataset_dir": train_result["dataset_dir"],
        "trainer": trainer_result,
    }


def _train_bulk_faces_job(actor_id, user_ids, job_id=None):
    items = []
    total_trained_embeddings = 0
    success_count = 0
    failure_count = 0

    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=len(user_ids),
            processed_faces=0,
            trained_faces=0,
            failed_faces=0,
            stage="training",
            message=f"Training 0 of {len(user_ids)} faces",
        )

    for index, sid in enumerate(user_ids, start=1):
        _raise_if_job_cancelled(job_id)
        user_id, _ = _resolve_user_identity(sid)
        if not user_id:
            items.append({"user_id": sid, "success": False, "error": "Student not found"})
            failure_count += 1
            if job_id:
                _update_training_job_progress(
                    job_id,
                    processed_faces=index,
                    trained_faces=success_count,
                    failed_faces=failure_count,
                    stage="training",
                    message=f"Training {success_count} of {len(user_ids)} faces",
                )
            continue

        try:
            result = _train_embeddings_from_dataset_for_user(user_id)
            total_trained_embeddings += int(result["trained_embeddings"])
            success_count += 1
            items.append(
                {
                    "user_id": _as_text(user_id),
                    "success": True,
                    "trained_embeddings": result["trained_embeddings"],
                    "skipped_images": result["skipped_images"],
                    "dataset_dir": result["dataset_dir"],
                }
            )
        except ValueError as exc:
            items.append({"user_id": _as_text(user_id), "success": False, "error": str(exc)})
            failure_count += 1

        if job_id:
            _update_training_job_progress(
                job_id,
                processed_faces=index,
                trained_faces=success_count,
                failed_faces=failure_count,
                stage="training",
                message=f"Training {success_count} of {len(user_ids)} faces",
            )

    _raise_if_job_cancelled(job_id)
    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=len(user_ids),
            processed_faces=len(user_ids),
            trained_faces=success_count,
            failed_faces=failure_count,
            stage="saving",
            message="Saving trainer artifact",
        )
    trainer_result = _refresh_face_trainer_artifact()
    _raise_if_job_cancelled(job_id)
    log_action(
        "BULK_TRAIN_FACE_FROM_DATASET",
        actor_id,
        details=(
            f"requested={len(user_ids)}, success={success_count}, "
            f"failed={failure_count}, trained_embeddings={total_trained_embeddings}, "
            f"trainer={trainer_result.get('model_path') or 'not_saved'}"
        ),
    )
    return {
        "message": "Bulk training completed",
        "requested_count": len(user_ids),
        "success_count": success_count,
        "failure_count": failure_count,
        "total_trained_embeddings": total_trained_embeddings,
        "items": items,
        "trainer": trainer_result,
    }


def _rebuild_all_faces_job(actor_id, job_id=None):
    profiles = get_all_profiles(["user_id"])
    if not profiles:
        return {"error": "No student profiles found"}

    items = []
    success_count = 0
    failure_count = 0
    total_trained_embeddings = 0

    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=len(profiles),
            processed_faces=0,
            trained_faces=0,
            failed_faces=0,
            stage="training",
            message=f"Training 0 of {len(profiles)} faces",
        )

    for index, profile in enumerate(profiles, start=1):
        _raise_if_job_cancelled(job_id)
        user_id = _as_text(profile.get("user_id"))
        if not user_id:
            failure_count += 1
            items.append({"user_id": None, "success": False, "error": "Missing user_id"})
            if job_id:
                _update_training_job_progress(
                    job_id,
                    processed_faces=index,
                    trained_faces=success_count,
                    failed_faces=failure_count,
                    stage="training",
                    message=f"Training {success_count} of {len(profiles)} faces",
                )
            continue

        try:
            result = _train_embeddings_from_dataset_for_user(user_id)
            success_count += 1
            total_trained_embeddings += int(result["trained_embeddings"])
            items.append(
                {
                    "user_id": user_id,
                    "success": True,
                    "trained_embeddings": result["trained_embeddings"],
                    "skipped_images": result["skipped_images"],
                    "dataset_dir": result["dataset_dir"],
                }
            )
        except Exception as exc:
            failure_count += 1
            items.append({"user_id": user_id, "success": False, "error": str(exc)})

        if job_id:
            _update_training_job_progress(
                job_id,
                processed_faces=index,
                trained_faces=success_count,
                failed_faces=failure_count,
                stage="training",
                message=f"Training {success_count} of {len(profiles)} faces",
            )

    _raise_if_job_cancelled(job_id)
    if job_id:
        _update_training_job_progress(
            job_id,
            total_faces=len(profiles),
            processed_faces=len(profiles),
            trained_faces=success_count,
            failed_faces=failure_count,
            stage="saving",
            message="Saving trainer artifact",
        )
    trainer_result = _refresh_face_trainer_artifact()
    _raise_if_job_cancelled(job_id)
    log_action(
        "REBUILD_ALL_FACE_EMBEDDINGS",
        actor_id,
        details=(
            f"requested={len(profiles)}, success={success_count}, failure={failure_count}, "
            f"trained_embeddings={total_trained_embeddings}, "
            f"trainer={trainer_result.get('model_path') or 'not_saved'}"
        ),
    )

    return {
        "message": "Face embeddings rebuilt",
        "requested_count": len(profiles),
        "success_count": success_count,
        "failure_count": failure_count,
        "total_trained_embeddings": total_trained_embeddings,
        "items": items,
        "trainer": trainer_result,
    }


@admin_bp.route("/students/<sid>/train-face", methods=["POST"])
@admin_bp.route("/students/<sid>/train", methods=["POST"])
@admin_bp.route("/student/<sid>/train-face", methods=["POST"])
@role_required("department_admin")
@validate_ids("sid")
def train_face_from_dataset(user, sid):
    """Train student face embeddings from dataset/<user_id> images and save to DB."""
    user_id, _ = _resolve_user_identity(sid)
    if not user_id:
        return jsonify({"error": "Student not found"}), 404

    d = request.get_json(silent=True) or {}
    async_requested = _to_bool(d.get("async", False))

    if async_requested:
        job_id = _launch_background_job(
            current_app._get_current_object(),
            "train_face_from_dataset",
            {
                "actor_id": str(user["_id"]),
                "user_id": str(user_id),
            },
        )
        _update_training_job_progress(
            job_id,
            total_faces=1,
            processed_faces=0,
            trained_faces=0,
            failed_faces=0,
            stage="queued",
            message="Queued",
        )
        return jsonify({
            "message": "Face training queued",
            "job_id": job_id,
            "status_url": f"/api/admin/jobs/{job_id}",
            "requested_count": 1,
        }), 202

    try:
        train_result = _train_single_face_job(str(user["_id"]), user_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({
        "message": "Face training completed",
        "trained_embeddings": train_result["trained_embeddings"],
        "skipped_images": train_result["skipped_images"],
        "dataset_dir": train_result["dataset_dir"],
    }), 200


@admin_bp.route("/students/train-face/bulk", methods=["POST"])
@admin_bp.route("/students/bulk-train-face", methods=["POST"])
@role_required("department_admin")
def bulk_train_face_from_dataset(user):
    """Train face embeddings in bulk for selected students from their dataset folders."""
    d = request.get_json(silent=True) or {}
    raw_ids = d.get("user_ids") or []
    user_ids = [_as_text(sid) for sid in raw_ids if _as_text(sid)]
    if not user_ids:
        return jsonify({"error": "user_ids is required"}), 400

    async_requested = _to_bool(d.get("async", False))
    if async_requested:
        job_id = _launch_background_job(
            current_app._get_current_object(),
            "bulk_train_face",
            {
                "requested_count": len(user_ids),
                "actor_id": str(user["_id"]),
                "user_ids": user_ids,
            },
        )
        _update_training_job_progress(
            job_id,
            total_faces=len(user_ids),
            processed_faces=0,
            trained_faces=0,
            failed_faces=0,
            stage="queued",
            message="Queued",
        )
        return jsonify({
            "message": "Bulk training queued",
            "job_id": job_id,
            "status_url": f"/api/admin/jobs/{job_id}",
            "requested_count": len(user_ids),
        }), 202

    result = _train_bulk_faces_job(str(user["_id"]), user_ids)
    _clear_query_cache()
    return jsonify(result), 200


@admin_bp.route("/students/train-face/rebuild-all", methods=["POST"])
@role_required("department_admin")
def rebuild_all_face_embeddings(user):
    """Rebuild face embeddings for every student profile from their dataset folders."""
    d = request.get_json(silent=True) or {}
    async_requested = _to_bool(d.get("async", False))

    if async_requested:
        job_id = _launch_background_job(
            current_app._get_current_object(),
            "rebuild_all_face_embeddings",
            {"actor_id": str(user["_id"])},
        )
        profiles = get_all_profiles(["user_id"])
        _update_training_job_progress(
            job_id,
            total_faces=len(profiles),
            processed_faces=0,
            trained_faces=0,
            failed_faces=0,
            stage="queued",
            message="Queued",
        )
        return jsonify({
            "message": "Face embeddings rebuild queued",
            "job_id": job_id,
            "status_url": f"/api/admin/jobs/{job_id}",
            "requested_count": len(profiles),
        }), 202

    result = _rebuild_all_faces_job(str(user["_id"]))
    if result.get("error"):
        return jsonify({"error": result["error"]}), 404
    _clear_query_cache()
    return jsonify(result), 200


@admin_bp.route("/jobs/<job_id>", methods=["GET"])
@role_required("department_admin")
def get_job_status(user, job_id):
    job = _get_background_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(sanitise_mongo_doc(job))

@admin_bp.route("/jobs/<job_id>/cancel", methods=["POST"])
@role_required("department_admin")
def cancel_background_job(user, job_id):
    job = _get_background_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    status = _as_text(job.get("status")).lower()
    if status in {"completed", "dead_letter", "cancelled"}:
        return jsonify({"error": f"Cannot cancel job in '{status}' state"}), 400

    now = _utcnow()
    if status == "queued":
        _update_background_job(
            job_id,
            status="cancelled",
            cancel_requested=True,
            cancelled_at=now,
            finished_at=now,
            training_stage="cancelled",
            training_message="Cancelled by user",
        )
    else:
        _update_background_job(
            job_id,
            cancel_requested=True,
            training_stage="cancelling",
            training_message="Cancellation requested",
        )

    log_action(
        "CANCEL_BACKGROUND_JOB",
        str(user["_id"]),
        details=f"job_id={job_id}, previous_status={status}",
    )

    updated = _get_background_job(job_id)
    return jsonify(
        {
            "message": "Cancellation requested",
            "job": sanitise_mongo_doc(updated) if updated else {"job_id": job_id},
        }
    ), 202


@admin_bp.route("/jobs/<job_id>/replay", methods=["POST"])
@role_required("department_admin")
def replay_dead_letter_job(user, job_id):
    jobs = get_collection("attendance", "background_jobs")
    job = jobs.find_one({"job_id": job_id})
    if not job:
        return jsonify({"error": "Job not found"}), 404

    status = _as_text(job.get("status")).lower()
    if status != "dead_letter":
        return jsonify({"error": "Only dead-letter jobs can be replayed"}), 400

    if not _requeue_dead_letter_job_by_id(job_id):
        return jsonify({"error": "Job replay failed due to concurrent update"}), 409

    log_action(
        "REPLAY_DEAD_LETTER_JOB",
        str(user["_id"]),
        details=f"job_id={job_id}, job_type={_as_text(job.get('job_type'))}",
    )

    return jsonify(
        {
            "message": "Job replay queued",
            "job_id": job_id,
            "status_url": f"/api/admin/jobs/{job_id}",
        }
    ), 202


def _requeue_dead_letter_job_by_id(job_id):
    jobs = get_collection("attendance", "background_jobs")
    now = _utcnow()
    updated = jobs.update_one(
        {"job_id": job_id, "status": "dead_letter"},
        {
            "$set": {
                "status": "queued",
                "attempts": 0,
                "error": None,
                "next_attempt_at": now,
                "updated_at": now,
                "started_at": None,
                "finished_at": None,
                "retry_count": 0,
                "retry_in_seconds": None,
                "last_error_at": None,
                "dead_lettered_at": None,
            }
        },
    )
    if not updated.modified_count:
        return False

    try:
        enqueued = _enqueue_background_job(job_id)
    except Exception:
        current_app.logger.exception("Replay enqueue failed for job %s", job_id)
        enqueued = False

    if not enqueued:
        _schedule_local_retry(job_id, 1)

    return True


def _fetch_dead_letter_rows(filters=None, include_pagination=True):
    filters = filters or {}
    q = _as_text(filters.get("q", "")).lower()
    job_type = _as_text(filters.get("job_type", "")).lower()
    from_raw = _as_text(filters.get("from", ""))
    to_raw = _as_text(filters.get("to", ""))
    sort_by = _as_text(filters.get("sort_by", "updated_at")).lower()
    sort_dir = _as_text(filters.get("sort_dir", "desc")).lower()
    tz_offset_minutes = _to_int(filters.get("tz_offset_minutes", 0), 0)

    allowed_sort_by = {"updated_at", "created_at", "attempts", "job_type"}
    if sort_by not in allowed_sort_by:
        raise ValueError("Invalid sort_by value")
    if sort_dir not in {"asc", "desc"}:
        raise ValueError("Invalid sort_dir value")

    query = {"status": "dead_letter"}
    if job_type:
        query["job_type"] = job_type

    ts_filter = {}
    from_local = _parse_iso_date(from_raw)
    to_local = _parse_iso_date(to_raw)
    if from_raw and not from_local:
        raise ValueError("Invalid from date format")
    if to_raw and not to_local:
        raise ValueError("Invalid to date format")

    if from_local:
        ts_filter["$gte"] = _local_midnight_to_utc(from_local, tz_offset_minutes)
    if to_local:
        to_local_exclusive = to_local + timedelta(days=1)
        ts_filter["$lt"] = _local_midnight_to_utc(to_local_exclusive, tz_offset_minutes)
    if ts_filter:
        query["updated_at"] = ts_filter

    sort_order = -1 if sort_dir == "desc" else 1
    jobs = get_collection("attendance", "background_jobs")
    rows = list(
        jobs.find(
            query,
            {
                "_id": 0,
                "job_id": 1,
                "job_type": 1,
                "error": 1,
                "payload": 1,
                "attempts": 1,
                "max_attempts": 1,
                "retry_count": 1,
                "last_error_at": 1,
                "dead_lettered_at": 1,
                "error_history": 1,
                "created_at": 1,
                "updated_at": 1,
            },
        ).sort(sort_by, sort_order)
    )

    if q:
        filtered = []
        for row in rows:
            if (
                q in _as_text(row.get("job_id")).lower()
                or q in _as_text(row.get("job_type")).lower()
                or q in _as_text(row.get("error")).lower()
            ):
                filtered.append(row)
        rows = filtered

    if not include_pagination:
        return rows, len(rows)

    page = max(1, _to_int(filters.get("page", 1), 1))
    per_page = max(1, min(_to_int(filters.get("per_page", 20), 20), 100))
    total = len(rows)
    start = (page - 1) * per_page
    end = start + per_page
    return rows[start:end], total


@admin_bp.route("/jobs/dead-letter", methods=["GET"])
@role_required("department_admin")
def list_dead_letter_jobs(user):
    filters = {
        "q": request.args.get("q", ""),
        "job_type": request.args.get("job_type", ""),
        "from": request.args.get("from", ""),
        "to": request.args.get("to", ""),
        "sort_by": request.args.get("sort_by", "updated_at"),
        "sort_dir": request.args.get("sort_dir", "desc"),
        "tz_offset_minutes": request.args.get("tz_offset_minutes", 0),
        "page": request.args.get("page", 1),
        "per_page": request.args.get("per_page", 20),
    }
    try:
        rows, total = _fetch_dead_letter_rows(filters, include_pagination=True)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    page = max(1, _to_int(filters.get("page", 1), 1))
    per_page = max(1, min(_to_int(filters.get("per_page", 20), 20), 100))
    return jsonify(
        {
            "items": sanitise_many(rows),
            "total": total,
            "page": page,
            "per_page": per_page,
        }
    )


@admin_bp.route("/jobs/dead-letter/replay-bulk", methods=["POST"])
@role_required("department_admin")
def replay_dead_letter_jobs_bulk(user):
    d = request.get_json(silent=True) or {}
    raw_ids = d.get("job_ids") or []
    job_ids = [_as_text(x) for x in raw_ids if _as_text(x)]
    if not job_ids:
        return jsonify({"error": "job_ids is required"}), 400

    replayed = 0
    skipped = 0
    for job_id in job_ids:
        if _requeue_dead_letter_job_by_id(job_id):
            replayed += 1
        else:
            skipped += 1

    log_action(
        "REPLAY_DEAD_LETTER_JOB_BULK",
        str(user["_id"]),
        details=f"requested={len(job_ids)}, replayed={replayed}, skipped={skipped}",
    )

    return jsonify(
        {
            "message": "Bulk dead-letter replay processed",
            "requested": len(job_ids),
            "replayed": replayed,
            "skipped": skipped,
        }
    ), 200


@admin_bp.route("/jobs/dead-letter/replay-filtered", methods=["POST"])
@role_required("department_admin")
def replay_dead_letter_jobs_filtered(user):
    d = request.get_json(silent=True) or {}
    filters = {
        "q": d.get("q", ""),
        "job_type": d.get("job_type", ""),
        "from": d.get("from", ""),
        "to": d.get("to", ""),
        "sort_by": d.get("sort_by", "updated_at"),
        "sort_dir": d.get("sort_dir", "desc"),
        "tz_offset_minutes": d.get("tz_offset_minutes", 0),
    }
    limit = max(1, min(_to_int(d.get("limit", 500), 500), 1000))

    try:
        rows, total_matched = _fetch_dead_letter_rows(filters, include_pagination=False)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    replayed = 0
    skipped = 0
    requested_rows = rows[:limit]
    for row in requested_rows:
        job_id = _as_text(row.get("job_id"))
        if not job_id:
            skipped += 1
            continue
        if _requeue_dead_letter_job_by_id(job_id):
            replayed += 1
        else:
            skipped += 1

    log_action(
        "REPLAY_DEAD_LETTER_JOB_FILTERED",
        str(user["_id"]),
        details=(
            f"matched={total_matched}, limit={limit}, requested={len(requested_rows)}, "
            f"replayed={replayed}, skipped={skipped}"
        ),
    )

    return jsonify(
        {
            "message": "Filtered dead-letter replay processed",
            "matched": total_matched,
            "limit": limit,
            "requested": len(requested_rows),
            "replayed": replayed,
            "skipped": skipped,
        }
    ), 200


@admin_bp.route("/jobs/metrics", methods=["GET"])
@role_required("department_admin")
def get_job_metrics(user):
    jobs = get_collection("attendance", "background_jobs")
    summary = {
        "queued": 0,
        "running": 0,
        "completed": 0,
        "dead_letter": 0,
    }

    for row in jobs.aggregate([
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]):
        status = _as_text(row.get("_id")).lower()
        if status in summary:
            summary[status] = int(row.get("count", 0) or 0)

    queue_depth = None
    delayed_queue_depth = None
    due_delayed_count = None
    client = _get_task_queue_client()
    if client is not None:
        try:
            queue_name, delayed_queue_name = _get_queue_names()
            queue_depth = int(client.llen(queue_name) or 0)
            delayed_queue_depth = int(client.zcard(delayed_queue_name) or 0)
            due_delayed_count = int(client.zcount(delayed_queue_name, 0, int(time.time())) or 0)
        except Exception:
            current_app.logger.exception("Unable to read queue metrics")

    running_timeout_seconds = max(30, _to_int(current_app.config.get("TASK_QUEUE_RUNNING_TIMEOUT_SECONDS", 900), 900))
    stale_cutoff = _utcnow() - timedelta(seconds=running_timeout_seconds)
    stale_running_count = int(
        jobs.count_documents({"status": "running", "updated_at": {"$lte": stale_cutoff}})
    )
    queued_retry_count = int(
        jobs.count_documents({
            "status": "queued",
            "retry_count": {"$gt": 0},
            "next_attempt_at": {"$ne": None},
        })
    )
    next_retry_candidates = list(
        jobs.find(
            {
                "status": "queued",
                "retry_count": {"$gt": 0},
                "next_attempt_at": {"$ne": None},
            },
            {
                "_id": 0,
                "job_id": 1,
                "job_type": 1,
                "next_attempt_at": 1,
                "retry_count": 1,
                "error": 1,
            },
        )
    )
    next_retry_candidates.sort(key=lambda row: row.get("next_attempt_at") or datetime.max)
    next_retry_job = next_retry_candidates[0] if next_retry_candidates else None
    dead_letter_last_24h = int(
        jobs.count_documents({"status": "dead_letter", "updated_at": {"$gte": _utcnow() - timedelta(hours=24)}})
    )
    recent_dead_letter_jobs = sanitise_many(
        list(
            jobs.find(
                {"status": "dead_letter"},
                {
                    "_id": 0,
                    "job_id": 1,
                    "job_type": 1,
                    "error": 1,
                    "updated_at": 1,
                    "attempts": 1,
                    "max_attempts": 1,
                },
            )
            .sort("updated_at", -1)
            .limit(5)
        )
    )

    return jsonify(
        {
            "jobs": {
                "total": int(sum(summary.values())),
                **summary,
                "stale_running": stale_running_count,
                "queued_retries": queued_retry_count,
                "next_retry_job": sanitise_mongo_doc(next_retry_job) if next_retry_job else None,
                "dead_letter_last_24h": dead_letter_last_24h,
                "recent_dead_letter_jobs": recent_dead_letter_jobs,
            },
            "queue": {
                "enabled": bool(current_app.config.get("TASK_QUEUE_ENABLED", False)),
                "depth": queue_depth,
                "delayed_depth": delayed_queue_depth,
                "due_delayed": due_delayed_count,
                "running_timeout_seconds": running_timeout_seconds,
            },
        }
    )


@admin_bp.route("/capture-faces", methods=["POST"])
@role_required("department_admin")
def capture_faces_dataset(user):
    """Capture webcam dataset images for a named user into dataset/<user_name>."""
    d = request.get_json(silent=True) or {}
    user_name = _as_text(d.get("user_name", ""))
    total_images = _to_int(d.get("total_images"), 50)

    if not user_name:
        return jsonify({"error": "user_name is required"}), 400
    if total_images <= 0:
        return jsonify({"error": "total_images must be greater than 0"}), 400

    try:
        saved_paths = capture_faces_for_user(
            user_name=user_name,
            dataset_root="dataset",
            total_images=total_images,
            delay_seconds=0.1,
        )
    except Exception as exc:
        current_app.logger.exception("Dataset capture failed")
        return jsonify({"error": f"Dataset capture failed: {exc}"}), 500

    log_action(
        "CAPTURE_FACE_DATASET",
        str(user["_id"]),
        details=f"Captured {len(saved_paths)} images for {user_name}",
    )

    return jsonify({
        "message": "Face dataset captured successfully",
        "captured_count": len(saved_paths),
        "dataset_folder": os.path.dirname(saved_paths[0]) if saved_paths else "dataset",
    }), 200


@admin_bp.route("/courses/reassign", methods=["POST"])
@role_required("department_admin")
def reassign_course_entities(user):
    """Batch move students/papers from one course to another active course."""
    d = request.get_json(silent=True) or {}
    from_course_id = _as_text(d.get("from_course_id"))
    to_course_id = _as_text(d.get("to_course_id"))
    move_students = _to_bool(d.get("move_students", True))
    move_papers = _to_bool(d.get("move_papers", True))

    if not from_course_id or not to_course_id:
        return jsonify({"error": "from_course_id and to_course_id are required"}), 400
    if from_course_id == to_course_id:
        return jsonify({"error": "from_course_id and to_course_id must be different"}), 400

    source_course = _safe_get_course(from_course_id)
    if not source_course:
        return jsonify({"error": "Source course not found"}), 404

    target_course, target_error = _get_active_course_or_error(to_course_id)
    if target_error:
        return target_error

    profile_col = get_collection("academic", "student_profiles")
    paper_col = get_collection("academic", "papers")

    moved_students = 0
    moved_papers = 0
    if move_students:
        res = profile_col.update_many({"course_id": from_course_id}, {"$set": {"course_id": to_course_id}})
        moved_students = int(res.modified_count)

    if move_papers:
        res = paper_col.update_many({"course_id": from_course_id}, {"$set": {"course_id": to_course_id}})
        moved_papers = int(res.modified_count)

    log_action(
        "REASSIGN_COURSE_ENTITIES",
        str(user["_id"]),
        details=(
            f"from={from_course_id}, to={to_course_id}, "
            f"move_students={move_students}, move_papers={move_papers}, "
            f"moved_students={moved_students}, moved_papers={moved_papers}"
        ),
    )
    _clear_query_cache()

    return jsonify(
        {
            "message": "Reassignment completed",
            "from_course": sanitise_mongo_doc(source_course),
            "to_course": sanitise_mongo_doc(target_course),
            "moved_students": moved_students,
            "moved_papers": moved_papers,
        }
    ), 200


# ─── Audit Trail ────────────────────────────────────────────────────────────

_AUDIT_EXCLUDED_ACTIONS = ["HEARTBEAT", "QUEUE_CHECK", "STATUS_CHECK"]

@admin_bp.route("/audit-logs", methods=["GET"])
@role_required("department_admin")
def list_audit_logs(user):
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    action = _as_text(request.args.get("action", "")).upper()
    date_from = _as_text(request.args.get("from", ""))
    date_to = _as_text(request.args.get("to", ""))
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    filters = {}
    if action:
        # Contains match allows flexible keyword search like OVERRIDE, CREATE, DELETE, etc.
        filters["action"] = {"$regex": re.escape(action), "$options": "i"}

    ts_filter = {}
    parsed_from = _parse_iso_date(date_from)
    parsed_to = _parse_iso_date(date_to)

    if parsed_from:
        ts_filter["$gte"] = _local_midnight_to_utc(parsed_from, tz_offset_minutes)
    if parsed_to:
        parsed_to_exclusive = parsed_to + timedelta(days=1)
        ts_filter["$lt"] = _local_midnight_to_utc(parsed_to_exclusive, tz_offset_minutes)

    if ts_filter:
        filters["timestamp"] = ts_filter

    if filters:
        filters = {
            "$and": [
                filters,
                {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}},
            ]
        }
    else:
        filters = {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}}

    dept_id_param = _as_text(request.args.get("department_id", ""))

    if is_super_admin(user):
        # Super admin: optionally scope to a department by ID
        dept_filter_id = None
        if dept_id_param:
            try:
                dept_filter_id = ObjectId(dept_id_param)
            except Exception:
                pass
        if dept_filter_id:
            # Get all user IDs in this department to filter logs
            users_col = get_collection("auth", "users")
            dept_user_ids = [
                str(u["_id"])
                for u in users_col.find({"department_id": dept_filter_id}, {"_id": 1})
            ]
            dept_user_ids_oids = [ObjectId(uid) for uid in dept_user_ids if ObjectId.is_valid(uid)]
            all_user_id_variants = dept_user_ids + dept_user_ids_oids  # type: ignore[operator]
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": dept_filter_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter
        logs, total = get_audit_logs(page, per_page, filters, department_id=None)
    else:
        # Department admin: locked to their own department — fetch dept user IDs
        user_dept_id = _user_dept_id(user)
        if user_dept_id:
            users_col = get_collection("auth", "users")
            dept_user_ids = [
                str(u["_id"])
                for u in users_col.find({"department_id": user_dept_id}, {"_id": 1})
            ]
            dept_user_ids_oids = [ObjectId(uid) for uid in dept_user_ids if ObjectId.is_valid(uid)]
            all_user_id_variants = dept_user_ids + dept_user_ids_oids  # type: ignore[operator]
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": user_dept_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter
        logs, total = get_audit_logs(page, per_page, filters, department_id=None)
    audit_user_ids = [
        item.get("performed_by") or item.get("actor_user_id")
        for item in logs
        if item.get("performed_by") or item.get("actor_user_id")
    ] + [
        item.get("target_user")
        or ((item.get("details") or {}).get("user_id") if isinstance(item.get("details"), dict) else None)
        for item in logs
        if item.get("target_user")
        or (isinstance(item.get("details"), dict) and (item.get("details") or {}).get("user_id"))
    ]
    user_map = get_users_by_ids(audit_user_ids)

    enriched = []
    for raw in logs:
        # We start with a copy for serialisation
        item = sanitise_mongo_doc(raw)

        # Re-fetch raw versions for local logic that requires datetime objects
        raw_ts = raw.get("timestamp")
        raw_rollback_until = raw.get("rollback_until")
        raw_rollback_payload = raw.get("rollback")
        raw_rolled_back = bool(raw.get("rolled_back"))

        actor_id = raw.get("performed_by") or raw.get("actor_user_id")
        details_user_id = (raw.get("details") or {}).get("user_id") if isinstance(raw.get("details"), dict) else None
        target_user_id = raw.get("target_user") or details_user_id

        actor = user_map.get(_as_text(actor_id)) if actor_id else None
        target_user = user_map.get(_as_text(target_user_id)) if target_user_id else None

        item["actor_name"] = (actor or {}).get("name") or ("System" if str(actor_id).lower() == "system" else "Unknown User")
        item["actor_email"] = (actor or {}).get("email") or ""
        item["role"] = (actor or {}).get("role") or raw.get("role") or "unknown"

        if target_user:
            item["target_type"] = f"{target_user.get('name', 'Unknown')} ({target_user.get('role', 'user')})"
            item["target_user_name"] = target_user.get("name")
            item["target_user_email"] = target_user.get("email")
            item["target_user_role"] = target_user.get("role")
        elif target_user_id:
            item["target_type"] = f"User {target_user_id}"
        else:
            item["target_type"] = _as_text(raw.get("details")) or "System"

        item["ip"] = raw.get("ip") or raw.get("ip_address") or ""

        # Handle time-sensitive rollback logic on raw datetime objects
        if raw_ts and raw_ts.tzinfo is None:
            raw_ts = raw_ts.replace(tzinfo=timezone.utc)
        
        if raw_rollback_payload and not raw_rollback_until and raw_ts:
            raw_rollback_until = raw_ts + timedelta(days=1)
        
        if raw_rollback_until and raw_rollback_until.tzinfo is None:
            raw_rollback_until = raw_rollback_until.replace(tzinfo=timezone.utc)

        now = datetime.now(timezone.utc)
        eligible = bool(raw_rollback_payload) and not raw_rolled_back and bool(raw_rollback_until) and now <= raw_rollback_until

        item["rollback_available"] = eligible
        item["rolled_back"] = raw_rolled_back
        if raw_rollback_until:
            item["rollback_until"] = raw_rollback_until.isoformat()

        # Ensure IDs are stringified
        item["performed_by"] = _as_text(actor_id) if actor_id else None
        item["target_user"] = _as_text(target_user_id) if target_user_id else None

        item.pop("rollback", None)
        enriched.append(item)

    return jsonify({
        "logs": enriched,
        "total": total,
        "page": page,
        "per_page": per_page,
    })


@admin_bp.route("/audit-logs/export", methods=["GET"])
@role_required("department_admin")
def export_audit_logs(user):
    """Export the filtered audit logs to an Excel file."""
    action = _as_text(request.args.get("action", "")).upper()
    date_from = _as_text(request.args.get("from", ""))
    date_to = _as_text(request.args.get("to", ""))
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    filters = {}
    if action:
        filters["action"] = {"$regex": re.escape(action), "$options": "i"}

    ts_filter = {}
    parsed_from = _parse_iso_date(date_from)
    parsed_to = _parse_iso_date(date_to)

    if parsed_from:
        ts_filter["$gte"] = _local_midnight_to_utc(parsed_from, tz_offset_minutes)
    if parsed_to:
        parsed_to_exclusive = parsed_to + timedelta(days=1)
        ts_filter["$lt"] = _local_midnight_to_utc(parsed_to_exclusive, tz_offset_minutes)

    if ts_filter:
        filters["timestamp"] = ts_filter

    if filters:
        filters = {
            "$and": [
                filters,
                {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}},
            ]
        }
    else:
        filters = {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}}

    dept_id_param = _as_text(request.args.get("department_id", ""))

    if is_super_admin(user):
        dept_filter_id = None
        if dept_id_param:
            try:
                dept_filter_id = ObjectId(dept_id_param)
            except Exception:
                pass
        if dept_filter_id:
            users_col = get_collection("auth", "users")
            dept_user_ids = [str(u["_id"]) for u in users_col.find({"department_id": dept_filter_id}, {"_id": 1})]
            dept_user_ids_oids = [ObjectId(uid) for uid in dept_user_ids if ObjectId.is_valid(uid)]
            all_user_id_variants = dept_user_ids + dept_user_ids_oids
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": dept_filter_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter
    else:
        user_dept_id = _user_dept_id(user)
        if user_dept_id:
            users_col = get_collection("auth", "users")
            dept_user_ids = [str(u["_id"]) for u in users_col.find({"department_id": user_dept_id}, {"_id": 1})]
            dept_user_ids_oids = [ObjectId(uid) for uid in dept_user_ids if ObjectId.is_valid(uid)]
            all_user_id_variants = dept_user_ids + dept_user_ids_oids
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": user_dept_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter

    # Fetch up to 10,000 logs for export to protect memory
    logs, _ = get_audit_logs(1, 10000, filters, department_id=None)
    
    audit_user_ids = [
        item.get("performed_by") or item.get("actor_user_id")
        for item in logs
        if item.get("performed_by") or item.get("actor_user_id")
    ] + [
        item.get("target_user") or ((item.get("details") or {}).get("user_id") if isinstance(item.get("details"), dict) else None)
        for item in logs
        if item.get("target_user") or (isinstance(item.get("details"), dict) and (item.get("details") or {}).get("user_id"))
    ]
    user_map = get_users_by_ids(audit_user_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    headers = ["Timestamp", "Actor Name", "Actor Email", "Role", "Action", "Target", "IP Address", "Details"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)

    def _safe_str(v):
        if not v:
            return ""
        try:
            return str(v)
        except Exception:
            return ""

    for raw in logs:
        raw_ts = raw.get("timestamp")
        if raw_ts and raw_ts.tzinfo is None:
            raw_ts = raw_ts.replace(tzinfo=timezone.utc)
            
        ts_str = raw_ts.isoformat() if raw_ts else ""
        
        actor_id = raw.get("performed_by") or raw.get("actor_user_id")
        details_user_id = (raw.get("details") or {}).get("user_id") if isinstance(raw.get("details"), dict) else None
        target_user_id = raw.get("target_user") or details_user_id

        actor = user_map.get(_as_text(actor_id)) if actor_id else None
        target_user = user_map.get(_as_text(target_user_id)) if target_user_id else None

        a_name = (actor or {}).get("name") or ("System" if str(actor_id).lower() == "system" else "Unknown User")
        a_mail = (actor or {}).get("email") or ""
        a_role = (actor or {}).get("role") or raw.get("role") or "unknown"
        action_name = raw.get("action", "")
        
        if target_user:
            t_type = f"{target_user.get('name', 'Unknown')} ({target_user.get('role', 'user')})"
        elif target_user_id:
            t_type = f"User {target_user_id}"
        else:
            t_type = _as_text(raw.get("details")) or "System"

        ip_addr = raw.get("ip") or raw.get("ip_address") or ""
        details = _safe_str(raw.get("details"))

        ws.append([ts_str, a_name, a_mail, a_role, action_name, t_type, ip_addr, details])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Audit_Trail_Export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route("/audit-logs/<log_id>/rollback", methods=["POST"])
@role_required("department_admin")
def rollback_audit_action(user, log_id):
    audit_log = get_audit_log_by_id(log_id)
    if not audit_log:
        return jsonify({"error": "Audit log not found"}), 404

    if audit_log.get("rolled_back"):
        return jsonify({"error": "This action has already been rolled back"}), 400

    rollback_payload = audit_log.get("rollback")
    if not rollback_payload:
        return jsonify({"error": "Rollback not available for this action"}), 400

    rollback_until = audit_log.get("rollback_until")
    raw_ts = audit_log.get("timestamp")

    # Normalize timestamp for derivation if needed
    if raw_ts and raw_ts.tzinfo is None:
        raw_ts = raw_ts.replace(tzinfo=timezone.utc)

    if not rollback_until:
        rollback_until = (raw_ts or datetime.now(timezone.utc)) + timedelta(days=1)

    # Normalize rollback_until for comparison
    if rollback_until and rollback_until.tzinfo is None:
        rollback_until = rollback_until.replace(tzinfo=timezone.utc)

    if datetime.now(timezone.utc) > rollback_until:
        return jsonify({"error": "Rollback window expired (1 day)"}), 403

    try:
        _execute_rollback_operation(rollback_payload)
    except Exception as exc:
        current_app.logger.exception("Audit rollback failed")
        return jsonify({"error": f"Rollback failed: {exc}"}), 500

    mark_audit_log_rolled_back(log_id, str(user["_id"]))
    log_action(
        "ROLLBACK_ACTION",
        str(user["_id"]),
        details=f"Rolled back audit log {log_id}",
    )
    return jsonify({"message": "Rollback completed successfully"}), 200


# ─── Attendance Override ────────────────────────────────────────────────────

@admin_bp.route("/attendance/override", methods=["POST"])
@role_required("department_admin")
def override_attendance(user):
    """Manually add or remove an attendance record (Special Exam Access)."""
    d = request.get_json(silent=True) or {}
    action = d.get("action", "add")  # "add" or "remove"

    if action == "add":
        log = log_attendance(
            d["paper_id"], d["user_id"], str(user["_id"]),
            session_id="manual-override", method="manual",
        )
        log_action("ATTENDANCE_OVERRIDE_ADD", str(user["_id"]),
                   target_user=d["user_id"],
                   details=f"Paper {d['paper_id']}")
        _clear_query_cache()
        return jsonify({"message": "Attendance added", "log": log}), 201
    else:
        from app.models.attendance import delete_attendance_log
        delete_attendance_log(d["log_id"])
        log_action("ATTENDANCE_OVERRIDE_REMOVE", str(user["_id"]),
                   details=f"Log {d['log_id']}")
        _clear_query_cache()
        return jsonify({"message": "Attendance removed"}), 200


@admin_bp.route("/exam-eligibility-summary", methods=["GET"])
@role_required("super_admin", "department_admin")
def exam_eligibility_summary(user):
    """Admin view of exam eligibility with filters and override states."""
    cache_key = (
        "exam_eligibility_summary",
        tuple(sorted((k, _as_text(v)) for k, v in request.args.items())),
    )
    cached_payload = _cache_get(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    department_filter = _as_text(request.args.get("department", ""))
    department_id_filter = _as_text(request.args.get("department_id", ""))
    course_id = _as_text(request.args.get("course_id", ""))
    paper_id = _as_text(request.args.get("paper_id", ""))
    academic_session = _normalise_year(request.args.get("academic_session", "")) or _normalise_year(request.args.get("academic_year", ""))
    semester_filter = _as_text(request.args.get("semester", ""))
    q = _as_text(request.args.get("q", "")).lower()
    final_eligible_filter = _as_text(request.args.get("final_eligible", ""))
    include_inactive = _to_bool(request.args.get("include_inactive", False))

    profiles_col = get_collection("academic", "student_profiles")

    # Build dept scope: dept admins are always scoped; super admins may filter by department_id
    if is_super_admin(user):
        dept_scope_id = department_id_filter or None
    else:
        dept_scope_id = _user_dept_id(user)

    courses = sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year"], department_id=dept_scope_id))
    papers = sanitise_many(get_all_papers(["name", "code", "semester", "course_id", "lecturer_id", "created_at"]))
    course_map = {c["_id"]: c for c in courses}
    paper_map = {p["_id"]: p for p in papers}

    # Restrict profiles to only courses visible to this admin
    visible_course_ids = list(course_map.keys())
    base_profile_query = {"course_id": {"$in": visible_course_ids}} if visible_course_ids else {"course_id": "never_match"}

    profiles = list(
        profiles_col.find(
            base_profile_query,
            {
                "user_id": 1,
                "course_id": 1,
                "academic_year": 1,
                "academic_session": 1,
                "year": 1,
                "current_semester": 1,
                "enrolled_papers": 1,
                "reg_number": 1,
                "roll_number": 1,
                "created_at": 1,
            },
        )
    )
    user_map = get_users_by_ids(profile.get("user_id") for profile in profiles)
    overrides_col = get_collection("attendance", "exam_eligibility_overrides")
    sessions_col = get_collection("attendance", "attendance_sessions")

    classes_happened_by_paper = {}
    for row in sessions_col.aggregate([
        {
            "$group": {
                "_id": {
                    "paper_id": "$paper_id",
                },
                "count": {"$sum": 1},
            }
        }
    ]):
        gid = row.get("_id") or {}
        gid_paper = _as_text(gid.get("paper_id"))
        count = int(row.get("count", 0) or 0)

        if gid_paper:
            classes_happened_by_paper[gid_paper] = classes_happened_by_paper.get(gid_paper, 0) + count

    selected_profiles = []
    relevant_user_ids = []
    relevant_paper_ids = set()

    for profile in profiles:
        uid = profile.get("user_id")
        if not uid:
            continue
        uid_text = _as_text(uid)
        student = user_map.get(uid_text)
        if not student:
            continue

        stu_course_id = _as_text(profile.get("course_id", ""))
        course_doc = course_map.get(stu_course_id)
        course_status = _as_text((course_doc or {}).get("status") or "active").lower() or "active"
        if course_status != "active" and not include_inactive:
            continue
        # Department filter – match on course's department name
        if department_filter:
            course_department = _as_text((course_doc or {}).get("department") or "")
            if course_department.lower() != department_filter.lower():
                continue
        stu_year = _as_text(profile.get("academic_session") or profile.get("academic_year") or profile.get("year"))
        enrolled = profile.get("enrolled_papers", []) or []
        stu_semester = _to_int(profile.get("current_semester"), 0) or None
        if stu_semester is None:
            derived_semesters = []
            for pid in enrolled:
                pdoc = paper_map.get(pid) or {}
                psem = _to_int(pdoc.get("semester"), 0)
                if psem > 0:
                    derived_semesters.append(psem)
            if derived_semesters:
                stu_semester = max(derived_semesters)

        if course_id and stu_course_id != course_id:
            continue
        if academic_session and stu_year != academic_session:
            continue

        selected_profiles.append((profile, student, uid_text, stu_course_id, stu_year, enrolled, stu_semester))
        relevant_user_ids.append(uid_text)
        for pid in enrolled:
            pid_text = _as_text(pid)
            if paper_id and pid_text != paper_id:
                continue
            paper = paper_map.get(pid_text) or paper_map.get(pid)
            if not paper:
                continue
            paper_semester = _to_int(paper.get("semester"), 0)
            if semester_filter and str(paper_semester) != semester_filter:
                continue
            relevant_paper_ids.add(pid_text)

    student_match_ids = []
    for sid in relevant_user_ids:
        student_match_ids.extend(_id_variants(sid))
    paper_match_ids = []
    for pid in relevant_paper_ids:
        paper_match_ids.extend(_id_variants(pid))

    attendance_count_map = {}
    if student_match_ids and paper_match_ids:
        attendance_logs = get_collection("attendance", "attendance_logs")
        for row in attendance_logs.aggregate([
            {
                "$match": {
                    "user_id": {"$in": student_match_ids},
                    "paper_id": {"$in": paper_match_ids},
                }
            },
            {
                "$group": {
                    "_id": {
                        "user_id": "$user_id",
                        "paper_id": "$paper_id",
                    },
                    "count": {"$sum": 1},
                }
            },
        ]):
            gid = row.get("_id") or {}
            attendance_count_map[(
                _as_text(gid.get("user_id")),
                _as_text(gid.get("paper_id")),
            )] = int(row.get("count", 0) or 0)

    override_map = {}
    if student_match_ids and paper_match_ids:
        for override in overrides_col.find(
            {
                "user_id": {"$in": student_match_ids},
                "paper_id": {"$in": paper_match_ids},
            },
            {
                "_id": 0,
                "user_id": 1,
                "paper_id": 1,
                "override_status": 1,
                "reason": 1,
            },
        ):
            key = (_as_text(override.get("user_id")), _as_text(override.get("paper_id")))
            override_map[key] = {
                "override_status": override.get("override_status"),
                "reason": _as_text(override.get("reason", "")),
            }

    items = []
    for profile, student, uid, stu_course_id, stu_year, enrolled, stu_semester in selected_profiles:
        course = course_map.get(stu_course_id)
        per_paper_rows = []
        total_attended_overall = 0
        total_classes_overall = 0

        for pid in enrolled:
            pid_text = _as_text(pid)
            if paper_id and pid_text != paper_id:
                continue

            paper = paper_map.get(pid_text) or paper_map.get(pid)
            if not paper:
                continue

            paper_semester = _to_int(paper.get("semester"), 0)
            if semester_filter and str(paper_semester) != semester_filter:
                continue

            lecturer_id_for_paper = _as_text(paper.get("lecturer_id", ""))
            profile_created_at = profile.get("created_at")

            # Count classes for this subject across all lecturers,
            # scoped to sessions after the student was enrolled.
            session_query = {"paper_id": {"$in": _id_variants(pid_text)}}

            if profile_created_at:
                session_query["$or"] = [
                    {"committed_at": {"$gte": profile_created_at}},
                    {
                        "committed_at": {"$exists": False},
                        "last_updated_at": {"$gte": profile_created_at},
                    },
                    {
                        "committed_at": {"$exists": False},
                        "last_updated_at": {"$exists": False},
                        "created_at": {"$gte": profile_created_at},
                    },
                ]
                classes_happened = int(sessions_col.count_documents(session_query) or 0)
            else:
                classes_happened = int(classes_happened_by_paper.get(pid_text, 0) or 0)

            attended = attendance_count_map.get((_as_text(uid), pid_text), 0)
            pct = round((attended / classes_happened) * 100, 2) if classes_happened > 0 else 0.0

            total_attended_overall += attended
            total_classes_overall += classes_happened

            override = override_map.get((_as_text(uid), pid_text))
            override_status = None if not override else override.get("override_status")
            override_reason = "" if not override else override.get("reason", "")

            if q and not (
                q in _as_text(student.get("name", "")).lower()
                or q in _as_text(student.get("email", "")).lower()
                or q in _as_text(profile.get("reg_number") or profile.get("roll_number")).lower()
                or q in _as_text(paper.get("name", "")).lower()
                or q in _as_text(paper.get("code", "")).lower()
            ):
                continue

            per_paper_rows.append({
                "user_id": uid,
                "student_name": student.get("name", "Unknown"),
                "student_email": student.get("email", ""),
                "reg_number": profile.get("reg_number") or profile.get("roll_number"),
                "course_id": stu_course_id,
                "course_name": (course or {}).get("name"),
                "course_department": (course or {}).get("department") or "",
                "student_semester": stu_semester,
                "paper_id": pid_text,
                "paper_name": paper.get("name", ""),
                "paper_code": paper.get("code", ""),
                "semester": paper_semester or None,
                "lecturer_id": lecturer_id_for_paper,
                "academic_year": stu_year,
                "academic_session": stu_year,
                "enrolled_since": profile_created_at,
                "attended": attended,
                "total_classes": classes_happened,
                "attended_classes": attended,
                "classes_happened": classes_happened,
                "attendance_percentage": pct,
                "override_status": override_status,
                "override_reason": override_reason,
            })

        overall_pct = round((total_attended_overall / total_classes_overall) * 100, 2) if total_classes_overall > 0 else 0.0
        has_lectures = total_classes_overall > 0
        overall_eligible = (overall_pct >= 75.0) if has_lectures else None

        for row in per_paper_rows:
            override_status = row.get("override_status")
            final_eligible = overall_eligible if override_status is None else bool(override_status)
            if final_eligible is None:
                eligibility_status = "no_lectures_yet"
            else:
                eligibility_status = "eligible" if final_eligible else "ineligible"

            if final_eligible_filter:
                required = _to_bool(final_eligible_filter)
                if final_eligible is None or final_eligible != required:
                    continue

            row["overall_attendance_percentage"] = overall_pct
            row["overall_attended_classes"] = total_attended_overall
            row["overall_total_classes"] = total_classes_overall
            row["eligible_by_attendance"] = overall_eligible
            row["final_eligible"] = final_eligible
            row["eligibility_status"] = eligibility_status
            items.append(row)

    payload = {
        "total": len(items),
        "eligible_count": sum(1 for x in items if x["final_eligible"] is True),
        "ineligible_count": sum(1 for x in items if x["final_eligible"] is False),
        "items": items,
    }
    _cache_set(cache_key, payload, _ELIGIBILITY_CACHE_TTL_SECONDS)
    return jsonify(payload)


@admin_bp.route("/exam-eligibility-override", methods=["PUT"])
@role_required("department_admin")
def set_exam_eligibility_override(user):
    """Override final exam eligibility status for a student-paper pair."""
    d = request.get_json(silent=True) or {}
    user_id = _as_text(d.get("user_id", ""))
    paper_id = _as_text(d.get("paper_id", ""))
    reason = _as_text(d.get("reason", ""))

    if not user_id or not paper_id:
        return jsonify({"error": "user_id and paper_id are required"}), 400

    if d.get("override_status", None) is None:
        return jsonify({"error": "override_status must be true or false"}), 400

    raw_status = d.get("override_status")
    if isinstance(raw_status, str):
        raw_lower = raw_status.strip().lower()
        if raw_lower not in {"1", "0", "true", "false", "yes", "no", "y", "n"}:
            return jsonify({"error": "override_status must be true or false"}), 400
    override_status = _to_bool(raw_status)

    overrides_col = get_collection("attendance", "exam_eligibility_overrides")
    overrides_col.update_one(
        {
            "user_id": {"$in": _id_variants(user_id)},
            "paper_id": {"$in": _id_variants(paper_id)},
        },
        {
            "$set": {
                "user_id": user_id,
                "paper_id": paper_id,
                "override_status": override_status,
                "reason": reason,
                "updated_by": str(user["_id"]),
                "updated_at": datetime.now(timezone.utc),
            }
        },
        upsert=True,
    )

    log_action(
        "EXAM_ELIGIBILITY_OVERRIDE",
        str(user["_id"]),
        target_user=user_id,
        details=f"Paper {paper_id}, override={override_status}, reason={reason}",
    )
    _clear_query_cache()
    return jsonify({"message": "Eligibility override updated"}), 200


@admin_bp.route("/exam-eligibility-override/bulk", methods=["PUT"])
@role_required("department_admin")
def set_exam_eligibility_override_bulk(user):
    """Bulk override final exam eligibility for multiple student-paper pairs."""
    d = request.get_json(silent=True) or {}
    overrides = d.get("overrides")

    if not isinstance(overrides, list) or len(overrides) == 0:
        return jsonify({"error": "overrides must be a non-empty list"}), 400

    sanitized = []
    for item in overrides:
        if not isinstance(item, dict):
            continue

        user_id = _as_text(item.get("user_id", ""))
        paper_id = _as_text(item.get("paper_id", ""))
        if not user_id or not paper_id:
            continue
        if item.get("override_status", None) is None:
            continue
        if isinstance(item.get("override_status"), str):
            raw_lower = item.get("override_status", "").strip().lower()
            if raw_lower not in {"1", "0", "true", "false", "yes", "no", "y", "n"}:
                continue

        sanitized.append({
            "user_id": user_id,
            "paper_id": paper_id,
            "override_status": _to_bool(item.get("override_status")),
            "reason": _as_text(item.get("reason", "")),
        })

    if not sanitized:
        return jsonify({"error": "No valid override items found"}), 400

    overrides_col = get_collection("attendance", "exam_eligibility_overrides")
    now = datetime.now(timezone.utc)
    admin_id = str(user["_id"])
    unique_pairs = set()

    for item in sanitized:
        pair = (item["user_id"], item["paper_id"])
        if pair in unique_pairs:
            continue
        unique_pairs.add(pair)
        overrides_col.update_one(
            {
                "user_id": {"$in": _id_variants(item["user_id"])},
                "paper_id": {"$in": _id_variants(item["paper_id"])},
            },
            {
                "$set": {
                    "user_id": item["user_id"],
                    "paper_id": item["paper_id"],
                    "override_status": item["override_status"],
                    "reason": item["reason"],
                    "updated_by": admin_id,
                    "updated_at": now,
                }
            },
            upsert=True,
        )

    log_action(
        "EXAM_ELIGIBILITY_OVERRIDE_BULK",
        admin_id,
        details=f"Bulk overrides applied: {len(unique_pairs)}",
    )
    _clear_query_cache()
    return jsonify({"message": "Bulk eligibility overrides updated", "updated": len(unique_pairs)}), 200


# ─── Leave Requests (Feature 3) ──────────────────────────────────────────────

@admin_bp.route("/leave-requests", methods=["GET"])
@role_required("department_admin")
def list_leave_requests(user):
    """List all leave requests, optionally filtered by status or student."""
    leaves_col = get_collection("academic", "leave_requests")
    status_filter = request.args.get("status", "").strip()
    user_id    = request.args.get("user_id", "").strip()

    query = {}
    if status_filter:
        query["status"] = status_filter
    if user_id:
        query["user_id"] = user_id

    docs = list(leaves_col.find(query).sort("created_at", -1).limit(200))

    # Enrich with student and paper info
    user_ids = list({d.get("user_id") for d in docs if d.get("user_id")})
    user_map  = get_users_by_ids(user_ids)
    
    paper_ids = list({d.get("paper_id") for d in docs if d.get("paper_id")})
    papers_col = get_collection("academic", "papers")
    paper_map = {str(p["_id"]): p for p in papers_col.find({"_id": {"$in": [ObjectId(pid) for pid in paper_ids]}})}

    for d in docs:
        ud = user_map.get(d.get("user_id")) or {}
        d["student_name"]  = _as_text(ud.get("name") or "Unknown")
        d["student_email"] = _as_text(ud.get("email") or "")
        
        pid = d.get("paper_id")
        if pid:
            pd = paper_map.get(str(pid)) or {}
            d["paper_name"] = _as_text(pd.get("name") or "Unknown")
            d["paper_code"] = _as_text(pd.get("code") or "")

    return jsonify(sanitise_many(docs))


@admin_bp.route("/leave-requests/<leave_id>/approve", methods=["PUT"])
@role_required("department_admin")
def approve_leave_request(user, leave_id):
    """Approve a leave request (marks as approved; attendance team can exclude those dates)."""
    leaves_col = get_collection("academic", "leave_requests")
    doc = leaves_col.find_one({"_id": ObjectId(leave_id)})
    if not doc:
        return jsonify({"error": "Leave request not found"}), 404

    now = datetime.now(timezone.utc)
    leaves_col.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": {
            "status":      "approved",
            "reviewed_by": str(user["_id"]),
            "reviewed_at": now,
        }},
    )
    log_action(
        "LEAVE_REQUEST_APPROVED",
        str(user["_id"]),
        target_user=doc.get("user_id"),
        details={"leave_id": leave_id, "date": doc.get("date"), "paper_id": doc.get("paper_id")},
    )
    return jsonify({"message": "Leave request approved"}), 200


@admin_bp.route("/leave-requests/<leave_id>/reject", methods=["PUT"])
@role_required("department_admin")
def reject_leave_request(user, leave_id):
    """Reject a leave request with an optional reason."""
    leaves_col = get_collection("academic", "leave_requests")
    doc = leaves_col.find_one({"_id": ObjectId(leave_id)})
    if not doc:
        return jsonify({"error": "Leave request not found"}), 404

    d      = request.get_json(silent=True) or {}
    remark = _as_text(d.get("remark", ""))
    now    = datetime.now(timezone.utc)

    leaves_col.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": {
            "status":      "rejected",
            "remark":      remark,
            "reviewed_by": str(user["_id"]),
            "reviewed_at": now,
        }},
    )
    log_action(
        "LEAVE_REQUEST_REJECTED",
        str(user["_id"]),
        target_user=doc.get("user_id"),
        details={"leave_id": leave_id, "remark": remark},
    )
    return jsonify({"message": "Leave request rejected"}), 200


def _parse_iso_date(value):
    text = _as_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except Exception:
        return None


def _local_midnight_to_utc(local_midnight, tz_offset_minutes):
    if not isinstance(local_midnight, datetime):
        return None
    return local_midnight + timedelta(minutes=_to_int(tz_offset_minutes, 0))


def _build_attendance_matrix_payload(args):
    department_filter = _as_text(args.get("department", ""))
    course_id = _as_text(args.get("course_id", ""))
    academic_session = _normalise_year(args.get("academic_session", "")) or _normalise_year(args.get("academic_year", ""))
    semester_filter = _as_text(args.get("semester", ""))
    tz_offset_minutes = _to_int(args.get("tz_offset_minutes", 0), 0)

    def _to_local(dt):
        if not isinstance(dt, datetime):
            return None
        # Browser sends JS getTimezoneOffset() minutes, so local = utc - offset.
        return dt - timedelta(minutes=tz_offset_minutes)

    from_date = _parse_iso_date(args.get("from_date", ""))
    to_date = _parse_iso_date(args.get("to_date", ""))

    range_start_utc = None
    range_end_utc = None
    if from_date:
        range_start_utc = _local_midnight_to_utc(from_date, tz_offset_minutes)
    if to_date:
        to_local_exclusive = to_date + timedelta(days=1)
        range_end_utc = _local_midnight_to_utc(to_local_exclusive, tz_offset_minutes)

    courses = sanitise_many(get_all_courses(["name", "code", "status", "course_duration", "department"]))
    papers = sanitise_many(get_all_papers(["name", "code", "semester", "course_id"]))
    paper_map = {p["_id"]: p for p in papers}

    dept_course_ids = set()
    if department_filter:
        for c in courses:
            if _as_text(c.get("department", "")).lower() == department_filter.lower():
                dept_course_ids.add(_as_text(c.get("_id")))

    allowed_papers = []
    for paper in papers:
        pid = _as_text(paper.get("_id"))
        if not pid:
            continue
        if department_filter and _as_text(paper.get("course_id")) not in dept_course_ids:
            continue
        if course_id and _as_text(paper.get("course_id")) != course_id:
            continue
        if semester_filter and _as_text(paper.get("semester")) != semester_filter:
            continue
        allowed_papers.append(pid)

    allowed_paper_set = set(allowed_papers)

    profiles_col = get_collection("academic", "student_profiles")
    profiles = list(
        profiles_col.find(
            {},
            {
                "_id": 0,
                "user_id": 1,
                "course_id": 1,
                "academic_session": 1,
                "academic_year": 1,
                "year": 1,
                "current_semester": 1,
                "roll_number": 1,
                "reg_number": 1,
                "enrolled_papers": 1,
            },
        )
    )

    available_sessions = set()
    for profile in profiles:
        profile_course_id = _as_text(profile.get("course_id"))
        if department_filter and profile_course_id not in dept_course_ids:
            continue
        if course_id and profile_course_id != course_id:
            continue
        profile_session = _as_text(profile.get("academic_session") or profile.get("academic_year") or profile.get("year"))
        if profile_session:
            available_sessions.add(profile_session)

    candidate_students = []
    for profile in profiles:
        user_id = _as_text(profile.get("user_id"))
        if not user_id:
            continue

        stu_course_id = _as_text(profile.get("course_id"))
        if department_filter and stu_course_id not in dept_course_ids:
            continue
        if course_id and stu_course_id != course_id:
            continue

        stu_session = _as_text(profile.get("academic_session") or profile.get("academic_year") or profile.get("year"))
        if academic_session and stu_session != academic_session:
            continue

        enrolled = [_as_text(pid) for pid in (profile.get("enrolled_papers") or []) if _as_text(pid)]
        if allowed_paper_set:
            enrolled = [pid for pid in enrolled if pid in allowed_paper_set]

        if semester_filter:
            current_sem = _as_text(profile.get("current_semester"))
            has_semester_paper = any(_as_text((paper_map.get(pid) or {}).get("semester")) == semester_filter for pid in enrolled)
            if current_sem != semester_filter and not has_semester_paper:
                continue

        candidate_students.append(
            {
                "user_id": user_id,
                "roll_no": _as_text(profile.get("roll_number") or profile.get("reg_number")),
                "enrolled_papers": enrolled,
            }
        )

    user_map = get_users_by_ids([s["user_id"] for s in candidate_students])

    students = []
    for stu in candidate_students:
        user_doc = user_map.get(stu["user_id"]) or {}
        students.append(
            {
                "user_id": stu["user_id"],
                "roll_no": stu["roll_no"] or "N/A",
                "name": _as_text(user_doc.get("name", "Unknown")) or "Unknown",
                "enrolled_papers": stu["enrolled_papers"],
            }
        )

    user_ids = set(s["user_id"] for s in students)

    if allowed_paper_set:
        paper_filter_set = set(allowed_paper_set)
    else:
        paper_filter_set = set()
        for stu in students:
            for pid in stu.get("enrolled_papers", []):
                paper_filter_set.add(pid)

    session_query = {}
    if not paper_filter_set:
        session_docs = []
    else:
        paper_match_ids = []
        for pid in paper_filter_set:
            paper_match_ids.extend(_id_variants(pid))
        session_query["paper_id"] = {"$in": paper_match_ids}

        committed_range = {}
        if range_start_utc:
            committed_range["$gte"] = range_start_utc
        if range_end_utc:
            committed_range["$lt"] = range_end_utc
        if committed_range:
            session_query["committed_at"] = committed_range

        sessions_col = get_collection("attendance", "attendance_sessions")
        session_docs = list(
            sessions_col.find(
                session_query,
                {
                    "_id": 0,
                    "session_id": 1,
                    "paper_id": 1,
                    "user_ids": 1,
                    "committed_at": 1,
                    "last_updated_at": 1,
                    "period_number": 1,
                    "period": 1,
                },
            )
        )

    date_subject_sessions = {}
    for doc in session_docs:
        paper_id = _as_text(doc.get("paper_id"))
        if not paper_id:
            continue
        if paper_filter_set and paper_id not in paper_filter_set:
            continue

        dt = doc.get("committed_at") or doc.get("last_updated_at")
        if not isinstance(dt, datetime):
            continue
        local_dt = _to_local(dt)
        date_key = local_dt.strftime("%Y-%m-%d")

        paper = paper_map.get(paper_id) or {}
        subject_code = _as_text(paper.get("code") or paper.get("name") or "SUB")
        subject_name = _as_text(paper.get("name") or paper.get("code") or "Subject")
        period_number = _as_text(doc.get("period_number") or doc.get("period"))
        session_id = _as_text(doc.get("session_id"))
        timestamp_key = local_dt.strftime("%Y%m%d%H%M%S%f")

        # Keep every committed class distinct, even for same subject/date/period.
        subject_key = f"{paper_id}::{period_number or 'NA'}::{session_id or timestamp_key}"
        compound_key = f"{date_key}::{subject_key}"

        if compound_key not in date_subject_sessions:
            date_subject_sessions[compound_key] = {
                "date": date_key,
                "paper_id": paper_id,
                "subject_code": subject_code,
                "subject_name": subject_name,
                "period_number": period_number,
                "column_key": compound_key,
                "present_set": set(),
            }

        for sid in (doc.get("user_ids") or []):
            sid_text = _as_text(sid)
            if sid_text:
                date_subject_sessions[compound_key]["present_set"].add(sid_text)

    ordered_columns = sorted(
        date_subject_sessions.values(),
        key=lambda x: (
            x.get("date") or "",
            _to_int(x.get("period_number"), 0),
            x.get("subject_code") or "",
            x.get("paper_id") or "",
        ),
    )

    # Global sequence across all class slots in selected range.
    for idx, col in enumerate(ordered_columns, start=1):
        col["global_sequence"] = idx

    grouped_dates = {}
    for col in ordered_columns:
        d = col["date"]
        date_bucket = grouped_dates.setdefault(d, [])
        slot_index = len(date_bucket)
        if slot_index < 26:
            subject_slot = chr(ord("A") + slot_index)
        else:
            subject_slot = f"S{slot_index + 1}"

        date_bucket.append(
            {
                "column_key": col["column_key"],
                "paper_id": col["paper_id"],
                "subject_code": col["subject_code"],
                "subject_name": col["subject_name"],
                "subject_slot": subject_slot,
                "period_number": col["period_number"],
                "sequence_number": col.get("global_sequence"),
                "label": f"{col['subject_code']} ({col['period_number']})" if col["period_number"] else col["subject_code"],
            }
        )

    dates = [{"date": d, "subjects": grouped_dates[d]} for d in sorted(grouped_dates.keys())]

    students.sort(key=lambda x: (x.get("roll_no") or "", x.get("name") or ""))
    rows = []
    for stu in students:
        cell_map = {}
        attended_counter = 0
        for col in ordered_columns:
            present = stu["user_id"] in col["present_set"]
            if present:
                attended_counter += 1
                cell_map[col["column_key"]] = _as_text(attended_counter) or "1"
            else:
                cell_map[col["column_key"]] = "X"

        date_summary = {}
        for date_entry in dates:
            parts = []
            for sub in date_entry["subjects"]:
                parts.append(cell_map.get(sub["column_key"], "X"))
            date_summary[date_entry["date"]] = " : ".join(parts) if parts else ""

        rows.append(
            {
                "user_id": stu["user_id"],
                "roll_no": stu["roll_no"],
                "name": stu["name"],
                "cells": cell_map,
                "date_summary": date_summary,
            }
        )

    available_semesters = set()
    for paper in papers:
        if course_id and _as_text(paper.get("course_id")) != course_id:
            continue
        sem = _as_text(paper.get("semester"))
        if sem:
            available_semesters.add(sem)

    if not available_semesters and course_id:
        course_doc = next((c for c in courses if _as_text(c.get("_id")) == course_id), None)
        duration_years = _to_int((course_doc or {}).get("course_duration"), 0)
        if duration_years > 0:
            for sem in range(1, duration_years * 2 + 1):
                available_semesters.add(str(sem))

    options_courses = []
    for course in courses:
        options_courses.append(
            {
                "_id": _as_text(course.get("_id")),
                "name": _as_text(course.get("name")),
                "code": _as_text(course.get("code")),
                "status": _as_text(course.get("status") or "active").lower() or "active",
                "course_duration": _to_int(course.get("course_duration"), 0),
            }
        )

    options_courses.sort(key=lambda x: (x.get("name") or "", x.get("code") or ""))

    return {
        "filters": {
            "course_id": course_id or None,
            "academic_session": academic_session or None,
            "semester": semester_filter or None,
            "tz_offset_minutes": tz_offset_minutes,
            "from_date": _as_text(args.get("from_date", "")) or None,
            "to_date": _as_text(args.get("to_date", "")) or None,
        },
        "options": {
            "courses": options_courses,
            "academic_sessions": sorted(available_sessions),
            "semesters": sorted(available_semesters, key=lambda x: _to_int(x, 0)),
        },
        "meta": {
            "students_count": len(rows),
            "dates_count": len(dates),
            "sessions_count": len(ordered_columns),
            "subject_columns_count": len(ordered_columns),
        },
        "dates": dates,
        "rows": rows,
    }


@admin_bp.route("/attendance-matrix", methods=["GET"])
@role_required("department_admin")
def attendance_matrix(user):
    cache_key = (
        "attendance_matrix",
        tuple(sorted((k, _as_text(v)) for k, v in request.args.items())),
    )
    cached_payload = _cache_get(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    payload = _build_attendance_matrix_payload(request.args)
    _cache_set(cache_key, payload, _QUERY_CACHE_TTL_SECONDS)
    return jsonify(payload)


@admin_bp.route("/attendance-matrix/export", methods=["GET"])
@role_required("department_admin")
def attendance_matrix_export(user):
    payload = _build_attendance_matrix_payload(request.args)
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"error": "openpyxl is required for Excel export. Install it in backend requirements."}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Matrix"

    dates = payload.get("dates") or []
    rows = payload.get("rows") or []

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(row=1, column=1, value="Roll No")
    ws.cell(row=1, column=2, value="Name")

    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_idx = 3
    ordered_subjects = []
    for date_entry in dates:
        subjects = date_entry.get("subjects") or []
        if not subjects:
            continue

        start_col = col_idx
        for sub in subjects:
            ws.cell(
                row=2,
                column=col_idx,
                value=_as_text(sub.get("subject_code") or sub.get("subject_name") or "SUB"),
            )
            ordered_subjects.append(sub)
            col_idx += 1

        end_col = col_idx - 1
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col, value=_as_text(date_entry.get("date")) or "Date")

    total_attended_col = col_idx
    total_held_col = col_idx + 1
    percentage_col = col_idx + 2

    ws.merge_cells(start_row=1, start_column=total_attended_col, end_row=2, end_column=total_attended_col)
    ws.merge_cells(start_row=1, start_column=total_held_col, end_row=2, end_column=total_held_col)
    ws.merge_cells(start_row=1, start_column=percentage_col, end_row=2, end_column=percentage_col)

    ws.cell(row=1, column=total_attended_col, value="TCA")
    ws.cell(row=1, column=total_held_col, value="TCH")
    ws.cell(row=1, column=percentage_col, value="%")

    body_start_row = 3
    for i, row in enumerate(rows, start=body_start_row):
        ws.cell(row=i, column=1, value=row.get("roll_no"))
        ws.cell(row=i, column=2, value=row.get("name"))

        total_attended = 0
        total_held = 0
        for j, sub in enumerate(ordered_subjects, start=3):
            value = _as_text((row.get("cells") or {}).get(sub.get("column_key"), "X")) or "X"
            ws.cell(row=i, column=j, value=value)
            total_held += 1
            if value.upper() != "X":
                total_attended += 1

        percentage = round((total_attended / total_held) * 100, 2) if total_held > 0 else 0
        ws.cell(row=i, column=total_attended_col, value=total_attended)
        ws.cell(row=i, column=total_held_col, value=total_held)
        ws.cell(row=i, column=percentage_col, value=f"{percentage}%")

    max_col = max(2, percentage_col)
    max_row = max(2, body_start_row + len(rows) - 1)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center
            if r <= 2:
                cell.fill = header_fill
                cell.font = header_font

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for c in range(3, total_attended_col):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions[get_column_letter(total_attended_col)].width = 20
    ws.column_dimensions[get_column_letter(total_held_col)].width = 18
    ws.column_dimensions[get_column_letter(percentage_col)].width = 14

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"attendance_matrix_{india_timestamp_token()}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/attendance-matrix/export-csv", methods=["GET"])
@role_required("department_admin")
def attendance_matrix_export_csv(user):
    payload = _build_attendance_matrix_payload(request.args)
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    dates = payload.get("dates") or []
    rows = payload.get("rows") or []

    output = StringIO()
    writer = csv.writer(output)

    header_row_1 = ["Roll No", "Name"]
    header_row_2 = ["", ""]
    ordered_subjects = []

    for date_entry in dates:
        subjects = date_entry.get("subjects") or []
        if not subjects:
            continue
        for _ in subjects:
            header_row_1.append(date_entry.get("date"))
        for sub in subjects:
            header_row_2.append(sub.get("subject_slot") or sub.get("subject_code") or sub.get("label") or "SUB")
            ordered_subjects.append(sub)

    writer.writerow(header_row_1)
    writer.writerow(header_row_2)

    for row in rows:
        line = [row.get("roll_no"), row.get("name")]
        for sub in ordered_subjects:
            line.append((row.get("cells") or {}).get(sub.get("column_key"), "X"))
        writer.writerow(line)

    csv_bytes = output.getvalue().encode("utf-8-sig")
    stream = BytesIO(csv_bytes)
    stream.seek(0)
    filename = f"attendance_matrix_{india_timestamp_token()}.csv"

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv",
    )


@admin_bp.route("/attendance-matrix/export-pdf", methods=["GET"])
@role_required("department_admin")
def attendance_matrix_export_pdf(user):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    payload = _build_attendance_matrix_payload(request.args)
    dates = payload.get("dates") or []
    rows = payload.get("rows") or []

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Official Attendance Matrix Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))

    header_row = ["Roll No", "Name"]
    # For a simple PDF, we'll flatten the columns
    ordered_subjects = []
    for date_entry in dates:
        subjects = date_entry.get("subjects") or []
        for sub in subjects:
            short_title = f"{date_entry.get('date')[5:]}\n{sub.get('subject_slot') or sub.get('subject_code')}"
            header_row.append(short_title)
            ordered_subjects.append(sub)

    header_row.extend(["TCA", "TCH", "%"])
    table_data = [header_row]

    for row in rows:
        line = [row.get("roll_no"), row.get("name")[:15] + ".." if len(row.get("name")) > 15 else row.get("name")]
        total_attended = 0
        total_held = 0
        for sub in ordered_subjects:
            val = (row.get("cells") or {}).get(sub.get("column_key"), "X")
            line.append(val)
            total_held += 1
            if val.upper() != "X":
                total_attended += 1
        
        percentage = round((total_attended / total_held) * 100, 2) if total_held > 0 else 0
        line.extend([str(total_attended), str(total_held), f"{percentage}%"])
        table_data.append(line)

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(t)
    
    elements.append(Spacer(1, 40))
    elements.append(Paragraph(f"Generated at: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Paragraph("Authorized Administrator Signature: _______________________", styles['Normal']))
    
    doc.build(elements)
    pdf_bytes = output.getvalue()
    stream = BytesIO(pdf_bytes)
    stream.seek(0)
    filename = f"attendance_matrix_{india_timestamp_token()}.pdf"

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )



# ─── Dashboard Stats ────────────────────────────────────────────────────────

@admin_bp.route("/stats", methods=["GET"])
@role_required("super_admin", "department_admin")
def dashboard_stats(user):
    dept_filter_id = _as_text(request.args.get("department_id", ""))
    
    user_dept = None
    if is_super_admin(user):
        user_dept = dept_filter_id or None
    else:
        user_dept = _user_dept_id(user)
        
    cache_key = ("dashboard_stats", user_dept)
    cached_payload = _cache_get(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    started_at = current_app.config.get("APP_STARTED_AT")
    uptime_seconds = int((datetime.now(timezone.utc) - started_at).total_seconds()) if started_at else 0
    uptime_days, remainder = divmod(max(uptime_seconds, 0), 86400)
    uptime_hours, remainder = divmod(remainder, 3600)
    uptime_minutes, uptime_seconds = divmod(remainder, 60)

    uptime_parts = []
    if uptime_days:
        uptime_parts.append(f"{uptime_days}d")
    if uptime_hours or uptime_parts:
        uptime_parts.append(f"{uptime_hours}h")
    uptime_parts.append(f"{uptime_minutes}m")
    system_uptime = " ".join(uptime_parts)

    profiles_col = get_collection("academic", "student_profiles")
    users_col = get_collection("auth", "users")
    courses_col = get_collection("academic", "courses")
    papers_col = get_collection("academic", "papers")
    attendance_col = get_collection("attendance", "attendance_logs")
    audit_col = get_collection("audit", "audit_logs")
    by_course = {}
    by_year = {}
    by_department = {}
    courses = sanitise_many(get_all_courses(["name", "code", "year", "status", "department"], department_id=user_dept))
    course_map = {c["_id"]: c for c in courses}
    course_ids = list(course_map.keys())

    active_course_ids = {
        c.get("_id")
        for c in courses
        if _as_text(c.get("status") or "active").lower() == "active"
    }

    # course_ids from sanitise_many are strings; profiles may store course_id as ObjectId or string
    course_ids_with_oids = []
    for cid in course_ids:
        course_ids_with_oids.append(cid)  # string
        if ObjectId.is_valid(cid):
            course_ids_with_oids.append(ObjectId(cid))  # ObjectId variant

    # Scope profiles to department's courses when filtering
    profile_query = {"course_id": {"$in": course_ids_with_oids}} if (user_dept and course_ids_with_oids) else {}
    profiles = list(
        profiles_col.find(
            profile_query,
            {
                "course_id": 1,
                "user_id": 1,
                "academic_session": 1,
                "academic_year": 1,
                "year": 1,
            },
        )
    )

    for profile in profiles:
        cid = _as_text(profile.get("course_id"))
        if cid and cid not in active_course_ids:
            continue

        course = course_map.get(cid)
        course_key = course.get("name") if course else "Unassigned"
        by_course[course_key] = by_course.get(course_key, 0) + 1

        dept_key = _as_text(course.get("department")) if course else ""
        if dept_key:
            if dept_key not in by_department:
                by_department[dept_key] = {"students": 0, "lecturers": 0}
            by_department[dept_key]["students"] += 1

        year_key = _normalise_year(
            profile.get("academic_session")
            or profile.get("academic_year")
            or profile.get("year")
        ) or "Unknown"
        by_year[year_key] = by_year.get(year_key, 0) + 1

    student_query = {"role": "student"}
    lecturer_query = {"role": "lecturer"}
    
    if user_dept:
        student_user_ids = [_as_text(p.get("user_id")) for p in profiles if p.get("user_id")]
        valid_oids = [ObjectId(u) for u in student_user_ids if ObjectId.is_valid(u)]
        student_query["$or"] = [
            {"_id": {"$in": valid_oids}},
            {"_id": {"$in": student_user_ids}},
        ]
        lecturer_query["department_id"] = user_dept

    for usr in users_col.find(lecturer_query, {"department": 1}):
        dept_key = _as_text(usr.get("department"))
        if dept_key:
            if dept_key not in by_department:
                by_department[dept_key] = {"students": 0, "lecturers": 0}
            by_department[dept_key]["lecturers"] += 1

    # Count queries - department_id must be sibling to $or, not inside it
    if user_dept:
        active_courses_count = courses_col.count_documents({
            "department_id": user_dept,
            "$or": [{"status": "active"}, {"status": {"$exists": False}}, {"status": ""}, {"status": None}],
        })
        inactive_courses_count = courses_col.count_documents({"department_id": user_dept, "status": "inactive"})
    else:
        active_courses_count = courses_col.count_documents({
            "$or": [{"status": "active"}, {"status": {"$exists": False}}, {"status": ""}, {"status": None}],
        })
        inactive_courses_count = courses_col.count_documents({"status": "inactive"})
    total_courses_count = active_courses_count + inactive_courses_count

    active_paper_count = 0
    inactive_paper_count = 0
    
    paper_query = {}
    if user_dept:
        paper_query["course_id"] = {"$in": course_ids}

    all_paper_ids = []
    for paper in papers_col.find(paper_query, {"course_id": 1}):
        if paper.get("_id"):
            all_paper_ids.append(paper["_id"])
            all_paper_ids.append(str(paper["_id"]))
        paper_course_id = _as_text(paper.get("course_id"))
        if not paper_course_id or paper_course_id in active_course_ids:
            active_paper_count += 1
        else:
            inactive_paper_count += 1

    audit_count = 0
    if not user_dept:
        audit_count = audit_col.count_documents({})
    else:
        dept_u_ids = [str(u["_id"]) for u in users_col.find({"department_id": user_dept}, {"_id": 1})]
        dept_oids = [ObjectId(u) for u in dept_u_ids if ObjectId.is_valid(u)]
        all_uid_variants = dept_u_ids + dept_oids
        audit_count = audit_col.count_documents({
            "$or": [
                {"performed_by": {"$in": all_uid_variants}},
                {"target_user": {"$in": all_uid_variants}},
                {"department_id": user_dept},
            ]
        })
        
    attendance_count = attendance_col.count_documents({"paper_id": {"$in": all_paper_ids}} if user_dept and all_paper_ids else {}) if not (user_dept and not all_paper_ids) else 0

    app_started_at = None
    if started_at:
        iso_started_at = started_at.isoformat()
        tz_part = iso_started_at[10:]
        has_tz = iso_started_at.endswith("Z") or "+" in tz_part or "-" in tz_part
        app_started_at = iso_started_at if has_tz else f"{iso_started_at}Z"

    payload = {
        "total_students": users_col.count_documents(student_query),
        "total_lecturers": users_col.count_documents(lecturer_query),
        "total_courses": total_courses_count,
        "active_courses": active_courses_count,
        "inactive_courses": inactive_courses_count,
        "total_papers": active_paper_count,
        "inactive_papers": inactive_paper_count,
        "total_attendance": attendance_count,
        "total_audit_logs": audit_count,
        "app_started_at": app_started_at,
        "system_uptime_seconds": max(int((datetime.now(timezone.utc) - started_at).total_seconds()), 0) if started_at else 0,
        "system_uptime": system_uptime,
        "students_by_course": by_course,
        "students_by_year": by_year,
        "departments_summary": by_department,
    }
    _cache_set(cache_key, payload, _QUERY_CACHE_TTL_SECONDS)
    return jsonify(payload)


@admin_bp.route("/stats/monthly-attendance", methods=["GET"])
@role_required("super_admin", "department_admin")
def monthly_attendance_trend_api(user):
    # Resolve which department_id to scope to
    if is_super_admin(user):
        # Super admin: optionally filter by department name param
        dept_name_filter = _as_text(request.args.get("department", ""))
        if dept_name_filter:
            dept_doc = get_collection("academic", "departments").find_one(
                {"name": {"$regex": f"^{dept_name_filter}$", "$options": "i"}}
            )
            scope_dept_id = dept_doc.get("_id") if dept_doc else None
        else:
            scope_dept_id = None  # No filter = global
    else:
        # Department admin: always locked to their own department_id
        scope_dept_id = _user_dept_id(user)

    attendance_col = get_collection("attendance", "attendance_logs")
    query = {}

    if scope_dept_id:
        # Filter by department_id (ObjectId) directly — reliable, no name mismatch
        dept_oid = ObjectId(str(scope_dept_id)) if not isinstance(scope_dept_id, ObjectId) else scope_dept_id
        dept_id_variants = [dept_oid, str(dept_oid)]

        courses = list(get_collection("academic", "courses").find(
            {"department_id": {"$in": dept_id_variants}}, {"_id": 1}
        ))
        course_ids = []
        for c in courses:
            course_ids.append(c["_id"])
            course_ids.append(str(c["_id"]))

        if not course_ids:
            # No courses for this department → return all-zero trend
            def _ms(dt): return datetime(dt.year, dt.month, 1)
            def _sm(dt, d):
                y = dt.year + ((dt.month - 1 + d) // 12)
                m = ((dt.month - 1 + d) % 12) + 1
                return datetime(y, m, 1)
            now = datetime.now(timezone.utc)
            sm = _sm(_ms(now), -5)
            return jsonify([
                {"key": f"{_sm(sm,i).year}-{_sm(sm,i).month:02d}",
                 "label": _sm(sm, i).strftime("%b"), "total": 0}
                for i in range(6)
            ])

        papers = list(get_collection("academic", "papers").find(
            {"course_id": {"$in": course_ids}}, {"_id": 1}
        ))
        paper_ids = []
        for p in papers:
            paper_ids.append(p["_id"])
            paper_ids.append(str(p["_id"]))

        if not paper_ids:
            # No papers → zero trend
            def _ms(dt): return datetime(dt.year, dt.month, 1)
            def _sm(dt, d):
                y = dt.year + ((dt.month - 1 + d) // 12)
                m = ((dt.month - 1 + d) % 12) + 1
                return datetime(y, m, 1)
            now = datetime.now(timezone.utc)
            sm = _sm(_ms(now), -5)
            return jsonify([
                {"key": f"{_sm(sm,i).year}-{_sm(sm,i).month:02d}",
                 "label": _sm(sm, i).strftime("%b"), "total": 0}
                for i in range(6)
            ])

        query["paper_id"] = {"$in": paper_ids}

    def _month_start(dt):
        return datetime(dt.year, dt.month, 1)

    def _shift_month(dt, delta):
        year = dt.year + ((dt.month - 1 + delta) // 12)
        month = ((dt.month - 1 + delta) % 12) + 1
        return datetime(year, month, 1)

    now = datetime.now(timezone.utc)
    current_month = _month_start(now)
    start_month = _shift_month(current_month, -5)

    query["timestamp"] = {"$gte": start_month}

    docs = attendance_col.find(query, {"timestamp": 1})
    count_map = {}
    for doc in docs:
        ts = doc.get("timestamp")
        if isinstance(ts, str):
            try:
                ts = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            except Exception:
                continue
        if not isinstance(ts, datetime):
            continue
        key = f"{ts.year}-{ts.month:02d}"
        count_map[key] = count_map.get(key, 0) + 1

    points = []
    for i in range(6):
        month_dt = _shift_month(start_month, i)
        key = f"{month_dt.year}-{month_dt.month:02d}"
        points.append({
            "key": key,
            "label": month_dt.strftime("%b"),
            "total": count_map.get(key, 0),
        })

    return jsonify(points)


@admin_bp.route("/attendance/send-shortage-alerts", methods=["POST"])
@role_required("department_admin")
def send_shortage_alerts(user):
    """Scan all students and send shortage alert emails to those < 75% attendance."""
    if not is_email_delivery_enabled():
        return jsonify({"error": "Email delivery is not configured on this server."}), 503

    d = request.get_json(silent=True) or {}
    course_id = d.get("course_id")
    paper_id_filter = d.get("paper_id")

    profiles_col = get_collection("academic", "student_profiles")
    sessions_col = get_collection("attendance", "attendance_sessions")
    
    query = {}
    if course_id:
        query["course_id"] = course_id

    profiles = list(profiles_col.find(query))
    if not profiles:
        return jsonify({"message": "No students found to check."}), 200

    alerts_sent = 0
    checked_count = 0

    threshold = float(current_app.config.get("ATTENDANCE_THRESHOLD", 75.0))
    threshold_dec = threshold / 100.0

    for profile in profiles:
        uid = str(profile.get("user_id"))
        enrolled_papers = profile.get("enrolled_papers", [])
        
        # Filter papers if requested
        if paper_id_filter:
            enrolled_papers = [p for p in enrolled_papers if str(p) == str(paper_id_filter)]
            
        if not enrolled_papers:
            continue
            
        checked_count += 1
        leave_map = get_approved_leave_dates(uid, enrolled_papers)
        user_doc = find_user_by_id(uid)
        if not user_doc or not user_doc.get("email"):
            continue

        for paper_id in enrolled_papers:
            paper_id_text = str(paper_id)
            paper = get_paper_by_id(paper_id_text)
            if not paper:
                continue

            # Fetch committed sessions for this paper
            paper_id_variants = [paper_id_text]
            try:
                paper_id_variants.append(ObjectId(paper_id_text))
            except Exception:
                pass
                
            committed_sessions = list(
                sessions_col.find(
                    {"paper_id": {"$in": paper_id_variants}},
                    {"user_ids": 1, "committed_at": 1, "last_updated_at": 1},
                )
            )
            
            if not committed_sessions:
                continue

            paper_leave_dates = leave_map.get(paper_id_text, set())
            attended = 0
            effective_total = 0
            for sess in committed_sessions:
                sess_date = session_date_str(sess)
                if sess_date and sess_date in paper_leave_dates:
                    continue
                effective_total += 1
                if uid in [str(sid) for sid in (sess.get("user_ids") or [])]:
                    attended += 1
            
            if effective_total == 0:
                continue
                
            pct = round((attended / effective_total) * 100, 2)
            if pct < threshold:
                # Calculate classes needed: (A + n) / (T + n) >= threshold_dec  =>  n >= (threshold_dec*T - A) / (1 - threshold_dec)
                divider = (1.0 - threshold_dec)
                needed_float = ((threshold_dec * effective_total) - attended) / divider if divider > 0 else 0
                classes_needed = max(0, int(needed_float) if needed_float.is_integer() else int(needed_float) + 1)
                
                send_shortage_alert_email(
                    to_email=user_doc["email"],
                    name=user_doc["name"],
                    paper_name=paper.get("name", "Unknown Paper"),
                    percentage=pct,
                    classes_needed=classes_needed
                )
                alerts_sent += 1

    return jsonify({
        "message": f"Alert scan completed. {checked_count} students checked.",
        "alerts_queued": alerts_sent
    })


# ─── Department Management (Super Admin only) ─────────────────────────────


@admin_bp.route("/departments", methods=["GET"])
@super_admin_required
def list_departments(user):
    """Return all departments."""
    include_inactive = _as_text(request.args.get("include_inactive", "")).lower() in ("1", "true", "yes")
    depts = get_all_departments(include_inactive=include_inactive)

    # Enrich with user counts
    users_col = get_collection("auth", "users")
    for dept in depts:
        dept["_id"] = str(dept["_id"])
        dept_oid = ObjectId(dept["_id"]) if dept["_id"] else None
        dept["admin_count"] = users_col.count_documents({"role": "department_admin", "department_id": dept_oid}) if dept_oid else 0
        dept["lecturer_count"] = users_col.count_documents({"role": "lecturer", "department_id": dept_oid}) if dept_oid else 0
        dept["student_count"] = users_col.count_documents({"role": "student", "department_id": dept_oid}) if dept_oid else 0
        dept["created_at"] = str(dept.get("created_at") or "")
        dept["updated_at"] = str(dept.get("updated_at") or "")

    return jsonify(depts)


@admin_bp.route("/departments", methods=["POST"])
@super_admin_required
def add_department(user):
    """Create a new department."""
    d = request.get_json(silent=True) or {}
    name = _as_text(d.get("name", "")).strip()
    code = _as_text(d.get("code", "")).strip().upper()

    if not name or not code:
        return jsonify({"error": "name and code are required"}), 400

    # Check uniqueness
    if get_department_by_code(code):
        return jsonify({"error": f"Department code '{code}' already exists"}), 409

    dept = create_department(name, code)
    log_action(
        "CREATE_DEPARTMENT",
        str(user["_id"]),
        details=f"Department {code} — {name}",
    )
    return jsonify(sanitise_mongo_doc(dept)), 201


@admin_bp.route("/departments/<dept_id>", methods=["GET"])
@super_admin_required
@validate_ids("dept_id")
def get_department(user, dept_id):
    """Return a single department."""
    dept = get_department_by_id(dept_id)
    if not dept:
        return jsonify({"error": "Department not found"}), 404
    return jsonify(sanitise_mongo_doc(dept))


@admin_bp.route("/departments/<dept_id>", methods=["PUT"])
@super_admin_required
@validate_ids("dept_id")
def edit_department(user, dept_id):
    """Update a department's name, code, or status."""
    d = request.get_json(silent=True) or {}
    fields = {}
    if "name" in d:
        fields["name"] = _as_text(d["name"]).strip()
    if "code" in d:
        new_code = _as_text(d["code"]).strip().upper()
        existing = get_department_by_code(new_code)
        if existing and str(existing["_id"]) != dept_id:
            return jsonify({"error": f"Department code '{new_code}' already in use"}), 409
        fields["code"] = new_code
    if "status" in d and d["status"] in ("active", "inactive"):
        fields["status"] = d["status"]

    if not fields:
        return jsonify({"error": "No valid fields to update"}), 400

    updated = update_department(dept_id, fields)
    if not updated:
        return jsonify({"error": "Department not found"}), 404

    log_action(
        "UPDATE_DEPARTMENT",
        str(user["_id"]),
        details=f"Department {dept_id}: {fields}",
    )
    return jsonify(sanitise_mongo_doc(updated))


@admin_bp.route("/departments/<dept_id>", methods=["DELETE"])
@super_admin_required
@validate_ids("dept_id")
def remove_department(user, dept_id):
    """Soft-delete a department (set status=inactive)."""
    dept = get_department_by_id(dept_id)
    if not dept:
        return jsonify({"error": "Department not found"}), 404

    # Prevent deletion if department has active users
    users_col = get_collection("auth", "users")
    active_users = users_col.count_documents({"department_id": ObjectId(dept_id)})
    if active_users > 0:
        return jsonify({
            "error": f"Cannot delete department with {active_users} active users. Reassign them first."
        }), 409

    soft_delete_department(dept_id)
    log_action(
        "DELETE_DEPARTMENT",
        str(user["_id"]),
        details=f"Department {dept.get('code')} deactivated",
    )
    return jsonify({"message": "Department deactivated successfully"})


# ─── Department Admin Management (Super Admin only) ────────────────────────


@admin_bp.route("/department-admins", methods=["GET"])
@super_admin_required
def list_department_admins(user):
    """List all department admin users with their department info."""
    dept_admins = sanitise_many(get_users_by_role("department_admin"))

    # Enrich with department info
    for admin in dept_admins:
        dept_id = admin.get("department_id")
        if dept_id:
            dept = get_department_by_id(str(dept_id))
            admin["department_name"] = dept.get("name") if dept else "Unknown"
            admin["department_code"] = dept.get("code") if dept else "?"
        else:
            admin["department_name"] = "Unassigned"
            admin["department_code"] = "—"
        # Remove sensitive fields
        admin.pop("password_hash", None)
        admin.pop("session_version", None)

    return jsonify(dept_admins)


@admin_bp.route("/department-admins", methods=["POST"])
@super_admin_required
def add_department_admin(user):
    """Create a new department admin user."""
    d = request.get_json(silent=True) or {}
    name = _as_text(d.get("name", "")).strip()
    email = _as_text(d.get("email", "")).strip().lower()
    department_id = _as_text(d.get("department_id", "")).strip()
    initial_password = _as_text(d.get("initial_password", "")).strip()

    if not name or not email or not department_id:
        return jsonify({"error": "name, email, and department_id are required"}), 400

    # Validate department exists
    dept = get_department_by_id(department_id)
    if not dept:
        return jsonify({"error": "Department not found"}), 404

    # Check email uniqueness
    if find_user_by_email(email):
        return jsonify({"error": "A user with this email already exists"}), 409

    # Generate temp password if not provided
    if not initial_password:
        initial_password = f"DeptAdmin{secrets.randbelow(90000) + 10000}!"

    # Validate password strength
    is_strong, pw_error = validate_password_strength(initial_password)
    if not is_strong:
        return jsonify({"error": pw_error}), 400

    new_admin = create_user(
        name=name,
        email=email,
        password=initial_password,
        role="department_admin",
        department=dept.get("name", ""),
        department_id=department_id,
        must_change_password=True,
    )

    log_action(
        "CREATE_DEPARTMENT_ADMIN",
        str(user["_id"]),
        target_user=str(new_admin.get("_id", "")),
        details=f"Dept admin {email} for {dept.get('code')}",
    )

    return jsonify({
        "message": f"Department admin created. Temp password: {initial_password}",
        "user": sanitise_mongo_doc(new_admin),
        "temp_password": initial_password,
    }), 201


@admin_bp.route("/department-admins/<uid>", methods=["PUT"])
@super_admin_required
@validate_ids("uid")
def edit_department_admin(user, uid):
    """Update a department admin's basic info or reassign their department."""
    d = request.get_json(silent=True) or {}
    target = find_user_by_id(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404
    if target.get("role") != "department_admin":
        return jsonify({"error": "User is not a department admin"}), 400

    update_fields = {}
    if "name" in d:
        update_fields["name"] = _as_text(d["name"]).strip()
    if "department_id" in d:
        new_dept_id = _as_text(d["department_id"]).strip()
        dept = get_department_by_id(new_dept_id)
        if not dept:
            return jsonify({"error": "Target department not found"}), 404
        update_fields["department_id"] = ObjectId(new_dept_id)
        update_fields["department"] = dept.get("name", "")

    if not update_fields:
        return jsonify({"error": "No valid fields to update"}), 400

    users_col = get_collection("auth", "users")
    users_col.update_one({"_id": ObjectId(uid)}, {"$set": update_fields})

    log_action(
        "UPDATE_DEPARTMENT_ADMIN",
        str(user["_id"]),
        target_user=uid,
        details=f"Updated dept admin {uid}: {update_fields}",
    )
    updated = find_user_by_id(uid)
    return jsonify(sanitise_mongo_doc(updated))


@admin_bp.route("/department-admins/<uid>", methods=["DELETE"])
@super_admin_required
@validate_ids("uid")
def remove_department_admin(user, uid):
    """Delete a department admin user."""
    target = find_user_by_id(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404
    if target.get("role") != "department_admin":
        return jsonify({"error": "User is not a department admin"}), 400

    delete_user(uid)
    log_action(
        "DELETE_DEPARTMENT_ADMIN",
        str(user["_id"]),
        target_user=uid,
        details=f"Deleted dept admin {target.get('email')}",
    )
    return jsonify({"message": "Department admin deleted"})


@admin_bp.route("/department-admins/<uid>/reset-password", methods=["POST"])
@super_admin_required
@validate_ids("uid")
def reset_department_admin_password(user, uid):
    """Reset a department admin's password."""
    target = find_user_by_id(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404
    if target.get("role") != "department_admin":
        return jsonify({"error": "User is not a department admin"}), 400

    d = request.get_json(silent=True) or {}
    temp_password = reset_user_password(uid, temp_password=_as_text(d.get("temp_password", "")).strip() or None)

    log_action(
        "RESET_PASSWORD",
        str(user["_id"]),
        target_user=uid,
        details=f"Password reset for dept admin {target.get('email')}",
    )
    return jsonify({"message": "Password reset", "temp_password": temp_password})
