from . import admin_bp
from ._helpers import *

@admin_bp.route("/courses/reassign", methods=["POST"])
@role_required("department_admin")
def reassign_course_entities(user):
    """Batch move students/papers from one course to another active course."""
    d = request.get_json(silent=True) or {}
    from_course_id = _as_text(d.get("from_course_id"))
    to_course_id = _as_text(d.get("to_course_id"))
    move_students = _to_bool(d.get("move_students", True))
    move_papers = _to_bool(d.get("move_papers", True))

    if not from_course_id or not to_course_id:
        return jsonify({"error": "from_course_id and to_course_id are required"}), 400
    if from_course_id == to_course_id:
        return jsonify({"error": "from_course_id and to_course_id must be different"}), 400

    source_course = _safe_get_course(from_course_id)
    if not source_course:
        return jsonify({"error": "Source course not found"}), 404

    target_course, target_error = _get_active_course_or_error(to_course_id)
    if target_error:
        return target_error

    profile_col = get_collection("academic", "student_profiles")
    paper_col = get_collection("academic", "papers")

    moved_students = 0
    moved_papers = 0
    if move_students:
        res = profile_col.update_many({"course_id": from_course_id}, {"$set": {"course_id": to_course_id}})
        moved_students = int(res.modified_count)

    if move_papers:
        res = paper_col.update_many({"course_id": from_course_id}, {"$set": {"course_id": to_course_id}})
        moved_papers = int(res.modified_count)

    log_action(
        "REASSIGN_COURSE_ENTITIES",
        str(user["_id"]),
        details=(
            f"from={from_course_id}, to={to_course_id}, "
            f"move_students={move_students}, move_papers={move_papers}, "
            f"moved_students={moved_students}, moved_papers={moved_papers}"
        ),
    )
    _clear_query_cache()

    return jsonify(
        {
            "message": "Reassignment completed",
            "from_course": sanitise_mongo_doc(source_course),
            "to_course": sanitise_mongo_doc(target_course),
            "moved_students": moved_students,
            "moved_papers": moved_papers,
        }
    ), 200


# ─── Audit Trail ────────────────────────────────────────────────────────────

# Merge additional high-level audit exclusions with the biometric read
# exclusions already defined in _helpers.py (imported via wildcard).
_AUDIT_EXCLUDED_ACTIONS = list(set(
    _AUDIT_EXCLUDED_ACTIONS + ["HEARTBEAT", "QUEUE_CHECK", "STATUS_CHECK"]
))

@admin_bp.route("/audit-logs", methods=["GET"])
@role_required("department_admin")
def list_audit_logs(user):
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    action = _as_text(request.args.get("action", "")).upper()
    date_from = _as_text(request.args.get("from", ""))
    date_to = _as_text(request.args.get("to", ""))
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    filters = {}
    if action:
        # Contains match allows flexible keyword search like OVERRIDE, CREATE, DELETE, etc.
        filters["action"] = {"$regex": re.escape(action), "$options": "i"}

    ts_filter = {}
    parsed_from = _parse_iso_date(date_from)
    parsed_to = _parse_iso_date(date_to)

    if parsed_from:
        ts_filter["$gte"] = _local_midnight_to_utc(parsed_from, tz_offset_minutes)
    if parsed_to:
        parsed_to_exclusive = parsed_to + timedelta(days=1)
        ts_filter["$lt"] = _local_midnight_to_utc(parsed_to_exclusive, tz_offset_minutes)

    if ts_filter:
        filters["timestamp"] = ts_filter

    if filters:
        filters = {
            "$and": [
                filters,
                {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}},
            ]
        }
    else:
        filters = {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}}

    dept_id_param = _as_text(request.args.get("department_id", ""))

    if is_super_admin(user):
        # Super admin: optionally scope to a department by ID
        dept_filter_id = None
        if dept_id_param:
            dept_filter_id = _to_oid(dept_id_param)
        if dept_filter_id:
            # Get all user IDs in this department to filter logs
            users_col = get_collection("auth", "users")
            all_user_id_variants = []
            for u in users_col.find({"department_id": dept_filter_id}, {"_id": 1}):
                all_user_id_variants.extend(_id_variants(u["_id"]))
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": dept_filter_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter
        logs, total = get_audit_logs(page, per_page, filters, department_id=None)
    else:
        # Department admin: locked to their own department — fetch dept user IDs
        user_dept_id = _user_dept_id(user)
        if user_dept_id:
            users_col = get_collection("auth", "users")
            all_user_id_variants = []
            for u in users_col.find({"department_id": user_dept_id}, {"_id": 1}):
                all_user_id_variants.extend(_id_variants(u["_id"]))
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": user_dept_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter
        logs, total = get_audit_logs(page, per_page, filters, department_id=None)
    audit_user_ids = [
        item.get("performed_by") or item.get("actor_user_id")
        for item in logs
        if item.get("performed_by") or item.get("actor_user_id")
    ] + [
        item.get("target_user")
        or ((item.get("details") or {}).get("user_id") if isinstance(item.get("details"), dict) else None)
        for item in logs
        if item.get("target_user")
        or (isinstance(item.get("details"), dict) and (item.get("details") or {}).get("user_id"))
    ]
    user_map = get_users_by_ids(audit_user_ids)

    enriched = []
    for raw in logs:
        # We start with a copy for serialisation
        item = sanitise_mongo_doc(raw)

        # Re-fetch raw versions for local logic that requires datetime objects
        raw_ts = raw.get("timestamp")
        raw_rollback_until = raw.get("rollback_until")
        raw_rollback_payload = raw.get("rollback")
        raw_rolled_back = bool(raw.get("rolled_back"))

        actor_id = raw.get("performed_by") or raw.get("actor_user_id")
        details_user_id = (raw.get("details") or {}).get("user_id") if isinstance(raw.get("details"), dict) else None
        target_user_id = raw.get("target_user") or details_user_id

        actor = user_map.get(_as_text(actor_id)) if actor_id else None
        target_user = user_map.get(_as_text(target_user_id)) if target_user_id else None

        item["actor_name"] = (actor or {}).get("name") or ("System" if str(actor_id).lower() == "system" else "Unknown User")
        item["actor_email"] = (actor or {}).get("email") or ""
        item["role"] = (actor or {}).get("role") or raw.get("role") or "unknown"

        if target_user:
            item["target_type"] = f"{target_user.get('name', 'Unknown')} ({target_user.get('role', 'user')})"
            item["target_user_name"] = target_user.get("name")
            item["target_user_email"] = target_user.get("email")
            item["target_user_role"] = target_user.get("role")
        elif target_user_id:
            item["target_type"] = f"User {target_user_id}"
        else:
            item["target_type"] = _as_text(raw.get("details")) or "System"

        item["ip"] = raw.get("ip") or raw.get("ip_address") or ""

        # Handle time-sensitive rollback logic on raw datetime objects
        if raw_ts and raw_ts.tzinfo is None:
            raw_ts = raw_ts.replace(tzinfo=timezone.utc)
        
        if raw_rollback_payload and not raw_rollback_until and raw_ts:
            raw_rollback_until = raw_ts + timedelta(days=1)
        
        if raw_rollback_until and raw_rollback_until.tzinfo is None:
            raw_rollback_until = raw_rollback_until.replace(tzinfo=timezone.utc)

        now = datetime.now(timezone.utc)
        eligible = bool(raw_rollback_payload) and not raw_rolled_back and bool(raw_rollback_until) and now <= raw_rollback_until

        item["rollback_available"] = eligible
        item["rolled_back"] = raw_rolled_back
        if raw_rollback_until:
            item["rollback_until"] = raw_rollback_until.isoformat()

        # Ensure IDs are stringified
        item["performed_by"] = _as_text(actor_id) if actor_id else None
        item["target_user"] = _as_text(target_user_id) if target_user_id else None

        item.pop("rollback", None)
        enriched.append(item)

    return jsonify({
        "logs": enriched,
        "total": total,
        "page": page,
        "per_page": per_page,
    })


@admin_bp.route("/audit-logs/export", methods=["GET"])
@role_required("department_admin")
def export_audit_logs(user):
    """Export the filtered audit logs to an Excel file."""
    action = _as_text(request.args.get("action", "")).upper()
    date_from = _as_text(request.args.get("from", ""))
    date_to = _as_text(request.args.get("to", ""))
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    filters = {}
    if action:
        filters["action"] = {"$regex": re.escape(action), "$options": "i"}

    ts_filter = {}
    parsed_from = _parse_iso_date(date_from)
    parsed_to = _parse_iso_date(date_to)

    if parsed_from:
        ts_filter["$gte"] = _local_midnight_to_utc(parsed_from, tz_offset_minutes)
    if parsed_to:
        parsed_to_exclusive = parsed_to + timedelta(days=1)
        ts_filter["$lt"] = _local_midnight_to_utc(parsed_to_exclusive, tz_offset_minutes)

    if ts_filter:
        filters["timestamp"] = ts_filter

    if filters:
        filters = {
            "$and": [
                filters,
                {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}},
            ]
        }
    else:
        filters = {"action": {"$nin": _AUDIT_EXCLUDED_ACTIONS}}

    dept_id_param = _as_text(request.args.get("department_id", ""))

    if is_super_admin(user):
        dept_filter_id = None
        if dept_id_param:
            dept_filter_id = _to_oid(dept_id_param)
        if dept_filter_id:
            users_col = get_collection("auth", "users")
            all_user_id_variants = []
            for u in users_col.find({"department_id": dept_filter_id}, {"_id": 1}):
                all_user_id_variants.extend(_id_variants(u["_id"]))
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": dept_filter_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter
    else:
        user_dept_id = _user_dept_id(user)
        if user_dept_id:
            users_col = get_collection("auth", "users")
            all_user_id_variants = []
            for u in users_col.find({"department_id": user_dept_id}, {"_id": 1}):
                all_user_id_variants.extend(_id_variants(u["_id"]))
            dept_user_filter = {"$or": [
                {"performed_by": {"$in": all_user_id_variants}},
                {"target_user": {"$in": all_user_id_variants}},
                {"department_id": user_dept_id},
            ]}
            if "$and" in filters:
                filters["$and"].append(dept_user_filter)
            elif filters:
                filters = {"$and": [filters, dept_user_filter]}
            else:
                filters = dept_user_filter

    # Fetch up to 10,000 logs for export to protect memory
    logs, _ = get_audit_logs(1, 10000, filters, department_id=None)
    
    audit_user_ids = [
        item.get("performed_by") or item.get("actor_user_id")
        for item in logs
        if item.get("performed_by") or item.get("actor_user_id")
    ] + [
        item.get("target_user") or ((item.get("details") or {}).get("user_id") if isinstance(item.get("details"), dict) else None)
        for item in logs
        if item.get("target_user") or (isinstance(item.get("details"), dict) and (item.get("details") or {}).get("user_id"))
    ]
    user_map = get_users_by_ids(audit_user_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    headers = ["Timestamp", "Actor Name", "Actor Email", "Role", "Action", "Target", "IP Address", "Details"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)

    def _safe_str(v):
        if not v:
            return ""
        try:
            return str(v)
        except Exception:
            return ""

    for raw in logs:
        raw_ts = raw.get("timestamp")
        if raw_ts and raw_ts.tzinfo is None:
            raw_ts = raw_ts.replace(tzinfo=timezone.utc)
            
        ts_str = raw_ts.isoformat() if raw_ts else ""
        
        actor_id = raw.get("performed_by") or raw.get("actor_user_id")
        details_user_id = (raw.get("details") or {}).get("user_id") if isinstance(raw.get("details"), dict) else None
        target_user_id = raw.get("target_user") or details_user_id

        actor = user_map.get(_as_text(actor_id)) if actor_id else None
        target_user = user_map.get(_as_text(target_user_id)) if target_user_id else None

        a_name = (actor or {}).get("name") or ("System" if str(actor_id).lower() == "system" else "Unknown User")
        a_mail = (actor or {}).get("email") or ""
        a_role = (actor or {}).get("role") or raw.get("role") or "unknown"
        action_name = raw.get("action", "")
        
        if target_user:
            t_type = f"{target_user.get('name', 'Unknown')} ({target_user.get('role', 'user')})"
        elif target_user_id:
            t_type = f"User {target_user_id}"
        else:
            t_type = _as_text(raw.get("details")) or "System"

        ip_addr = raw.get("ip") or raw.get("ip_address") or ""
        details = _safe_str(raw.get("details"))

        ws.append([ts_str, a_name, a_mail, a_role, action_name, t_type, ip_addr, details])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Audit_Trail_Export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route("/audit-logs/<log_id>/rollback", methods=["POST"])
@role_required("department_admin")
def rollback_audit_action(user, log_id):
    audit_log = get_audit_log_by_id(log_id)
    if not audit_log:
        return jsonify({"error": "Audit log not found"}), 404

    if audit_log.get("rolled_back"):
        return jsonify({"error": "This action has already been rolled back"}), 400

    rollback_payload = audit_log.get("rollback")
    if not rollback_payload:
        return jsonify({"error": "Rollback not available for this action"}), 400

    rollback_until = audit_log.get("rollback_until")
    raw_ts = audit_log.get("timestamp")

    # Normalize timestamp for derivation if needed
    if raw_ts and raw_ts.tzinfo is None:
        raw_ts = raw_ts.replace(tzinfo=timezone.utc)

    if not rollback_until:
        rollback_until = (raw_ts or datetime.now(timezone.utc)) + timedelta(days=1)

    # Normalize rollback_until for comparison
    if rollback_until and rollback_until.tzinfo is None:
        rollback_until = rollback_until.replace(tzinfo=timezone.utc)

    if datetime.now(timezone.utc) > rollback_until:
        return jsonify({"error": "Rollback window expired (1 day)"}), 403

    try:
        _execute_rollback_operation(rollback_payload)
    except Exception as exc:
        current_app.logger.exception("Audit rollback failed")
        return jsonify({"error": f"Rollback failed: {exc}"}), 500

    mark_audit_log_rolled_back(log_id, str(user["_id"]))
    log_action(
        "ROLLBACK_ACTION",
        str(user["_id"]),
        details=f"Rolled back audit log {log_id}",
    )
    return jsonify({"message": "Rollback completed successfully"}), 200


# ─── Attendance Override ────────────────────────────────────────────────────

@admin_bp.route("/attendance/override", methods=["POST"])
@role_required("department_admin")
def override_attendance(user):
    """Manually add or remove an attendance record (Special Exam Access)."""
    d = request.get_json(silent=True) or {}
    action = d.get("action", "add")  # "add" or "remove"

    if action == "add":
        log = log_attendance(
            d["paper_id"], d["user_id"], str(user["_id"]),
            session_id="manual-override", method="manual",
        )
        log_action("ATTENDANCE_OVERRIDE_ADD", str(user["_id"]),
                   target_user=d["user_id"],
                   details=f"Paper {d['paper_id']}")
        _clear_query_cache()
        return jsonify({"message": "Attendance added", "log": log}), 201
    else:
        from app.models.attendance import delete_attendance_log
        delete_attendance_log(d["log_id"])
        log_action("ATTENDANCE_OVERRIDE_REMOVE", str(user["_id"]),
                   details=f"Log {d['log_id']}")
        _clear_query_cache()
        return jsonify({"message": "Attendance removed"}), 200


@admin_bp.route("/exam-eligibility-summary", methods=["GET"])
@role_required("super_admin", "department_admin")
def exam_eligibility_summary(user):
    """Admin view of exam eligibility with filters and override states."""
    cache_key = (
        "exam_eligibility_summary",
        tuple(sorted((k, _as_text(v)) for k, v in request.args.items())),
    )
    cached_payload = _cache_get(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    department_filter = _as_text(request.args.get("department", ""))
    department_id_filter = _as_text(request.args.get("department_id", ""))
    course_id = _as_text(request.args.get("course_id", ""))
    paper_id = _as_text(request.args.get("paper_id", ""))
    academic_session = _normalise_year(request.args.get("academic_session", "")) or _normalise_year(request.args.get("academic_year", ""))
    semester_filter = _as_text(request.args.get("semester", ""))
    q = _as_text(request.args.get("q", "")).lower()
    final_eligible_filter = _as_text(request.args.get("final_eligible", ""))
    include_inactive = _to_bool(request.args.get("include_inactive", False))

    profiles_col = get_collection("academic", "student_profiles")

    # Build dept scope: dept admins are always scoped; super admins may filter by department_id
    if is_super_admin(user):
        dept_scope_id = department_id_filter or None
    else:
        dept_scope_id = _user_dept_id(user)

    # Fetch courses by ID scope
    courses = sanitise_many(get_all_courses(["name", "code", "status", "department", "course_duration", "year"], department_id=dept_scope_id))
    
    # Defensive fix: If we have a specific department selected, ensure we also include courses that might only have the department name set but missing the ID
    if dept_scope_id:
        from app.models.department import get_department_by_id
        target_dept = get_department_by_id(dept_scope_id)
        if target_dept and target_dept.get("name"):
            target_name = target_dept["name"]
            # Get all courses matching this department name to catch any data inconsistencies
            all_courses_col = get_collection("academic", "courses")
            name_matches = sanitise_many(list(all_courses_col.find({"department": target_name})))
            
            # Merge results to avoid duplicates
            existing_ids = {str(c["_id"]) for c in courses}
            for nm in name_matches:
                if str(nm["_id"]) not in existing_ids:
                    courses.append(nm)

    papers = sanitise_many(get_all_papers(["name", "code", "semester", "course_id", "lecturer_id", "created_at"]))
    course_map = {c["_id"]: c for c in courses}
    paper_map = {p["_id"]: p for p in papers}

    # Restrict profiles to only courses visible to this admin
    visible_course_ids = list(course_map.keys())
    base_profile_query = {"course_id": {"$in": visible_course_ids}} if visible_course_ids else {"course_id": "never_match"}

    profiles = list(
        profiles_col.find(
            base_profile_query,
            {
                "user_id": 1,
                "course_id": 1,
                "academic_year": 1,
                "academic_session": 1,
                "year": 1,
                "current_semester": 1,
                "enrolled_papers": 1,
                "reg_number": 1,
                "roll_number": 1,
                "created_at": 1,
            },
        )
    )
    user_map = get_users_by_ids(profile.get("user_id") for profile in profiles)
    overrides_col = get_collection("attendance", "exam_eligibility_overrides")
    sessions_col = get_collection("attendance", "attendance_sessions")

    classes_happened_by_paper = {}
    for row in sessions_col.aggregate([
        {
            "$group": {
                "_id": {
                    "paper_id": "$paper_id",
                },
                "count": {"$sum": 1},
            }
        }
    ]):
        gid = row.get("_id") or {}
        gid_paper = _as_text(gid.get("paper_id"))
        count = int(row.get("count", 0) or 0)

        if gid_paper:
            classes_happened_by_paper[gid_paper] = classes_happened_by_paper.get(gid_paper, 0) + count

    selected_profiles = []
    relevant_user_ids = []
    relevant_paper_ids = set()

    for profile in profiles:
        uid = profile.get("user_id")
        if not uid:
            continue
        uid_text = _as_text(uid)
        student = user_map.get(uid_text)
        if not student:
            continue

        stu_course_id = _as_text(profile.get("course_id", ""))
        course_doc = course_map.get(stu_course_id)
        course_status = _as_text((course_doc or {}).get("status") or "active").lower() or "active"
        if course_status != "active" and not include_inactive:
            continue
        # Department filter – match on course's department name
        if department_filter:
            course_department = _as_text((course_doc or {}).get("department") or "")
            if course_department.lower() != department_filter.lower():
                continue
        stu_year = _as_text(profile.get("academic_session") or profile.get("academic_year") or profile.get("year"))
        enrolled = profile.get("enrolled_papers", []) or []
        stu_semester = _to_int(profile.get("current_semester"), 0) or None
        if stu_semester is None:
            derived_semesters = []
            for pid in enrolled:
                pdoc = paper_map.get(pid) or {}
                psem = _to_int(pdoc.get("semester"), 0)
                if psem > 0:
                    derived_semesters.append(psem)
            if derived_semesters:
                stu_semester = max(derived_semesters)

        if course_id and stu_course_id != course_id:
            continue
        if academic_session and stu_year != academic_session:
            continue

        selected_profiles.append((profile, student, uid_text, stu_course_id, stu_year, enrolled, stu_semester))
        relevant_user_ids.append(uid_text)
        for pid in enrolled:
            pid_text = _as_text(pid)
            if paper_id and pid_text != paper_id:
                continue
            paper = paper_map.get(pid_text) or paper_map.get(pid)
            if not paper:
                continue
            paper_semester = _to_int(paper.get("semester"), 0)
            if semester_filter and str(paper_semester) != semester_filter:
                continue
            relevant_paper_ids.add(pid_text)

    student_match_ids = []
    for sid in relevant_user_ids:
        student_match_ids.extend(_id_variants(sid))
    paper_match_ids = []
    for pid in relevant_paper_ids:
        paper_match_ids.extend(_id_variants(pid))

    attendance_count_map = {}
    if student_match_ids and paper_match_ids:
        attendance_logs = get_collection("attendance", "attendance_logs")
        for row in attendance_logs.aggregate([
            {
                "$match": {
                    "user_id": {"$in": student_match_ids},
                    "paper_id": {"$in": paper_match_ids},
                }
            },
            {
                "$group": {
                    "_id": {
                        "user_id": "$user_id",
                        "paper_id": "$paper_id",
                    },
                    "count": {"$sum": 1},
                }
            },
        ]):
            gid = row.get("_id") or {}
            attendance_count_map[(
                _as_text(gid.get("user_id")),
                _as_text(gid.get("paper_id")),
            )] = int(row.get("count", 0) or 0)

    # Pre-fetch all relevant sessions to eliminate N+1 queries in the loop
    sessions_by_paper = {}
    if paper_match_ids:
        all_sessions = list(sessions_col.find(
            {"paper_id": {"$in": paper_match_ids}},
            {"paper_id": 1, "committed_at": 1, "last_updated_at": 1, "created_at": 1}
        ))
        for s in all_sessions:
            pid_str = _as_text(s.get("paper_id"))
            ts = s.get("committed_at") or s.get("last_updated_at") or s.get("created_at")
            if ts:
                if ts.tzinfo is None:
                    ts = ts.replace(tzinfo=timezone.utc)
                if pid_str not in sessions_by_paper:
                    sessions_by_paper[pid_str] = []
                sessions_by_paper[pid_str].append(ts)

    override_map = {}
    if student_match_ids and paper_match_ids:
        for override in overrides_col.find(
            {
                "user_id": {"$in": student_match_ids},
                "paper_id": {"$in": paper_match_ids},
            },
            {
                "_id": 0,
                "user_id": 1,
                "paper_id": 1,
                "override_status": 1,
                "reason": 1,
            },
        ):
            key = (_as_text(override.get("user_id")), _as_text(override.get("paper_id")))
            override_map[key] = {
                "override_status": override.get("override_status"),
                "reason": _as_text(override.get("reason", "")),
            }

    items = []
    for profile, student, uid, stu_course_id, stu_year, enrolled, stu_semester in selected_profiles:
        course = course_map.get(stu_course_id)
        per_paper_rows = []
        total_attended_overall = 0
        total_classes_overall = 0

        for pid in enrolled:
            pid_text = _as_text(pid)
            if paper_id and pid_text != paper_id:
                continue

            paper = paper_map.get(pid_text) or paper_map.get(pid)
            if not paper:
                continue

            paper_semester = _to_int(paper.get("semester"), 0)
            if semester_filter and str(paper_semester) != semester_filter:
                continue

            lecturer_id_for_paper = _as_text(paper.get("lecturer_id", ""))
            profile_created_at = profile.get("created_at")

            # Count classes for this subject across all lecturers,
            # scoped to sessions after the student was enrolled.
            if profile_created_at:
                if profile_created_at.tzinfo is None:
                    profile_created_at = profile_created_at.replace(tzinfo=timezone.utc)
                paper_sessions = sessions_by_paper.get(pid_text, [])
                classes_happened = sum(1 for ts in paper_sessions if ts >= profile_created_at)
            else:
                classes_happened = int(classes_happened_by_paper.get(pid_text, 0) or 0)

            attended = attendance_count_map.get((_as_text(uid), pid_text), 0)
            pct = round((attended / classes_happened) * 100, 2) if classes_happened > 0 else 0.0

            total_attended_overall += attended
            total_classes_overall += classes_happened

            override = override_map.get((_as_text(uid), pid_text))
            override_status = None if not override else override.get("override_status")
            override_reason = "" if not override else override.get("reason", "")

            if q and not (
                q in _as_text(student.get("name", "")).lower()
                or q in _as_text(student.get("email", "")).lower()
                or q in _as_text(profile.get("reg_number") or profile.get("roll_number")).lower()
                or q in _as_text(paper.get("name", "")).lower()
                or q in _as_text(paper.get("code", "")).lower()
            ):
                continue

            per_paper_rows.append({
                "user_id": uid,
                "student_name": student.get("name", "Unknown"),
                "student_email": student.get("email", ""),
                "reg_number": profile.get("reg_number") or profile.get("roll_number"),
                "course_id": stu_course_id,
                "course_name": (course or {}).get("name"),
                "course_department": (course or {}).get("department") or "",
                "student_semester": stu_semester,
                "paper_id": pid_text,
                "paper_name": paper.get("name", ""),
                "paper_code": paper.get("code", ""),
                "semester": paper_semester or None,
                "lecturer_id": lecturer_id_for_paper,
                "academic_year": stu_year,
                "academic_session": stu_year,
                "enrolled_since": profile_created_at,
                "attended": attended,
                "total_classes": classes_happened,
                "attended_classes": attended,
                "classes_happened": classes_happened,
                "attendance_percentage": pct,
                "override_status": override_status,
                "override_reason": override_reason,
            })

        overall_pct = round((total_attended_overall / total_classes_overall) * 100, 2) if total_classes_overall > 0 else 0.0
        has_lectures = total_classes_overall > 0
        overall_eligible = (overall_pct >= 75.0) if has_lectures else None

        for row in per_paper_rows:
            override_status = row.get("override_status")
            final_eligible = overall_eligible if override_status is None else bool(override_status)
            if final_eligible is None:
                eligibility_status = "no_lectures_yet"
            else:
                eligibility_status = "eligible" if final_eligible else "ineligible"

            if final_eligible_filter:
                required = _to_bool(final_eligible_filter)
                if final_eligible is None or final_eligible != required:
                    continue

            row["overall_attendance_percentage"] = overall_pct
            row["overall_attended_classes"] = total_attended_overall
            row["overall_total_classes"] = total_classes_overall
            row["eligible_by_attendance"] = overall_eligible
            row["final_eligible"] = final_eligible
            row["eligibility_status"] = eligibility_status
            items.append(row)

    payload = {
        "total": len(items),
        "eligible_count": sum(1 for x in items if x["final_eligible"] is True),
        "ineligible_count": sum(1 for x in items if x["final_eligible"] is False),
        "items": items,
    }
    _cache_set(cache_key, payload, _ELIGIBILITY_CACHE_TTL_SECONDS)
    return jsonify(payload)


@admin_bp.route("/exam-eligibility-override", methods=["PUT"])
@role_required("department_admin")
def set_exam_eligibility_override(user):
    """Override final exam eligibility status for a student-paper pair."""
    d = request.get_json(silent=True) or {}
    user_id = _as_text(d.get("user_id", ""))
    paper_id = _as_text(d.get("paper_id", ""))
    reason = _as_text(d.get("reason", ""))

    if not user_id or not paper_id:
        return jsonify({"error": "user_id and paper_id are required"}), 400

    if d.get("override_status", None) is None:
        return jsonify({"error": "override_status must be true or false"}), 400

    raw_status = d.get("override_status")
    if isinstance(raw_status, str):
        raw_lower = raw_status.strip().lower()
        if raw_lower not in {"1", "0", "true", "false", "yes", "no", "y", "n"}:
            return jsonify({"error": "override_status must be true or false"}), 400
    override_status = _to_bool(raw_status)

    overrides_col = get_collection("attendance", "exam_eligibility_overrides")
    overrides_col.update_one(
        {
            "user_id": {"$in": _id_variants(user_id)},
            "paper_id": {"$in": _id_variants(paper_id)},
        },
        {
            "$set": {
                "user_id": user_id,
                "paper_id": paper_id,
                "override_status": override_status,
                "reason": reason,
                "updated_by": str(user["_id"]),
                "updated_at": datetime.now(timezone.utc),
            }
        },
        upsert=True,
    )

    log_action(
        "EXAM_ELIGIBILITY_OVERRIDE",
        str(user["_id"]),
        target_user=user_id,
        details=f"Paper {paper_id}, override={override_status}, reason={reason}",
    )
    _clear_query_cache()
    return jsonify({"message": "Eligibility override updated"}), 200


@admin_bp.route("/exam-eligibility-override/bulk", methods=["PUT"])
@role_required("department_admin")
def set_exam_eligibility_override_bulk(user):
    """Bulk override final exam eligibility for multiple student-paper pairs."""
    d = request.get_json(silent=True) or {}
    overrides = d.get("overrides")

    if not isinstance(overrides, list) or len(overrides) == 0:
        return jsonify({"error": "overrides must be a non-empty list"}), 400

    sanitized = []
    for item in overrides:
        if not isinstance(item, dict):
            continue

        user_id = _as_text(item.get("user_id", ""))
        paper_id = _as_text(item.get("paper_id", ""))
        if not user_id or not paper_id:
            continue
        if item.get("override_status", None) is None:
            continue
        if isinstance(item.get("override_status"), str):
            raw_lower = item.get("override_status", "").strip().lower()
            if raw_lower not in {"1", "0", "true", "false", "yes", "no", "y", "n"}:
                continue

        sanitized.append({
            "user_id": user_id,
            "paper_id": paper_id,
            "override_status": _to_bool(item.get("override_status")),
            "reason": _as_text(item.get("reason", "")),
        })

    if not sanitized:
        return jsonify({"error": "No valid override items found"}), 400

    overrides_col = get_collection("attendance", "exam_eligibility_overrides")
    now = datetime.now(timezone.utc)
    admin_id = str(user["_id"])
    unique_pairs = set()

    for item in sanitized:
        pair = (item["user_id"], item["paper_id"])
        if pair in unique_pairs:
            continue
        unique_pairs.add(pair)
        overrides_col.update_one(
            {
                "user_id": {"$in": _id_variants(item["user_id"])},
                "paper_id": {"$in": _id_variants(item["paper_id"])},
            },
            {
                "$set": {
                    "user_id": item["user_id"],
                    "paper_id": item["paper_id"],
                    "override_status": item["override_status"],
                    "reason": item["reason"],
                    "updated_by": admin_id,
                    "updated_at": now,
                }
            },
            upsert=True,
        )

    log_action(
        "EXAM_ELIGIBILITY_OVERRIDE_BULK",
        admin_id,
        details=f"Bulk overrides applied: {len(unique_pairs)}",
    )
    _clear_query_cache()
    return jsonify({"message": "Bulk eligibility overrides updated", "updated": len(unique_pairs)}), 200


# ─── Leave Requests (Feature 3) ──────────────────────────────────────────────

@admin_bp.route("/leave-requests", methods=["GET"])
@role_required("department_admin")
def list_leave_requests(user):
    """List all leave requests, optionally filtered by status or student."""
    leaves_col = get_collection("academic", "leave_requests")
    status_filter = request.args.get("status", "").strip()
    user_id    = request.args.get("user_id", "").strip()

    query = {}
    if status_filter:
        query["status"] = status_filter
    if user_id:
        query["user_id"] = user_id

    docs = list(leaves_col.find(query).sort("created_at", -1).limit(200))

    # Enrich with student and paper info
    user_ids = list({d.get("user_id") for d in docs if d.get("user_id")})
    user_map  = get_users_by_ids(user_ids)
    
    paper_ids = list({d.get("paper_id") for d in docs if d.get("paper_id")})
    papers_col = get_collection("academic", "papers")
    paper_ids_with_oids = []
    for pid in paper_ids:
        paper_ids_with_oids.extend(_id_variants(pid))
    paper_map = {str(p["_id"]): p for p in papers_col.find({"_id": {"$in": paper_ids_with_oids}})}

    for d in docs:
        ud = user_map.get(d.get("user_id")) or {}
        d["student_name"]  = _as_text(ud.get("name") or "Unknown")
        d["student_email"] = _as_text(ud.get("email") or "")
        
        pid = d.get("paper_id")
        if pid:
            pd = paper_map.get(str(pid)) or {}
            d["paper_name"] = _as_text(pd.get("name") or "Unknown")
            d["paper_code"] = _as_text(pd.get("code") or "")

    return jsonify(sanitise_many(docs))


@admin_bp.route("/leave-requests/<leave_id>/approve", methods=["PUT"])
@role_required("department_admin")
def approve_leave_request(user, leave_id):
    """Approve a leave request (marks as approved; attendance team can exclude those dates)."""
    leaves_col = get_collection("academic", "leave_requests")
    doc = leaves_col.find_one({"_id": _to_oid(leave_id)})
    if not doc:
        return jsonify({"error": "Leave request not found"}), 404

    now = datetime.now(timezone.utc)
    leaves_col.update_one(
        {"_id": _to_oid(leave_id)},
        {"$set": {
            "status":      "approved",
            "reviewed_by": str(user["_id"]),
            "reviewed_at": now,
        }},
    )
    log_action(
        "LEAVE_REQUEST_APPROVED",
        str(user["_id"]),
        target_user=doc.get("user_id"),
        details={"leave_id": leave_id, "date": doc.get("date"), "paper_id": doc.get("paper_id")},
    )
    return jsonify({"message": "Leave request approved"}), 200


@admin_bp.route("/leave-requests/<leave_id>/reject", methods=["PUT"])
@role_required("department_admin")
def reject_leave_request(user, leave_id):
    """Reject a leave request with an optional reason."""
    leaves_col = get_collection("academic", "leave_requests")
    doc = leaves_col.find_one({"_id": _to_oid(leave_id)})
    if not doc:
        return jsonify({"error": "Leave request not found"}), 404

    d      = request.get_json(silent=True) or {}
    remark = _as_text(d.get("remark", ""))
    now    = datetime.now(timezone.utc)

    leaves_col.update_one(
        {"_id": _to_oid(leave_id)},
        {"$set": {
            "status":      "rejected",
            "remark":      remark,
            "reviewed_by": str(user["_id"]),
            "reviewed_at": now,
        }},
    )
    log_action(
        "LEAVE_REQUEST_REJECTED",
        str(user["_id"]),
        target_user=doc.get("user_id"),
        details={"leave_id": leave_id, "remark": remark},
    )
    return jsonify({"message": "Leave request rejected"}), 200


def _parse_iso_date(value):
    text = _as_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except Exception:
        return None


def _local_midnight_to_utc(local_midnight, tz_offset_minutes):
    if not isinstance(local_midnight, datetime):
        return None
    return local_midnight + timedelta(minutes=_to_int(tz_offset_minutes, 0))


def _build_attendance_matrix_payload(args):
    department_filter = _as_text(args.get("department", ""))
    course_id = _as_text(args.get("course_id", ""))
    academic_session = _normalise_year(args.get("academic_session", "")) or _normalise_year(args.get("academic_year", ""))
    semester_filter = _as_text(args.get("semester", ""))
    tz_offset_minutes = _to_int(args.get("tz_offset_minutes", 0), 0)

    def _to_local(dt):
        if not isinstance(dt, datetime):
            return None
        # Browser sends JS getTimezoneOffset() minutes, so local = utc - offset.
        return dt - timedelta(minutes=tz_offset_minutes)

    from_date = _parse_iso_date(args.get("from_date", ""))
    to_date = _parse_iso_date(args.get("to_date", ""))

    range_start_utc = None
    range_end_utc = None
    if from_date:
        range_start_utc = _local_midnight_to_utc(from_date, tz_offset_minutes)
    if to_date:
        to_local_exclusive = to_date + timedelta(days=1)
        range_end_utc = _local_midnight_to_utc(to_local_exclusive, tz_offset_minutes)

    courses = sanitise_many(get_all_courses(["name", "code", "status", "course_duration", "department"]))
    papers = sanitise_many(get_all_papers(["name", "code", "semester", "course_id"]))
    paper_map = {p["_id"]: p for p in papers}

    dept_course_ids = set()
    if department_filter:
        for c in courses:
            if _as_text(c.get("department", "")).lower() == department_filter.lower():
                dept_course_ids.add(_as_text(c.get("_id")))

    allowed_papers = []
    for paper in papers:
        pid = _as_text(paper.get("_id"))
        if not pid:
            continue
        if department_filter and _as_text(paper.get("course_id")) not in dept_course_ids:
            continue
        if course_id and _as_text(paper.get("course_id")) != course_id:
            continue
        if semester_filter and _as_text(paper.get("semester")) != semester_filter:
            continue
        allowed_papers.append(pid)

    allowed_paper_set = set(allowed_papers)

    profiles_col = get_collection("academic", "student_profiles")
    profiles = list(
        profiles_col.find(
            {},
            {
                "_id": 0,
                "user_id": 1,
                "course_id": 1,
                "academic_session": 1,
                "academic_year": 1,
                "year": 1,
                "current_semester": 1,
                "roll_number": 1,
                "reg_number": 1,
                "enrolled_papers": 1,
            },
        )
    )

    available_sessions = set()
    for profile in profiles:
        profile_course_id = _as_text(profile.get("course_id"))
        if department_filter and profile_course_id not in dept_course_ids:
            continue
        if course_id and profile_course_id != course_id:
            continue
        profile_session = _as_text(profile.get("academic_session") or profile.get("academic_year") or profile.get("year"))
        if profile_session:
            available_sessions.add(profile_session)

    candidate_students = []
    for profile in profiles:
        user_id = _as_text(profile.get("user_id"))
        if not user_id:
            continue

        stu_course_id = _as_text(profile.get("course_id"))
        if department_filter and stu_course_id not in dept_course_ids:
            continue
        if course_id and stu_course_id != course_id:
            continue

        stu_session = _as_text(profile.get("academic_session") or profile.get("academic_year") or profile.get("year"))
        if academic_session and stu_session != academic_session:
            continue

        enrolled = [_as_text(pid) for pid in (profile.get("enrolled_papers") or []) if _as_text(pid)]
        if allowed_paper_set:
            enrolled = [pid for pid in enrolled if pid in allowed_paper_set]

        if semester_filter:
            current_sem = _as_text(profile.get("current_semester"))
            has_semester_paper = any(_as_text((paper_map.get(pid) or {}).get("semester")) == semester_filter for pid in enrolled)
            if current_sem != semester_filter and not has_semester_paper:
                continue

        candidate_students.append(
            {
                "user_id": user_id,
                "roll_no": _as_text(profile.get("roll_number") or profile.get("reg_number")),
                "enrolled_papers": enrolled,
            }
        )

    user_map = get_users_by_ids([s["user_id"] for s in candidate_students])

    students = []
    for stu in candidate_students:
        user_doc = user_map.get(stu["user_id"]) or {}
        students.append(
            {
                "user_id": stu["user_id"],
                "roll_no": stu["roll_no"] or "N/A",
                "name": _as_text(user_doc.get("name", "Unknown")) or "Unknown",
                "enrolled_papers": stu["enrolled_papers"],
            }
        )

    user_ids = set(s["user_id"] for s in students)

    if allowed_paper_set:
        paper_filter_set = set(allowed_paper_set)
    else:
        paper_filter_set = set()
        for stu in students:
            for pid in stu.get("enrolled_papers", []):
                paper_filter_set.add(pid)

    session_query = {}
    if not paper_filter_set:
        session_docs = []
    else:
        paper_match_ids = []
        for pid in paper_filter_set:
            paper_match_ids.extend(_id_variants(pid))
        session_query["paper_id"] = {"$in": paper_match_ids}

        committed_range = {}
        if range_start_utc:
            committed_range["$gte"] = range_start_utc
        if range_end_utc:
            committed_range["$lt"] = range_end_utc
        if committed_range:
            session_query["committed_at"] = committed_range

        sessions_col = get_collection("attendance", "attendance_sessions")
        session_docs = list(
            sessions_col.find(
                session_query,
                {
                    "_id": 0,
                    "session_id": 1,
                    "paper_id": 1,
                    "user_ids": 1,
                    "committed_at": 1,
                    "last_updated_at": 1,
                    "period_number": 1,
                    "period": 1,
                },
            )
        )

    date_subject_sessions = {}
    for doc in session_docs:
        paper_id = _as_text(doc.get("paper_id"))
        if not paper_id:
            continue
        if paper_filter_set and paper_id not in paper_filter_set:
            continue

        dt = doc.get("committed_at") or doc.get("last_updated_at")
        if not isinstance(dt, datetime):
            continue
        local_dt = _to_local(dt)
        date_key = local_dt.strftime("%Y-%m-%d")

        paper = paper_map.get(paper_id) or {}
        subject_code = _as_text(paper.get("code") or paper.get("name") or "SUB")
        subject_name = _as_text(paper.get("name") or paper.get("code") or "Subject")
        period_number = _as_text(doc.get("period_number") or doc.get("period"))
        session_id = _as_text(doc.get("session_id"))
        timestamp_key = local_dt.strftime("%Y%m%d%H%M%S%f")

        # Keep every committed class distinct, even for same subject/date/period.
        subject_key = f"{paper_id}::{period_number or 'NA'}::{session_id or timestamp_key}"
        compound_key = f"{date_key}::{subject_key}"

        if compound_key not in date_subject_sessions:
            date_subject_sessions[compound_key] = {
                "date": date_key,
                "paper_id": paper_id,
                "subject_code": subject_code,
                "subject_name": subject_name,
                "period_number": period_number,
                "column_key": compound_key,
                "present_set": set(),
            }

        for sid in (doc.get("user_ids") or []):
            sid_text = _as_text(sid)
            if sid_text:
                date_subject_sessions[compound_key]["present_set"].add(sid_text)

    ordered_columns = sorted(
        date_subject_sessions.values(),
        key=lambda x: (
            x.get("date") or "",
            _to_int(x.get("period_number"), 0),
            x.get("subject_code") or "",
            x.get("paper_id") or "",
        ),
    )

    # Global sequence across all class slots in selected range.
    for idx, col in enumerate(ordered_columns, start=1):
        col["global_sequence"] = idx

    grouped_dates = {}
    for col in ordered_columns:
        d = col["date"]
        date_bucket = grouped_dates.setdefault(d, [])
        slot_index = len(date_bucket)
        if slot_index < 26:
            subject_slot = chr(ord("A") + slot_index)
        else:
            subject_slot = f"S{slot_index + 1}"

        date_bucket.append(
            {
                "column_key": col["column_key"],
                "paper_id": col["paper_id"],
                "subject_code": col["subject_code"],
                "subject_name": col["subject_name"],
                "subject_slot": subject_slot,
                "period_number": col["period_number"],
                "sequence_number": col.get("global_sequence"),
                "label": f"{col['subject_code']} ({col['period_number']})" if col["period_number"] else col["subject_code"],
            }
        )

    dates = [{"date": d, "subjects": grouped_dates[d]} for d in sorted(grouped_dates.keys())]

    students.sort(key=lambda x: (x.get("roll_no") or "", x.get("name") or ""))
    rows = []
    for stu in students:
        cell_map = {}
        attended_counter = 0
        for col in ordered_columns:
            present = stu["user_id"] in col["present_set"]
            if present:
                attended_counter += 1
                cell_map[col["column_key"]] = _as_text(attended_counter) or "1"
            else:
                cell_map[col["column_key"]] = "X"

        date_summary = {}
        for date_entry in dates:
            parts = []
            for sub in date_entry["subjects"]:
                parts.append(cell_map.get(sub["column_key"], "X"))
            date_summary[date_entry["date"]] = " : ".join(parts) if parts else ""

        rows.append(
            {
                "user_id": stu["user_id"],
                "roll_no": stu["roll_no"],
                "name": stu["name"],
                "cells": cell_map,
                "date_summary": date_summary,
            }
        )

    available_semesters = set()
    for paper in papers:
        if course_id and _as_text(paper.get("course_id")) != course_id:
            continue
        sem = _as_text(paper.get("semester"))
        if sem:
            available_semesters.add(sem)

    if not available_semesters and course_id:
        course_doc = next((c for c in courses if _as_text(c.get("_id")) == course_id), None)
        duration_years = _to_int((course_doc or {}).get("course_duration"), 0)
        if duration_years > 0:
            for sem in range(1, duration_years * 2 + 1):
                available_semesters.add(str(sem))

    options_courses = []
    for course in courses:
        options_courses.append(
            {
                "_id": _as_text(course.get("_id")),
                "name": _as_text(course.get("name")),
                "code": _as_text(course.get("code")),
                "status": _as_text(course.get("status") or "active").lower() or "active",
                "course_duration": _to_int(course.get("course_duration"), 0),
            }
        )

    options_courses.sort(key=lambda x: (x.get("name") or "", x.get("code") or ""))

    return {
        "filters": {
            "course_id": course_id or None,
            "academic_session": academic_session or None,
            "semester": semester_filter or None,
            "tz_offset_minutes": tz_offset_minutes,
            "from_date": _as_text(args.get("from_date", "")) or None,
            "to_date": _as_text(args.get("to_date", "")) or None,
        },
        "options": {
            "courses": options_courses,
            "academic_sessions": sorted(available_sessions),
            "semesters": sorted(available_semesters, key=lambda x: _to_int(x, 0)),
        },
        "meta": {
            "students_count": len(rows),
            "dates_count": len(dates),
            "sessions_count": len(ordered_columns),
            "subject_columns_count": len(ordered_columns),
        },
        "dates": dates,
        "rows": rows,
    }


@admin_bp.route("/attendance-matrix", methods=["GET"])
@role_required("department_admin")
def attendance_matrix(user):
    cache_key = (
        "attendance_matrix",
        tuple(sorted((k, _as_text(v)) for k, v in request.args.items())),
    )
    cached_payload = _cache_get(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    payload = _build_attendance_matrix_payload(request.args)
    _cache_set(cache_key, payload, _QUERY_CACHE_TTL_SECONDS)
    return jsonify(payload)


@admin_bp.route("/attendance-matrix/export", methods=["GET"])
@role_required("department_admin")
def attendance_matrix_export(user):
    payload = _build_attendance_matrix_payload(request.args)
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"error": "openpyxl is required for Excel export. Install it in backend requirements."}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Matrix"

    dates = payload.get("dates") or []
    rows = payload.get("rows") or []

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(row=1, column=1, value="Roll No")
    ws.cell(row=1, column=2, value="Name")

    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_idx = 3
    ordered_subjects = []
    for date_entry in dates:
        subjects = date_entry.get("subjects") or []
        if not subjects:
            continue

        start_col = col_idx
        for sub in subjects:
            ws.cell(
                row=2,
                column=col_idx,
                value=_as_text(sub.get("subject_code") or sub.get("subject_name") or "SUB"),
            )
            ordered_subjects.append(sub)
            col_idx += 1

        end_col = col_idx - 1
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col, value=_as_text(date_entry.get("date")) or "Date")

    total_attended_col = col_idx
    total_held_col = col_idx + 1
    percentage_col = col_idx + 2

    ws.merge_cells(start_row=1, start_column=total_attended_col, end_row=2, end_column=total_attended_col)
    ws.merge_cells(start_row=1, start_column=total_held_col, end_row=2, end_column=total_held_col)
    ws.merge_cells(start_row=1, start_column=percentage_col, end_row=2, end_column=percentage_col)

    ws.cell(row=1, column=total_attended_col, value="TCA")
    ws.cell(row=1, column=total_held_col, value="TCH")
    ws.cell(row=1, column=percentage_col, value="%")

    body_start_row = 3
    for i, row in enumerate(rows, start=body_start_row):
        ws.cell(row=i, column=1, value=row.get("roll_no"))
        ws.cell(row=i, column=2, value=row.get("name"))

        total_attended = 0
        total_held = 0
        for j, sub in enumerate(ordered_subjects, start=3):
            value = _as_text((row.get("cells") or {}).get(sub.get("column_key"), "X")) or "X"
            ws.cell(row=i, column=j, value=value)
            total_held += 1
            if value.upper() != "X":
                total_attended += 1

        percentage = round((total_attended / total_held) * 100, 2) if total_held > 0 else 0
        ws.cell(row=i, column=total_attended_col, value=total_attended)
        ws.cell(row=i, column=total_held_col, value=total_held)
        ws.cell(row=i, column=percentage_col, value=f"{percentage}%")

    max_col = max(2, percentage_col)
    max_row = max(2, body_start_row + len(rows) - 1)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center
            if r <= 2:
                cell.fill = header_fill
                cell.font = header_font

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for c in range(3, total_attended_col):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions[get_column_letter(total_attended_col)].width = 20
    ws.column_dimensions[get_column_letter(total_held_col)].width = 18
    ws.column_dimensions[get_column_letter(percentage_col)].width = 14

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"attendance_matrix_{india_timestamp_token()}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/attendance-matrix/export-csv", methods=["GET"])
@role_required("department_admin")
def attendance_matrix_export_csv(user):
    payload = _build_attendance_matrix_payload(request.args)
    tz_offset_minutes = _to_int(request.args.get("tz_offset_minutes", 0), 0)

    dates = payload.get("dates") or []
    rows = payload.get("rows") or []

    output = StringIO()
    writer = csv.writer(output)

    header_row_1 = ["Roll No", "Name"]
    header_row_2 = ["", ""]
    ordered_subjects = []

    for date_entry in dates:
        subjects = date_entry.get("subjects") or []
        if not subjects:
            continue
        for _ in subjects:
            header_row_1.append(date_entry.get("date"))
        for sub in subjects:
            header_row_2.append(sub.get("subject_slot") or sub.get("subject_code") or sub.get("label") or "SUB")
            ordered_subjects.append(sub)

    writer.writerow(header_row_1)
    writer.writerow(header_row_2)

    for row in rows:
        line = [row.get("roll_no"), row.get("name")]
        for sub in ordered_subjects:
            line.append((row.get("cells") or {}).get(sub.get("column_key"), "X"))
        writer.writerow(line)

    csv_bytes = output.getvalue().encode("utf-8-sig")
    stream = BytesIO(csv_bytes)
    stream.seek(0)
    filename = f"attendance_matrix_{india_timestamp_token()}.csv"

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv",
    )


@admin_bp.route("/attendance-matrix/export-pdf", methods=["GET"])
@role_required("department_admin")
def attendance_matrix_export_pdf(user):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    payload = _build_attendance_matrix_payload(request.args)
    dates = payload.get("dates") or []
    rows = payload.get("rows") or []

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Official Attendance Matrix Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))

    header_row = ["Roll No", "Name"]
    # For a simple PDF, we'll flatten the columns
    ordered_subjects = []
    for date_entry in dates:
        subjects = date_entry.get("subjects") or []
        for sub in subjects:
            short_title = f"{date_entry.get('date')[5:]}\n{sub.get('subject_slot') or sub.get('subject_code')}"
            header_row.append(short_title)
            ordered_subjects.append(sub)

    header_row.extend(["TCA", "TCH", "%"])
    table_data = [header_row]

    for row in rows:
        line = [row.get("roll_no"), row.get("name")[:15] + ".." if len(row.get("name")) > 15 else row.get("name")]
        total_attended = 0
        total_held = 0
        for sub in ordered_subjects:
            val = (row.get("cells") or {}).get(sub.get("column_key"), "X")
            line.append(val)
            total_held += 1
            if val.upper() != "X":
                total_attended += 1
        
        percentage = round((total_attended / total_held) * 100, 2) if total_held > 0 else 0
        line.extend([str(total_attended), str(total_held), f"{percentage}%"])
        table_data.append(line)

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(t)
    
    elements.append(Spacer(1, 40))
    elements.append(Paragraph(f"Generated at: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Paragraph("Authorized Administrator Signature: _______________________", styles['Normal']))
    
    doc.build(elements)
    pdf_bytes = output.getvalue()
    stream = BytesIO(pdf_bytes)
    stream.seek(0)
    filename = f"attendance_matrix_{india_timestamp_token()}.pdf"

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )



# ─── Dashboard Stats ────────────────────────────────────────────────────────

@admin_bp.route("/stats", methods=["GET"])
@role_required("super_admin", "department_admin")
def dashboard_stats(user):
    dept_filter_id = _as_text(request.args.get("department_id", ""))
    
    user_dept = None
    if is_super_admin(user):
        user_dept = _to_oid(dept_filter_id) if dept_filter_id else None
    else:
        user_dept = _user_dept_id(user)
    
    cache_key = ("dashboard_stats", user_dept)
    cached_payload = _cache_get(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    started_at = current_app.config.get("APP_STARTED_AT")
    uptime_seconds = int((datetime.now(timezone.utc) - started_at).total_seconds()) if started_at else 0
    uptime_days, remainder = divmod(max(uptime_seconds, 0), 86400)
    uptime_hours, remainder = divmod(remainder, 3600)
    uptime_minutes, uptime_seconds = divmod(remainder, 60)

    uptime_parts = []
    if uptime_days:
        uptime_parts.append(f"{uptime_days}d")
    if uptime_hours or uptime_parts:
        uptime_parts.append(f"{uptime_hours}h")
    uptime_parts.append(f"{uptime_minutes}m")
    system_uptime = " ".join(uptime_parts)

    profiles_col = get_collection("academic", "student_profiles")
    users_col = get_collection("auth", "users")
    courses_col = get_collection("academic", "courses")
    papers_col = get_collection("academic", "papers")
    attendance_col = get_collection("attendance", "attendance_logs")
    audit_col = get_collection("audit", "audit_logs")
    by_course = {}
    by_year = {}
    by_department = {}
    courses = sanitise_many(get_all_courses(["name", "code", "year", "status", "department"], department_id=user_dept))
    course_map = {c["_id"]: c for c in courses}
    course_ids = list(course_map.keys())

    active_course_ids = {
        c.get("_id")
        for c in courses
        if _as_text(c.get("status") or "active").lower() == "active"
    }

    # course_ids from sanitise_many are strings; profiles may store course_id as ObjectId or string
    course_ids_with_oids = []
    for cid in course_ids:
        course_ids_with_oids.extend(_id_variants(cid))

    # Scope profiles to department's courses when filtering
    profile_query = {"course_id": {"$in": course_ids_with_oids}} if (user_dept and course_ids_with_oids) else {}
    profiles = list(
        profiles_col.find(
            profile_query,
            {
                "course_id": 1,
                "user_id": 1,
                "academic_session": 1,
                "academic_year": 1,
                "year": 1,
            },
        )
    )

    for profile in profiles:
        cid = _as_text(profile.get("course_id"))
        if cid and cid not in active_course_ids:
            continue

        course = course_map.get(cid)
        course_key = course.get("name") if course else "Unassigned"
        by_course[course_key] = by_course.get(course_key, 0) + 1

        dept_key = _as_text(course.get("department")) if course else ""
        if dept_key:
            if dept_key not in by_department:
                by_department[dept_key] = {"students": 0, "lecturers": 0}
            by_department[dept_key]["students"] += 1

        year_key = _normalise_year(
            profile.get("academic_session")
            or profile.get("academic_year")
            or profile.get("year")
        ) or "Unknown"
        by_year[year_key] = by_year.get(year_key, 0) + 1

    student_query = {"role": "student"}
    lecturer_query = {"role": "lecturer"}
    
    if user_dept:
        student_user_ids = [_as_text(p.get("user_id")) for p in profiles if p.get("user_id")]
        valid_oids = []
        for u in student_user_ids:
            valid_oids.extend(_id_variants(u))
        student_query["$or"] = [
            {"_id": {"$in": valid_oids}}
        ]

        dept_name = ""
        dept_doc = get_department_by_id(str(user_dept)) if user_dept else None
        if dept_doc:
            dept_name = _as_text(dept_doc.get("name", "")).strip()

        lecturer_query["$or"] = [{"department_id": user_dept}]
        if dept_name:
            lecturer_query["$or"].append({
                "department": {
                    "$regex": f"^{re.escape(dept_name)}$",
                    "$options": "i",
                },
            })

    for usr in users_col.find(lecturer_query, {"department": 1}):
        dept_key = _as_text(usr.get("department"))
        if dept_key:
            if dept_key not in by_department:
                by_department[dept_key] = {"students": 0, "lecturers": 0}
            by_department[dept_key]["lecturers"] += 1

    # Count queries - department_id must be sibling to $or, not inside it
    if user_dept:
        active_courses_count = courses_col.count_documents({
            "department_id": user_dept,
            "$or": [{"status": "active"}, {"status": {"$exists": False}}, {"status": ""}, {"status": None}],
        })
        inactive_courses_count = courses_col.count_documents({"department_id": user_dept, "status": "inactive"})
    else:
        active_courses_count = courses_col.count_documents({
            "$or": [{"status": "active"}, {"status": {"$exists": False}}, {"status": ""}, {"status": None}],
        })
        inactive_courses_count = courses_col.count_documents({"status": "inactive"})
    total_courses_count = active_courses_count + inactive_courses_count

    active_paper_count = 0
    inactive_paper_count = 0
    
    paper_query = {}
    if user_dept:
        paper_query["course_id"] = {"$in": course_ids}

    all_paper_ids = []
    for paper in papers_col.find(paper_query, {"course_id": 1}):
        if paper.get("_id"):
            all_paper_ids.extend(_id_variants(paper["_id"]))
        paper_course_id = _as_text(paper.get("course_id"))
        if not paper_course_id or paper_course_id in active_course_ids:
            active_paper_count += 1
        else:
            inactive_paper_count += 1

    audit_count = 0
    if not user_dept:
        audit_count = audit_col.count_documents({})
    else:
        all_uid_variants = []
        for u in users_col.find({"department_id": user_dept}, {"_id": 1}):
            all_uid_variants.extend(_id_variants(u["_id"]))
        audit_count = audit_col.count_documents({
            "$or": [
                {"performed_by": {"$in": all_uid_variants}},
                {"target_user": {"$in": all_uid_variants}},
                {"department_id": user_dept},
            ]
        })
        
    attendance_count = attendance_col.count_documents({"paper_id": {"$in": all_paper_ids}} if user_dept and all_paper_ids else {}) if not (user_dept and not all_paper_ids) else 0

    app_started_at = None
    if started_at:
        iso_started_at = started_at.isoformat()
        tz_part = iso_started_at[10:]
        has_tz = iso_started_at.endswith("Z") or "+" in tz_part or "-" in tz_part
        app_started_at = iso_started_at if has_tz else f"{iso_started_at}Z"

    payload = {
        "total_students": users_col.count_documents(student_query),
        "total_lecturers": users_col.count_documents(lecturer_query),
        "total_courses": total_courses_count,
        "active_courses": active_courses_count,
        "inactive_courses": inactive_courses_count,
        "total_papers": active_paper_count,
        "inactive_papers": inactive_paper_count,
        "total_attendance": attendance_count,
        "total_audit_logs": audit_count,
        "app_started_at": app_started_at,
        "system_uptime_seconds": max(int((datetime.now(timezone.utc) - started_at).total_seconds()), 0) if started_at else 0,
        "system_uptime": system_uptime,
        "students_by_course": by_course,
        "students_by_year": by_year,
        "departments_summary": by_department,
    }
    _cache_set(cache_key, payload, _QUERY_CACHE_TTL_SECONDS)
    return jsonify(payload)


@admin_bp.route("/stats/monthly-attendance", methods=["GET"])
@role_required("super_admin", "department_admin")
def monthly_attendance_trend_api(user):
    # Resolve which department_id to scope to
    if is_super_admin(user):
        # Super admin: optionally filter by department name param
        dept_name_filter = _as_text(request.args.get("department", ""))
        if dept_name_filter:
            dept_doc = get_collection("academic", "departments").find_one(
                {"name": {"$regex": f"^{dept_name_filter}$", "$options": "i"}}
            )
            scope_dept_id = dept_doc.get("_id") if dept_doc else None
        else:
            scope_dept_id = None  # No filter = global
    else:
        # Department admin: always locked to their own department_id
        scope_dept_id = _user_dept_id(user)

    attendance_col = get_collection("attendance", "attendance_logs")
    query = {}

    if scope_dept_id:
        # Filter by department_id (ObjectId) directly — reliable, no name mismatch
        dept_oid = _to_oid(scope_dept_id)
        dept_id_variants = [dept_oid, str(dept_oid)]

        courses = list(get_collection("academic", "courses").find(
            {"department_id": {"$in": dept_id_variants}}, {"_id": 1}
        ))
        course_ids = []
        for c in courses:
            course_ids.append(c["_id"])
            course_ids.append(str(c["_id"]))

        if not course_ids:
            def _ms(dt): return datetime(dt.year, dt.month, 1)
            def _sm(dt, d):
                y = dt.year + ((dt.month - 1 + d) // 12)
                m = ((dt.month - 1 + d) % 12) + 1
                return datetime(y, m, 1)
            now = datetime.now(timezone.utc)
            sm = _sm(_ms(now), -5)
            return jsonify([
                {"key": f"{_sm(sm,i).year}-{_sm(sm,i).month:02d}",
                 "label": _sm(sm, i).strftime("%b"), "total": 0, "sessions": 0, "students": 0}
                for i in range(6)
            ])

        papers = list(get_collection("academic", "papers").find(
            {"course_id": {"$in": course_ids}}, {"_id": 1}
        ))
        paper_ids = []
        for p in papers:
            paper_ids.append(p["_id"])
            paper_ids.append(str(p["_id"]))

        if not paper_ids:
            def _ms(dt): return datetime(dt.year, dt.month, 1)
            def _sm(dt, d):
                y = dt.year + ((dt.month - 1 + d) // 12)
                m = ((dt.month - 1 + d) % 12) + 1
                return datetime(y, m, 1)
            now = datetime.now(timezone.utc)
            sm = _sm(_ms(now), -5)
            return jsonify([
                {"key": f"{_sm(sm,i).year}-{_sm(sm,i).month:02d}",
                 "label": _sm(sm, i).strftime("%b"), "total": 0, "sessions": 0, "students": 0}
                for i in range(6)
            ])

        query["paper_id"] = {"$in": paper_ids}

    def _month_start(dt):
        return datetime(dt.year, dt.month, 1)

    def _shift_month(dt, delta):
        year = dt.year + ((dt.month - 1 + delta) // 12)
        month = ((dt.month - 1 + delta) % 12) + 1
        return datetime(year, month, 1)

    now = datetime.now(timezone.utc)
    current_month = _month_start(now)
    start_month = _shift_month(current_month, -5)

    query["timestamp"] = {"$gte": start_month}

    # ── Attendance logs: total records + unique students per month ──
    att_pipeline = [
        {"$match": query},
        {"$group": {
            "_id":      {"year": {"$year": "$timestamp"}, "month": {"$month": "$timestamp"}},
            "total":    {"$sum": 1},
            "students": {"$addToSet": "$user_id"},
        }},
    ]
    att_map = {}
    for row in attendance_col.aggregate(att_pipeline):
        g = row["_id"]
        key = f"{g['year']}-{g['month']:02d}"
        att_map[key] = {
            "total":    row["total"],
            "students": len([u for u in row["students"] if u is not None]),
        }

    # ── Committed sessions: unique sessions held per month ──
    sessions_col = get_collection("attendance", "attendance_sessions")
    sess_query = {"committed_at": {"$gte": start_month}}
    if "paper_id" in query:          # scoped to dept papers
        sess_query["paper_id"] = query["paper_id"]
    sess_pipeline = [
        {"$match": sess_query},
        {"$group": {
            "_id": {"year": {"$year": "$committed_at"}, "month": {"$month": "$committed_at"}},
            "sessions": {"$sum": 1},
        }},
    ]
    sess_map = {}
    for row in sessions_col.aggregate(sess_pipeline):
        g = row["_id"]
        key = f"{g['year']}-{g['month']:02d}"
        sess_map[key] = row["sessions"]

    points = []
    for i in range(6):
        month_dt = _shift_month(start_month, i)
        key = f"{month_dt.year}-{month_dt.month:02d}"
        a = att_map.get(key, {})
        points.append({
            "key":      key,
            "label":    month_dt.strftime("%b"),
            "total":    a.get("total", 0),
            "sessions": sess_map.get(key, 0),
            "students": a.get("students", 0),
        })

    return jsonify(points)



@admin_bp.route("/attendance/send-shortage-alerts", methods=["POST"])
@role_required("department_admin")
def send_shortage_alerts(user):
    """Scan all students and send shortage alert emails to those < 75% attendance."""
    if not is_email_delivery_enabled():
        return jsonify({"error": "Email delivery is not configured on this server."}), 503

    d = request.get_json(silent=True) or {}
    course_id = d.get("course_id")
    paper_id_filter = d.get("paper_id")

    profiles_col = get_collection("academic", "student_profiles")
    sessions_col = get_collection("attendance", "attendance_sessions")
    
    query = {}
    if course_id:
        query["course_id"] = course_id

    profiles = list(profiles_col.find(query))
    if not profiles:
        return jsonify({"message": "No students found to check."}), 200

    alerts_sent = 0
    checked_count = 0

    threshold = float(current_app.config.get("ATTENDANCE_THRESHOLD", 75.0))
    threshold_dec = threshold / 100.0

    for profile in profiles:
        uid = str(profile.get("user_id"))
        enrolled_papers = profile.get("enrolled_papers", [])
        
        # Filter papers if requested
        if paper_id_filter:
            enrolled_papers = [p for p in enrolled_papers if str(p) == str(paper_id_filter)]
            
        if not enrolled_papers:
            continue
            
        checked_count += 1
        leave_map = get_approved_leave_dates(uid, enrolled_papers)
        user_doc = find_user_by_id(uid)
        if not user_doc or not user_doc.get("email"):
            continue

        for paper_id in enrolled_papers:
            paper_id_text = str(paper_id)
            paper = get_paper_by_id(paper_id_text)
            if not paper:
                continue

            # Fetch committed sessions for this paper
            paper_id_variants = [paper_id_text]
            try:
                paper_id_variants.extend(_id_variants(paper_id_text))
            except Exception:
                pass
                
            committed_sessions = list(
                sessions_col.find(
                    {"paper_id": {"$in": paper_id_variants}},
                    {"user_ids": 1, "committed_at": 1, "last_updated_at": 1},
                )
            )
            
            if not committed_sessions:
                continue

            paper_leave_dates = leave_map.get(paper_id_text, set())
            attended = 0
            effective_total = 0
            for sess in committed_sessions:
                sess_date = session_date_str(sess)
                if sess_date and sess_date in paper_leave_dates:
                    continue
                effective_total += 1
                if uid in [str(sid) for sid in (sess.get("user_ids") or [])]:
                    attended += 1
            
            if effective_total == 0:
                continue
                
            pct = round((attended / effective_total) * 100, 2)
            if pct < threshold:
                # Calculate classes needed: (A + n) / (T + n) >= threshold_dec  =>  n >= (threshold_dec*T - A) / (1 - threshold_dec)
                divider = (1.0 - threshold_dec)
                needed_float = ((threshold_dec * effective_total) - attended) / divider if divider > 0 else 0
                classes_needed = max(0, int(needed_float) if needed_float.is_integer() else int(needed_float) + 1)
                
                send_shortage_alert_email(
                    to_email=user_doc["email"],
                    name=user_doc["name"],
                    paper_name=paper.get("name", "Unknown Paper"),
                    percentage=pct,
                    classes_needed=classes_needed
                )
                alerts_sent += 1

    return jsonify({
        "message": f"Alert scan completed. {checked_count} students checked.",
        "alerts_queued": alerts_sent
    })


# ─── Department Management (Super Admin only) ─────────────────────────────


